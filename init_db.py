"""
init_db.py
----------
Initialize the database and seed initial users.
DEVELOPMENT USE ONLY
"""

import os
import re
import time
import shutil
from sqlalchemy import text, or_
from werkzeug.security import generate_password_hash

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

# IMPORTANT (Windows/SQLite):
# Do NOT import the Flask app at module import time.
# app.py runs a runtime schema sync on import which opens SQLite and LOCKS the file.
# We need to delete/replace the DB file first, then import app/db/models.

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
os.makedirs(INSTANCE_DIR, exist_ok=True)
DB_PATH = os.path.join(INSTANCE_DIR, "workflow.db")

def safe_remove_db(db_path: str) -> None:
    """Safely remove (or rename) SQLite DB on Windows when the file is locked."""
    # Remove WAL/SHM first (they can also be locked)
    for suffix in ("-wal", "-shm"):
        p = db_path + suffix
        try:
            if os.path.exists(p):
                os.remove(p)
        except PermissionError:
            pass

    if not os.path.exists(db_path):
        return

    # ✅ Create a backup copy before deleting (safety net)
    try:
        backup_dir = os.path.join(os.path.dirname(db_path), "backups")
        os.makedirs(backup_dir, exist_ok=True)
        ts = time.strftime("%Y%m%d_%H%M%S")
        backup_path = os.path.join(backup_dir, f"workflow_backup_{ts}.db")
        shutil.copy2(db_path, backup_path)
        print(f"✅ Backup created: {backup_path}")
    except Exception as e:
        print(f"⚠️ Backup failed (continuing): {e}")

    # Try a few times to delete
    for _ in range(6):
        try:
            os.remove(db_path)
            return
        except PermissionError:
            time.sleep(0.5)

    # Still locked: rename so init can continue
    ts = time.strftime("%Y%m%d_%H%M%S")
    backup_path = db_path.replace(".db", f".locked_{ts}.db")
    try:
        shutil.move(db_path, backup_path)
        print(f"⚠️ DB is locked. Renamed to: {backup_path}")
    except PermissionError:
        print("❌ قاعدة البيانات ما زالت مفتوحة من برنامج آخر (Flask/DB Browser/PyCharm).")
        print("🔧 أغلق أي برنامج فاتح workflow.db ثم شغّل: python init_db.py")
        raise SystemExit(1)


def init_database():
    # =========================
    # 1️⃣ حذف قاعدة البيانات القديمة (قبل استيراد app)
    # =========================
    if os.path.exists(DB_PATH):
        print("Removing existing database...")
        safe_remove_db(DB_PATH)

    # منع runtime schema sync عند استيراد app.py (حتى لا يقفل SQLite على Windows)
    os.environ["SKIP_RUNTIME_SCHEMA"] = "1"

    from app import app
    from extensions import db

    # ⚠️ مهم: استيراد كل الـ Models لتسجيل الجداول بعد تهيئة db
    from models import (
        User, ArchivedFile, FilePermission, AuditLog, Notification,
        WorkflowRequest, RequestAttachment,
        WorkflowTemplate, WorkflowTemplateStep, WorkflowInstance, WorkflowInstanceStep,
        Organization, Directorate, Unit, Department,
        Role, RequestType, WorkflowRoutingRule, Committee, CommitteeAssignee
    )

    with app.app_context():

        import models
        # =========================
        # 2️⃣ إنشاء الجداول
        # =========================
        print("Creating database tables...")
        db.create_all()

        # =========================
        # Helper: list existing tables (SQLite)
        # =========================
        tables = {
            row[0]
            for row in db.session.execute(
                text("SELECT name FROM sqlite_master WHERE type='table'")
            ).all()
        }

        def has_table(name: str) -> bool:
            return name in tables


        # =========================
        # 2️⃣٫٥ Seed: Employee File Lookups (HRLookupItem)
        # =========================
        try:
            from models import HRLookupItem

            if has_table("hr_lookup_item") and HRLookupItem.query.count() == 0:
                print("Seeding HRLookupItem (employee file lookups)...")
                seed = {
                    "GENDER": [("M", "ذكر", "Male"), ("F", "أنثى", "Female")],
                    "MARITAL_STATUS": [("SINGLE", "أعزب", "Single"), ("MARRIED", "متزوج", "Married"), ("DIVORCED", "مطلق", "Divorced"), ("WIDOWED", "أرمل", "Widowed")],
                    "IDENTITY_TYPE": [("ID", "هوية", "ID"), ("PASS", "جواز سفر", "Passport")],
                    "EMP_STATUS": [("ACTIVE", "على رأس عمله", "Active"), ("ON_LEAVE", "إجازة", "On leave"), ("SUSPENDED", "موقوف", "Suspended"), ("ENDED", "منتهية خدمته", "Ended")],
                    "SHIFT": [("MORNING", "صباحية", "Morning"), ("EVENING", "مسائية", "Evening"), ("NIGHT", "ليلية", "Night")],
                    "ATTACH_TYPE": [("GEN", "عام", "General"), ("ID_DOC", "هوية", "ID Document"), ("CERT", "شهادة", "Certificate"), ("CONTRACT", "عقد", "Contract")],
                }
                for cat, items in seed.items():
                    for i, (code, ar, en) in enumerate(items):
                        db.session.add(HRLookupItem(category=cat, code=code, name_ar=ar, name_en=en, sort_order=i*10, is_active=True))
                db.session.commit()
        except Exception as e:
            print("HRLookupItem seed skipped:", e)


        # =========================
        # 3️⃣ إنشاء INDEXES
        # =========================
        print("Creating indexes...")

        # ===== Notification =====
        if has_table("notification"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_notification_user_read
                ON notification (user_id, is_read);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_notification_created
                ON notification (created_at DESC);
            """))

        # ===== Audit Log =====
        if has_table("audit_log"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_audit_user_created
                ON audit_log (user_id, created_at DESC);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_audit_target_created
                ON audit_log (target_type, target_id, created_at DESC);
            """))

        # ===== Archive =====
        if has_table("archived_file"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_archived_file_owner
                ON archived_file (owner_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_archived_file_created
                ON archived_file (upload_date DESC);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_archived_file_deleted
                ON archived_file (is_deleted, upload_date DESC);
            """))

        # ===== File Permission =====
        if has_table("file_permission"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_file_permission_user
                ON file_permission (user_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_file_permission_file
                ON file_permission (file_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_file_permission_expiry
                ON file_permission (expires_at);
            """))

        # ===== WorkflowRequest (default table name if no __tablename__) =====
        # if your model has __tablename__ then adjust here.
        if has_table("workflow_request"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_request_requester
                ON workflow_request (requester_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_request_status
                ON workflow_request (status);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_request_created
                ON workflow_request (created_at DESC);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_request_role
                ON workflow_request (current_role);
            """))

        # ===== Attachments table (we set __tablename__ = workflow_request_attachments) =====
        if has_table("workflow_request_attachments"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_wra_request
                ON workflow_request_attachments (request_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_wra_archived_file
                ON workflow_request_attachments (archived_file_id);
            """))

        # ===== Optional: workflow instances/steps if you created them =====
        if has_table("workflow_instances"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_instances_request
                ON workflow_instances (request_id);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_instances_current
                ON workflow_instances (current_step_order, is_completed);
            """))

        if has_table("workflow_instance_steps"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_wis_instance_order
                ON workflow_instance_steps (instance_id, step_order);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_wis_status_due
                ON workflow_instance_steps (status, due_at);
            """))

        if has_table("workflow_templates"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_templates_active
                ON workflow_templates (is_active);
            """))
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_workflow_templates_created
                ON workflow_templates (created_at DESC);
            """))

        if has_table("workflow_template_steps"):
            db.session.execute(text("""
                CREATE INDEX IF NOT EXISTS ix_wts_template_order
                ON workflow_template_steps (template_id, step_order);
            """))

        
        # =========================
        # =========================
        # 4) Seed Organization / Directorates / Departments + Users from seed.xlsx
        # =========================
        # Notes:
        # - The script will create the DB at: <instance_path>/workflow.db
        # - Place seed.xlsx next to init_db.py (recommended) OR set env var SEED_XLSX to a full path.
        # - If openpyxl is missing, seeding will STOP by default (to avoid silent empty seed).
        print(f"Database path: {DB_PATH}")
        print(f"Instance path: {app.instance_path}")

        SEED_OPTIONAL = os.environ.get("SEED_OPTIONAL", "").strip() in ("1", "true", "TRUE", "yes", "YES")
        SEED_XLSX_CANDIDATES = [
            os.environ.get("SEED_XLSX"),
            os.path.join(os.path.dirname(__file__), "seed.xlsx"),
            os.path.join(os.getcwd(), "seed.xlsx"),
        ]
        SEED_XLSX_CANDIDATES = [p for p in SEED_XLSX_CANDIDATES if p]

        def _norm(v):
            if v is None:
                return ""
            s = str(v).strip()
            s = re.sub(r"\s+", " ", s)
            return s

        def _norm_header(h: str) -> str:
            s = _norm(h)
            # Arabic normalization for headers
            s = s.replace("إ", "ا").replace("أ", "ا").replace("آ", "ا")
            s = s.replace("ى", "ي").replace("ة", "ه")
            s = re.sub(r"[\u0640]", "", s)   # tatweel
            s = re.sub(r"[\s\-_/]+", "", s)
            return s.lower()

        def _guess_role(job_title: str) -> str:
            t = _norm(job_title)
            t2 = t.replace("مديرعام", "مدير عام")
            if not t2:
                return "USER"
            # secretary general
            if ("امين عام" in t2) or ("الأمين العام" in t2) or ("الامين العام" in t2):
                return "SECRETARY_GENERAL"
            # directorate head (general manager)
            if "مدير عام" in t2:
                return "DIRECTORATE_HEAD"
            # finance keyword (optional)
            if "مالية" in t2 or "المالي" in t2:
                return "FINANCE"
            # department head
            if "مدير" in t2:
                return "DEPT_HEAD"
            return "USER"

        def _resolve_seed_path() -> str | None:
            for p in SEED_XLSX_CANDIDATES:
                try:
                    if p and os.path.exists(p):
                        return p
                except Exception:
                    continue
            return None

        def seed_from_excel() -> int:
            seed_path = _resolve_seed_path()
            if not seed_path:
                msg = (
                    "[seed] seed.xlsx NOT FOUND.\n"
                    "  - Put seed.xlsx next to init_db.py OR\n"
                    "  - Set env var SEED_XLSX to the full path of the file.\n"
                    "  - Current candidates: " + ", ".join(SEED_XLSX_CANDIDATES)
                )
                if SEED_OPTIONAL:
                    print(msg)
                    print("[seed] SEED_OPTIONAL=1 -> skipping seeding.")
                    return 0
                raise SystemExit(msg)

            if load_workbook is None:
                msg = (
                    "[seed] openpyxl is required to read seed.xlsx.\n"
                    "Install it then re-run:  pip install openpyxl\n"
                    f"Seed file path: {seed_path}"
                )
                if SEED_OPTIONAL:
                    print(msg)
                    print("[seed] SEED_OPTIONAL=1 -> skipping seeding.")
                    return 0
                raise SystemExit(msg)

            print(f"[seed] Using seed file: {seed_path}")

            wb = load_workbook(seed_path, data_only=True)

            # Find a sheet that contains required headers
            ws = None
            needed = {
                _norm_header("الإسم"): ("الإسم", "الاسم", "الإسم الكامل"),
                _norm_header("الإدارة"): ("الإدارة", "الادارة"),
                _norm_header("الدور"): ("الدور", "المسمى الوظيفي", "الوظيفة"),
                _norm_header("الايميل"): ("الايميل", "الإيميل", "البريد الالكتروني", "البريد الإلكتروني", "email"),
            }

            def build_header_map(sheet):
                headers_norm = {}
                for col in range(1, sheet.max_column + 1):
                    h = sheet.cell(1, col).value
                    if not h:
                        continue
                    headers_norm[_norm_header(h)] = col
                return headers_norm

            for name in wb.sheetnames:
                sh = wb[name]
                hm = build_header_map(sh)
                ok = True
                for k_norm, aliases in needed.items():
                    found = False
                    for a in aliases:
                        if _norm_header(a) in hm:
                            found = True
                            break
                    if not found:
                        ok = False
                        break
                if ok:
                    ws = sh
                    break

            if ws is None:
                # fall back to active sheet
                ws = wb.active

            headers_norm = build_header_map(ws)

            def col_of(*aliases):
                for a in aliases:
                    k = _norm_header(a)
                    if k in headers_norm:
                        return headers_norm[k]
                return None

            c_name = col_of("الإسم", "الاسم", "الإسم الكامل")
            c_dir = col_of("الإدارة", "الادارة")
            c_dept = col_of("الدائرة", "الدائره", "القسم")
            c_title = col_of("الدور", "المسمى الوظيفي", "الوظيفة")
            c_email = col_of("الايميل", "الإيميل", "البريد الالكتروني", "البريد الإلكتروني", "email")

            if not all([c_name, c_dir, c_title, c_email]):
                msg = (
                    "[seed] Missing required columns in seed.xlsx.\n"
                    "Expected headers like: (الإسم/الاسم), (الإدارة/الادارة), (الدور/المسمى الوظيفي), (الايميل/البريد الإلكتروني)."
                )
                raise SystemExit(msg)

            # Create (or get) base organization
            org = Organization.query.filter_by(code="PNCECS").first()
            if not org:
                org = Organization(
                    name_ar="اللجنة الوطنية الفلسطينية للتربية والثقافة والعلوم",
                    name_en="Palestinian National Commission for Education, Culture and Science",
                    code="PNCECS",
                    is_active=True,
                )
                db.session.add(org)
                db.session.flush()

            # caches
            dir_cache = {}    # directorate name -> Directorate
            dept_cache = {}   # (directorate_id, dept_name) -> Department

            def get_or_create_directorate(name_ar: str) -> Directorate:
                name_ar = _norm(name_ar)
                if name_ar == "الأمين العام":
                    name_ar = "الأمانة العامة"
                if not name_ar:
                    name_ar = "غير محددة"

                if name_ar in dir_cache:
                    return dir_cache[name_ar]

                d = Directorate.query.filter_by(organization_id=org.id, name_ar=name_ar).first()
                if not d:
                    d = Directorate(organization_id=org.id, name_ar=name_ar, is_active=True)
                    db.session.add(d)
                    db.session.flush()
                dir_cache[name_ar] = d
                return d

            def get_or_create_department(directorate_id: int, name_ar: str) -> Department:
                name_ar = _norm(name_ar)
                if not name_ar:
                    name_ar = "غير محددة"
                key = (directorate_id, name_ar)
                if key in dept_cache:
                    return dept_cache[key]

                dep = Department.query.filter_by(directorate_id=directorate_id, name_ar=name_ar).first()
                if not dep:
                    dep = Department(directorate_id=directorate_id, name_ar=name_ar, is_active=True)
                    db.session.add(dep)
                    db.session.flush()
                dept_cache[key] = dep
                return dep

            created_users = 0

            for r in range(2, ws.max_row + 1):
                name = _norm(ws.cell(r, c_name).value)
                dir_name = _norm(ws.cell(r, c_dir).value)
                dept_name = _norm(ws.cell(r, c_dept).value) if c_dept else ""
                title = _norm(ws.cell(r, c_title).value)
                email = _norm(ws.cell(r, c_email).value).lower()

                # skip empty rows
                if not any([name, dir_name, dept_name, title, email]):
                    continue

                dir_obj = get_or_create_directorate(dir_name)

                # if department is missing, create an "office" department so department_id is always set
                if not dept_name:
                    dept_name = f"مكتب {dir_obj.name_ar}"
                dept_obj = get_or_create_department(dir_obj.id, dept_name)

                if not email:
                    safe = re.sub(r"[^a-z0-9]+", ".", name.lower(), flags=re.I).strip(".")
                    email = f"{safe or 'user'}{r}@pncecs.local"

                role_code = _guess_role(title)

                u = User.query.filter_by(email=email).first()
                if not u:
                    u = User(
                        email=email,
                        name=name or None,
                        job_title=title or None,
                        password_hash=generate_password_hash("123"),
                        role=role_code,
                        department_id=dept_obj.id,
                    )
                    db.session.add(u)
                    created_users += 1
                else:
                    u.name = name or u.name
                    u.job_title = title or u.job_title
                    u.department_id = dept_obj.id
                    u.role = u.role or role_code

            print(f"[seed] Done. Created users: {created_users}")
            return created_users

        # =========================
        # Seed: Departments (دوائر) fixed list
        # =========================
        def seed_departments_fixed_list() -> int:
            """Upsert departments based on the fixed table provided by the admin.

            This complements seed_from_excel():
            - Adds code + English name + active flag
            - Avoids duplicates by preferring match on code, otherwise match on (directorate + Arabic name) where code is empty.
            """
            org = Organization.query.filter_by(code="PNCECS").first()
            if not org:
                org = Organization(
                    name_ar="اللجنة الوطنية الفلسطينية للتربية والثقافة والعلوم",
                    name_en="Palestinian National Commission for Education, Culture and Science",
                    code="PNCECS",
                    is_active=True,
                )
                db.session.add(org)
                db.session.flush()

            rows = [
                # (directorate_ar, dept_ar, dept_en, code, is_active)
                ("الإدارة العامة للمنظمات والمتخصصة", "دائرة الألكسو", "Alecso Dept", "3", True),
                ("الأمانة العامة", "دائرة الأمانة العامة", "General Secretariat Dept", "5", True),
                ("الأمانة العامة", "دائرة الإعلام", "Media Dept", "10", True),
                ("الإدارة العامة للبرامج والمشاريع", "دائرة البرامج", "Programs Dept", "14", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة التخطيط", "Planning Dept", "11", True),
                ("الإدارة العامة للمنظمات والمتخصصة", "دائرة التربية", "Education Dept", "18", True),
                ("الإدارة العامة للمنظمات والمتخصصة", "دائرة الثقافة", "Culture Dept", "17", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة الرقابة الداخلية", "Internal Audit", "9", True),
                ("الأمانة العامة", "دائرة السكرتاريا", "Secretary Dept", "13", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة الشؤون الإدارية", "Administrative Affairs Dept", "8", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة الشؤون الإدارية", "Administrative Affairs Dept", "12", True),
                ("الإدارة العامة للمنظمات والمتخصصة", "دائرة العلوم الإنسانية", "Humanities Dept", "19", True),
                ("الإدارة العامة للمنظمات والمتخصصة", "دائرة العلوم البيئية", "Environmental Sciences Dept", "15", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة المراسلات", "Correspondence Dept", "7", True),
                ("الإدارة العامة للبرامج والمشاريع", "دائرة المشاريع", "Project Management Dept", "1", True),
                ("الإدارة العامة للمطبوعات", "دائرة المطبوعات", "Publications Dept", "2", True),
                ("الإدارة العامة للمطبوعات", "دائرة المطبوعات", "Publications Department", "6", True),
                ("الإدارة العامة للشؤون الإدارية", "دائرة مالية اللجنة الوطنية الفلسطينية للتربية والثقافة والعلوم", "Finance Dept", "20", True),
                ("الإدارة العامة للبرامج والمشاريع", "دائرة مالية المشاريع", "Project Finance Dept", "16", True),
            ]

            # cache directorates by Arabic name
            dir_cache: dict[str, Directorate] = {}

            def get_or_create_dir(name_ar: str) -> Directorate:
                key = _norm(name_ar)
                if key in dir_cache:
                    return dir_cache[key]
                d = Directorate.query.filter_by(organization_id=org.id, name_ar=key).first()
                if not d:
                    d = Directorate(organization_id=org.id, name_ar=key, is_active=True)
                    db.session.add(d)
                    db.session.flush()
                dir_cache[key] = d
                return d

            changed = 0
            for dir_ar, dept_ar, dept_en, code, active in rows:
                d = get_or_create_dir(dir_ar)
                dept_ar_n = _norm(dept_ar)
                dept_en_n = _norm(dept_en) or None
                code_n = _norm(code) or None

                dep = None
                if code_n:
                    dep = Department.query.filter_by(directorate_id=d.id, code=code_n).first()

                if not dep:
                    # Prefer matching an existing unnamed-code row with same Arabic name
                    dep = Department.query.filter(
                        Department.directorate_id == d.id,
                        Department.name_ar == dept_ar_n,
                        or_(Department.code.is_(None), Department.code == "")
                    ).first()

                if not dep:
                    # Fallback: exact match by name + code if exists
                    q = Department.query.filter(
                        Department.directorate_id == d.id,
                        Department.name_ar == dept_ar_n,
                    )
                    if code_n:
                        q = q.filter(Department.code == code_n)
                    dep = q.first()

                if not dep:
                    dep = Department(
                        directorate_id=d.id,
                        name_ar=dept_ar_n,
                        name_en=dept_en_n,
                        code=code_n,
                        is_active=bool(active),
                    )
                    db.session.add(dep)
                    changed += 1
                else:
                    # update
                    dep.name_ar = dept_ar_n
                    dep.name_en = dept_en_n
                    dep.code = code_n
                    dep.is_active = bool(active)
                    changed += 1

            db.session.flush()
            return changed

        

        # =========================
        # 3.5️⃣ Seed Extra Master Data (Roles + Request Types + Sample Template)
        # =========================
        def seed_extra_master_data():
            # ---- Roles (Master Data) ----
            roles = [
                # code, name_ar, name_en, active
                ("sect_head", "رئيس قسم", "Section Head", True),
                ("devision_head", "رئيس شعبة", "Division Head", True),
                ("General_secretary", "الأمين العام", "General Secretary", True),
                ("dept_head", "مدير دائرة", "Department Head", True),
                ("directorate_head", "مدير عام الإدارة", "Directorate Head", True),
                ("PNCECS_Head", "رئيس اللجنة الوطنية الفلسطينية للتربية والثقافة والعلوم", "President of National Commission", True),
                ("employee", "موظف", "Employee", True),
                ("ADMIN", "مدير النظام", "System Admin", True),
                ("SUPER_ADMIN", "سوبر أدمن", "Super Admin", True),
                ("USER", "مستخدم", "User", True),
            ]

            for code, name_ar, name_en, active in roles:
                r = Role.query.filter_by(code=code).first()
                if not r:
                    r = Role(code=code)
                    db.session.add(r)
                r.name_ar = name_ar
                r.name_en = name_en
                r.is_active = bool(active)

            db.session.flush()

            # ---- Request Types ----
            request_types = [
                ("purchase_request", "طلب شراء", "Purchase request", True),
                ("project_review", "طلب مراجعة مشروع", "Project Review", True),
                ("signature_request", "طلب توقيع", "Signature Request", True),
                ("approval_request", "طلب إعتماد", "Application for approval", True),
            ]
            for code, name_ar, name_en, active in request_types:
                rt = RequestType.query.filter_by(code=code).first()
                if not rt:
                    rt = RequestType(code=code, name_ar=name_ar, name_en=name_en, is_active=bool(active))
                    db.session.add(rt)
                else:
                    rt.name_ar = name_ar
                    rt.name_en = name_en
                    rt.is_active = bool(active)

            db.session.flush()

            # ---- Sample Template: "مسار طلب مالي مشاريع" ----
            template_name = "مسار طلب مالي مشاريع"
            t = WorkflowTemplate.query.filter_by(name=template_name).first()
            if not t:
                creator = User.query.filter(User.role.in_(["SUPER_ADMIN", "ADMIN"])).order_by(User.id.asc()).first()
                created_by_id = creator.id if creator else None

                t = WorkflowTemplate(
                    name=template_name,
                    sla_days_default=5,
                    is_active=True,
                    created_by_id=created_by_id
                )
                db.session.add(t)
                db.session.flush()

                # Lookups
                user1 = User.query.filter_by(email="adham.pncecs@gmail.com").first()

                org = Organization.query.filter_by(code="PNCECS").first()

                # Ensure directorates exist (create if missing)
                dir_programs_projects = Directorate.query.filter(
                    Directorate.name_ar.ilike("%الإدارة العامة للبرامج%")
                ).first()
                if not dir_programs_projects and org:
                    dir_programs_projects = Directorate(
                        organization_id=org.id,
                        name_ar="الإدارة العامة للبرامج والمشاريع",
                        name_en="General Directorate of Programs and Projects",
                        is_active=True
                    )
                    db.session.add(dir_programs_projects)
                    db.session.flush()

                dir_admin_affairs = Directorate.query.filter(
                    Directorate.name_ar.ilike("%الشؤون الإدارية%")
                ).first()
                if not dir_admin_affairs and org:
                    dir_admin_affairs = Directorate(
                        organization_id=org.id,
                        name_ar="الإدارة العامة للشؤون الإدارية",
                        name_en="General Directorate of Administrative Affairs",
                        is_active=True
                    )
                    db.session.add(dir_admin_affairs)
                    db.session.flush()

                # Ensure department exists (create if missing)
                dept_fin_projects = Department.query.filter(
                    Department.name_ar.ilike("%دائرة مالية المشاريع%")
                ).first()
                if not dept_fin_projects and dir_programs_projects:
                    dept_fin_projects = Department(
                        directorate_id=dir_programs_projects.id,
                        name_ar="دائرة مالية المشاريع",
                        name_en="Projects Financial Department",
                        is_active=True
                    )
                    db.session.add(dept_fin_projects)
                    db.session.flush()

                steps = [
                    ("USER", {"user": user1}, 1),
                    ("DEPARTMENT", {"dept": dept_fin_projects}, 1),
                    ("DIRECTORATE", {"dir": dir_programs_projects}, 1),
                    ("DIRECTORATE", {"dir": dir_admin_affairs}, 1),
                    ("ROLE", {"role": "General_secretary"}, 1),
                ]

                order = 1
                for kind, target, sla_days in steps:
                    s = WorkflowTemplateStep(
                        template_id=t.id,
                        step_order=order,
                        approver_kind=kind,
                        approver_user_id=(target.get("user").id if kind == "USER" and target.get("user") else None),
                        approver_department_id=(target.get("dept").id if kind == "DEPARTMENT" and target.get("dept") else None),
                        approver_directorate_id=(target.get("dir").id if kind == "DIRECTORATE" and target.get("dir") else None),
                        approver_role=(target.get("role") if kind == "ROLE" else None),
                        sla_days=sla_days
                    )
                    db.session.add(s)
                    order += 1

                                # Optional: map request types to this template (helps testing)
                for code in ["purchase_request", "project_review", "signature_request", "approval_request"]:
                    rt = RequestType.query.filter_by(code=code).first()
                    if not rt:
                        continue
                    rule = WorkflowRoutingRule.query.filter_by(
                        request_type_id=rt.id,
                        template_id=t.id,
                        is_active=True
                    ).first()
                    if not rule:
                        rule = WorkflowRoutingRule(
                            request_type_id=rt.id,
                            template_id=t.id,
                            organization_id=None,   # global fallback (works even if user has no dept/org)
                            directorate_id=None,
                            department_id=None,
                            priority=10,
                            is_active=True
                        )
                        db.session.add(rule)


# Run the Excel seeder first (so org structure exists)
        seed_from_excel()

        # Seed departments list (codes + EN names) as per admin fixed table
        try:
            changed = seed_departments_fixed_list()
            if changed:
                print(f"[seed] Departments fixed list applied. Changed/Upserted: {changed}")
        except Exception as e:
            db.session.rollback()
            print(f"[seed] Departments fixed list failed: {e}")

        # 4️⃣ Seed Users (FIXED)
        # =========================
        def get_or_create_user(email, password, role="USER", department_id=None):
            u = User.query.filter_by(email=email).first()
            if u:
                return u, False
            u = User(
                email=email,
                password_hash=generate_password_hash(password),
                role=role,
                department_id=department_id
            )
            db.session.add(u)
            return u, True

        seeded = []

        admin_email = "admin@pncecs.org"
        admin_password = "123"
        u, created = get_or_create_user(admin_email, admin_password, role="ADMIN")
        if created:
            print("Admin user created")
        seeded.append(("ADMIN", admin_email, admin_password))

        # Default SUPER_ADMIN (inherits all Admin permissions)
        super_email = "superadmin@pncecs.org"
        super_password = "123"
        u, created = get_or_create_user(super_email, super_password, role="SUPER_ADMIN")
        if created:
            print("SUPER_ADMIN user created")
        seeded.append(("SUPER_ADMIN", super_email, super_password))

        user1_email = "adham.pncecs@gmail.com"
        user1_password = "123"
        u, created = get_or_create_user(user1_email, user1_password, role="USER")
        if created:
            print("User created:", user1_email)
        seeded.append(("USER", user1_email, user1_password))

        user2_email = "mo@gmail.com"
        user2_password = "123"
        u, created = get_or_create_user(user2_email, user2_password, role="USER")
        if created:
            print("User created:", user2_email)
        seeded.append(("USER", user2_email, user2_password))

        user3_email = "ta.pncecs@gmail.com"
        user3_password = "123"
        u, created = get_or_create_user(user3_email, user3_password, role="USER")
        if created:
            print("User created:", user3_email)
        seeded.append(("USER", user3_email, user3_password))


        # =========================
        # 4.9️⃣ Seed Extra Master Data (after users)
        # =========================
        seed_extra_master_data()

        # =========================
        # 4.95️⃣ Apply User Role Overrides (By Email)
        # =========================
        def apply_user_role_overrides():
            mapping = {
                "irene2.pncecs@mail.com": "directorate_head",
                "belal.pncecs@gmail.com": "dept_head",
                "bayan.pncecs@gmail.com": "USER",
                "khlud.pncecs@gmail.com": "directorate_head",
                "jihad.pncecs@gmail.com": "General_secretary",
                "raed.pncecs@gmail.com": "directorate_head",
                "saleem.pncecs@gmail.com": "USER",
                "samir.pncecs@gmail.com": "USER",
                "sawsan.pncecs@gmail.com": "dept_head",
                "fadi.pncecs@gmail.com": "dept_head",
                "hamdan.pncecs@gmail.com": "dept_head",
                "harfoush82@yahoo.com": "dept_head",
                "motaz.elbh1@gmail.com": "directorate_head",
                "safi.pncecs2@gmail.com": "directorate_head",
                "dua.pncecs@gmail.com": "USER",
                "ruba.pncecs@gmail.com": "USER",
                "adham.pncecs@gmail.com": "USER",
                "raheeq.pncecs@gmail.com": "USER",
                "shorouq.pncecs@gmail.com": "dept_head",
                "ayman.pncecs@gmail.com": "USER",
                "noor.pncecs@gmail.com": "USER",
                "shawkat.pncecs@gmail.com": "USER",
                "majd.pncecs@gmail.com": "dept_head",
                "admin@pncecs.org": "ADMIN",
                "superadmin@pncecs.org": "SUPER_ADMIN",
                "mo@gmail.com": "USER",
                "ta.pncecs@gmail.com": "USER",
            }

            updated = 0
            for email, role in mapping.items():
                u = User.query.filter_by(email=email).first()
                if not u:
                    continue
                u.role = role
                updated += 1
            print(f"[seed] Role overrides applied: {updated}")

        apply_user_role_overrides()


        # =========================
        # 5️⃣ Commit نهائي
        # =========================

        db.session.commit()

        print("===================================")
        print("Database initialized successfully")
        print("===================================")
        print("Login credentials:")
        for role, em, pw in seeded:
            print(f"{role:<12} -> {em} / {pw}")


if __name__ == "__main__":
    init_database()