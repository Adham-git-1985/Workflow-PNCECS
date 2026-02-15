# reset_and_import_masterdata.py
# تشغيل:
#   python reset_and_import_masterdata.py --dir seed_excels --force
#   أو:
#   python reset_and_import_masterdata.py --files organizations_*.xlsx directorates_*.xlsx ...
#
# ملاحظات:
# - لا يحذف teams ولا workflow_routing_rules
# - يعطّل فقط قواعد التوجيه التي تشير إلى بيانات سيتم حذفها (templates/org units)
# - استخدم --dry-run لعمل اختبار بدون حفظ

import argparse
import glob
import os
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy import inspect, text


# -----------------------
# 1) تحميل Flask app + db + models
# -----------------------
def load_app_db_and_models():
    app = None
    db = None

    try:
        from app import app as _app  # type: ignore
        app = _app
    except Exception:
        try:
            from app import create_app  # type: ignore
            app = create_app()
        except Exception as e:
            raise RuntimeError("لم أستطع تحميل Flask app. عدّل load_app_db_and_models.") from e

    try:
        from app import db as _db  # type: ignore
        db = _db
    except Exception:
        try:
            from extensions import db as _db  # type: ignore
            db = _db
        except Exception as e:
            raise RuntimeError("لم أستطع تحميل db (SQLAlchemy). عدّل load_app_db_and_models.") from e

    try:
        import models as models_mod  # type: ignore
    except Exception as e:
        raise RuntimeError("لم أستطع استيراد models.py.") from e

    Organization = getattr(models_mod, "Organization", None)
    Directorate = getattr(models_mod, "Directorate", None)
    Department = getattr(models_mod, "Department", None)
    Section = getattr(models_mod, "Section", None)
    WorkflowTemplate = getattr(models_mod, "WorkflowTemplate", None) or getattr(models_mod, "WorkflowDefinition", None)

    missing = [k for k, v in {
        "Organization": Organization,
        "Directorate": Directorate,
        "Department": Department,
        "Section": Section,
        "WorkflowTemplate/WorkflowDefinition": WorkflowTemplate,
    }.items() if v is None]

    if missing:
        raise RuntimeError("موديلات ناقصة: " + ", ".join(missing))

    return app, db, Organization, Directorate, Department, Section, WorkflowTemplate


# -----------------------
# 2) Helpers (import)
# -----------------------
YES_VALUES = {"نعم", "yes", "true", "1", 1, True}
NO_VALUES = {"لا", "no", "false", "0", 0, False}

def to_bool(val: Any) -> Optional[bool]:
    if val is None:
        return None
    if isinstance(val, str):
        v = val.strip().lower()
        if v in {x.lower() for x in YES_VALUES if isinstance(x, str)}:
            return True
        if v in {x.lower() for x in NO_VALUES if isinstance(x, str)}:
            return False
    if val in YES_VALUES:
        return True
    if val in NO_VALUES:
        return False
    return None

def to_int(val: Any) -> Optional[int]:
    if val is None or val == "":
        return None
    try:
        return int(val)
    except Exception:
        try:
            return int(float(val))
        except Exception:
            return None

def normalize_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None

def read_excel_rows(path: str) -> Tuple[str, List[Dict[str, Any]]]:
    wb = load_workbook(path)
    ws = wb[wb.sheetnames[0]]
    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        row = {}
        for h, v in zip(headers, r):
            if h:
                row[h] = v
        rows.append(row)

    return ws.title, rows

def set_attr_if_exists(obj: Any, attr: str, value: Any) -> None:
    if value is None:
        return
    if hasattr(obj, attr):
        setattr(obj, attr, value)

def get_first_attr_name(obj_or_cls: Any, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if hasattr(obj_or_cls, c):
            return c
    return None

def upsert_by(db, Model, where: Dict[str, Any], values: Dict[str, Any]):
    q = db.session.query(Model)
    for k, v in where.items():
        if not hasattr(Model, k):
            continue
        q = q.filter(getattr(Model, k) == v)
    obj = q.first()

    if obj is None:
        obj = Model()
        for k, v in {**where, **values}.items():
            set_attr_if_exists(obj, k, v)
        db.session.add(obj)
    else:
        for k, v in values.items():
            set_attr_if_exists(obj, k, v)

    return obj


# -----------------------
# 3) Importers لكل شاشة
# -----------------------
def import_organizations(db, Organization, rows: List[Dict[str, Any]]):
    name_ar_key = "الاسم (AR)"
    name_en_key = "الاسم (EN)"
    code_key = "Code"
    active_key = "نشط"

    field_name_ar = get_first_attr_name(Organization, ["name_ar", "ar_name", "title_ar"])
    field_name_en = get_first_attr_name(Organization, ["name_en", "en_name", "title_en"])
    field_code = get_first_attr_name(Organization, ["code"])
    field_active = get_first_attr_name(Organization, ["is_active", "active", "enabled"])

    for r in rows:
        name_ar = normalize_str(r.get(name_ar_key))
        name_en = normalize_str(r.get(name_en_key))
        code = normalize_str(r.get(code_key))
        is_active = to_bool(r.get(active_key))

        where = {}
        if field_code and code:
            where[field_code] = code
        elif field_name_ar and name_ar:
            where[field_name_ar] = name_ar
        else:
            continue

        values = {}
        if field_name_ar: values[field_name_ar] = name_ar
        if field_name_en: values[field_name_en] = name_en
        if field_code: values[field_code] = code
        if field_active and is_active is not None: values[field_active] = is_active

        upsert_by(db, Organization, where, values)

def import_directorates(db, Organization, Directorate, rows: List[Dict[str, Any]]):
    org_name_key = "المنظمة"
    name_ar_key = "الاسم (AR)"
    name_en_key = "الاسم (EN)"
    code_key = "Code"
    active_key = "نشط"

    dir_name_ar = get_first_attr_name(Directorate, ["name_ar", "ar_name", "title_ar"])
    dir_name_en = get_first_attr_name(Directorate, ["name_en", "en_name", "title_en"])
    dir_code = get_first_attr_name(Directorate, ["code"])
    dir_active = get_first_attr_name(Directorate, ["is_active", "active", "enabled"])
    org_fk = get_first_attr_name(Directorate, ["organization_id", "org_id"])

    org_name_ar_field = get_first_attr_name(Organization, ["name_ar", "ar_name", "title_ar"])
    org_code_field = get_first_attr_name(Organization, ["code"])

    for r in rows:
        org_name = normalize_str(r.get(org_name_key))
        name_ar = normalize_str(r.get(name_ar_key))
        name_en = normalize_str(r.get(name_en_key))
        code = normalize_str(r.get(code_key))
        is_active = to_bool(r.get(active_key))

        if not (org_fk and org_name and org_name_ar_field and name_ar and dir_name_ar):
            continue

        org = db.session.query(Organization).filter(getattr(Organization, org_name_ar_field) == org_name).first()
        if org is None:
            org_where = {org_name_ar_field: org_name}
            org_values = {}
            if org_code_field: org_values[org_code_field] = None
            org = upsert_by(db, Organization, org_where, org_values)

        where = {org_fk: org.id, dir_name_ar: name_ar}
        values = {org_fk: org.id, dir_name_ar: name_ar}
        if dir_name_en: values[dir_name_en] = name_en
        if dir_code: values[dir_code] = code
        if dir_active and is_active is not None: values[dir_active] = is_active

        upsert_by(db, Directorate, where, values)

def import_departments(db, Directorate, Department, rows: List[Dict[str, Any]]):
    directorate_key = "الإدارة"
    name_ar_key = "الدائرة (AR)"
    name_en_key = "الدائرة (EN)"
    code_key = "Code"
    active_key = "نشط"

    dep_name_ar = get_first_attr_name(Department, ["name_ar", "ar_name", "title_ar"])
    dep_name_en = get_first_attr_name(Department, ["name_en", "en_name", "title_en"])
    dep_code = get_first_attr_name(Department, ["code"])
    dep_active = get_first_attr_name(Department, ["is_active", "active", "enabled"])
    dir_fk = get_first_attr_name(Department, ["directorate_id", "dir_id"])

    dir_name_ar = get_first_attr_name(Directorate, ["name_ar", "ar_name", "title_ar"])

    for r in rows:
        dir_name = normalize_str(r.get(directorate_key))
        name_ar = normalize_str(r.get(name_ar_key))
        name_en = normalize_str(r.get(name_en_key))
        code = normalize_str(r.get(code_key))
        is_active = to_bool(r.get(active_key))

        if not (dir_fk and dir_name_ar and dir_name and name_ar and dep_name_ar):
            continue

        d = db.session.query(Directorate).filter(getattr(Directorate, dir_name_ar) == dir_name).first()
        if d is None:
            continue

        where = {dir_fk: d.id, dep_name_ar: name_ar}
        values = {dir_fk: d.id, dep_name_ar: name_ar}
        if dep_name_en: values[dep_name_en] = name_en
        if dep_code: values[dep_code] = code
        if dep_active and is_active is not None: values[dep_active] = is_active

        upsert_by(db, Department, where, values)

def import_sections(db, Directorate, Department, Section, rows: List[Dict[str, Any]]):
    directorate_key = "الإدارة"
    department_key = "الدائرة"
    name_ar_key = "القسم (AR)"
    name_en_key = "القسم (EN)"
    code_key = "Code"
    active_key = "نشط"

    sec_name_ar = get_first_attr_name(Section, ["name_ar", "ar_name", "title_ar"])
    sec_name_en = get_first_attr_name(Section, ["name_en", "en_name", "title_en"])
    sec_code = get_first_attr_name(Section, ["code"])
    sec_active = get_first_attr_name(Section, ["is_active", "active", "enabled"])
    dep_fk = get_first_attr_name(Section, ["department_id", "dept_id"])

    dir_name_ar = get_first_attr_name(Directorate, ["name_ar", "ar_name", "title_ar"])
    dep_name_ar = get_first_attr_name(Department, ["name_ar", "ar_name", "title_ar"])
    dep_dir_fk = get_first_attr_name(Department, ["directorate_id", "dir_id"])

    for r in rows:
        dir_name = normalize_str(r.get(directorate_key))
        dep_name = normalize_str(r.get(department_key))
        name_ar = normalize_str(r.get(name_ar_key))
        name_en = normalize_str(r.get(name_en_key))
        code = normalize_str(r.get(code_key))
        is_active = to_bool(r.get(active_key))

        if not (dep_fk and dir_name and dep_name and name_ar and dir_name_ar and dep_name_ar and dep_dir_fk and sec_name_ar):
            continue

        d = db.session.query(Directorate).filter(getattr(Directorate, dir_name_ar) == dir_name).first()
        if d is None:
            continue

        dep_q = db.session.query(Department).filter(getattr(Department, dep_name_ar) == dep_name)
        dep_q = dep_q.filter(getattr(Department, dep_dir_fk) == d.id)
        dep = dep_q.first()
        if dep is None:
            continue

        where = {dep_fk: dep.id, sec_name_ar: name_ar}
        values = {dep_fk: dep.id, sec_name_ar: name_ar}
        if sec_name_en: values[sec_name_en] = name_en
        if sec_code: values[sec_code] = code
        if sec_active and is_active is not None: values[sec_active] = is_active

        upsert_by(db, Section, where, values)

def import_workflow_templates(db, WorkflowTemplate, rows: List[Dict[str, Any]]):
    name_key = "اسم المسار"
    active_key = "نشط"
    sla_key = "SLA الافتراضي (أيام)"

    wt_name = get_first_attr_name(WorkflowTemplate, ["name", "name_ar", "title", "title_ar"])
    wt_active = get_first_attr_name(WorkflowTemplate, ["is_active", "active", "enabled"])
    wt_sla = get_first_attr_name(WorkflowTemplate, ["default_sla_days", "sla_days", "sla_default_days"])

    for r in rows:
        name = normalize_str(r.get(name_key))
        is_active = to_bool(r.get(active_key))
        sla_days = to_int(r.get(sla_key))

        if not (wt_name and name):
            continue

        where = {wt_name: name}
        values = {}
        if wt_active and is_active is not None: values[wt_active] = is_active
        if wt_sla and sla_days is not None: values[wt_sla] = sla_days

        upsert_by(db, WorkflowTemplate, where, values)


# -----------------------
# 4) File routing
# -----------------------
def detect_screen_from_filename(filename: str) -> Optional[str]:
    base = os.path.basename(filename).lower()
    for key in ["organizations", "directorates", "departments", "sections", "workflow_templates"]:
        if base.startswith(key + "_") or base == (key + ".xlsx") or base.startswith(key + "-"):
            return key
    return None

IMPORT_ORDER = ["organizations", "directorates", "departments", "sections", "workflow_templates"]


# -----------------------
# 5) CLEANUP (delete + selective disable rules)
# -----------------------
def table_exists(db, table_name: str) -> bool:
    return table_name in inspect(db.engine).get_table_names()

def col_exists(db, table_name: str, col: str) -> bool:
    if not table_exists(db, table_name):
        return False
    cols = [c["name"] for c in inspect(db.engine).get_columns(table_name)]
    return col in cols

def sqlite_toggle_fk(db, on: bool) -> None:
    try:
        if db.engine.dialect.name == "sqlite":
            db.session.execute(text(f"PRAGMA foreign_keys={'ON' if on else 'OFF'}"))
    except Exception:
        pass

def delete_all(db, table_name: str) -> int:
    before = int(db.session.execute(text(f"SELECT COUNT(*) FROM {table_name}")).scalar() or 0)
    db.session.execute(text(f"DELETE FROM {table_name}"))
    return before

def set_null_if_exists(db, table_name: str, col: str) -> None:
    if table_exists(db, table_name) and col_exists(db, table_name, col):
        db.session.execute(text(f"UPDATE {table_name} SET {col} = NULL"))

def get_active_col(db, table_name: str) -> Optional[str]:
    for c in ["is_active", "active", "enabled"]:
        if col_exists(db, table_name, c):
            return c
    return None

def selective_disable_broken_routing_rules(db) -> None:
    """
    يعطّل فقط القواعد التي تشير إلى بيانات سنحذفها:
    - template_id IN (SELECT id FROM workflow_templates)
    - organization_id/directorate_id/department_id/section_id IN (SELECT id FROM ...)

    لا يحذف rules إطلاقاً.
    """
    t = "workflow_routing_rules"
    if not table_exists(db, t):
        return

    active_col = get_active_col(db, t)
    if not active_col:
        print("⚠️ workflow_routing_rules موجودة ولكن لا يوجد عمود is_active/active/enabled لتعطيلها.")
        return

    conditions = []

    # templates
    if col_exists(db, t, "template_id") and table_exists(db, "workflow_templates"):
        conditions.append(f"template_id IN (SELECT id FROM workflow_templates)")

    # org units
    if col_exists(db, t, "organization_id") and table_exists(db, "organizations"):
        conditions.append(f"organization_id IN (SELECT id FROM organizations)")
    if col_exists(db, t, "directorate_id") and table_exists(db, "directorates"):
        conditions.append(f"directorate_id IN (SELECT id FROM directorates)")
    if col_exists(db, t, "department_id") and table_exists(db, "departments"):
        conditions.append(f"department_id IN (SELECT id FROM departments)")
    if col_exists(db, t, "section_id") and table_exists(db, "sections"):
        conditions.append(f"section_id IN (SELECT id FROM sections)")

    if not conditions:
        print("ℹ️ لا توجد أعمدة (template_id/org_unit ids) في workflow_routing_rules لتعطيل القواعد المرتبطة.")
        return

    where_sql = " OR ".join(conditions)
    sql = f"UPDATE {t} SET {active_col}=0 WHERE ({where_sql})"
    db.session.execute(text(sql))
    print("✅ Disabled broken routing rules (only those referencing deleted templates/org units).")


def cleanup_masterdata(db) -> None:
    """
    يحذف: templates + steps + parallel_assignees + org structure (بدون teams) + org_unit_manager
    ويبقي: teams + workflow_routing_rules (مع تعطيل القواعد المكسورة فقط).
    """
    # مهم لتفادي FK blockers لو عندك runtime قديم
    set_null_if_exists(db, "workflow_instances", "template_id")

    # عطّل القواعد المكسورة قبل الحذف
    selective_disable_broken_routing_rules(db)

    delete_plan = [
        "workflow_template_parallel_assignees",
        "workflow_template_steps",
        "workflow_templates",
        "org_unit_manager",
        "sections",
        "departments",
        "directorates",
        "organizations",
    ]

    print("\n>>> Deleting selected tables (teams + routing rules are kept)...")
    for t in delete_plan:
        if table_exists(db, t):
            before = delete_all(db, t)
            print(f"Deleted {t} (was {before})")


# -----------------------
# 6) MAIN
# -----------------------
def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dir", help="مجلد يحتوي ملفات xlsx", default=None)
    parser.add_argument("--files", nargs="*", help="مسارات ملفات xlsx مباشرة", default=None)
    parser.add_argument("--dry-run", action="store_true", help="قراءة فقط بدون commit")
    parser.add_argument("--force", action="store_true", help="نفّذ التنظيف + الاستيراد فعلياً")

    args = parser.parse_args()

    paths: List[str] = []
    if args.files:
        for p in args.files:
            paths.extend(glob.glob(p))
    if args.dir:
        paths.extend(glob.glob(os.path.join(args.dir, "*.xlsx")))

    paths = sorted(set(paths))
    if not paths:
        raise SystemExit("لم يتم العثور على أي ملفات xlsx. استخدم --dir أو --files")

    app, db, Organization, Directorate, Department, Section, WorkflowTemplate = load_app_db_and_models()

    # صنّف الملفات حسب screen
    by_screen: Dict[str, List[str]] = {}
    for p in paths:
        screen = detect_screen_from_filename(p)
        if not screen:
            continue
        by_screen.setdefault(screen, []).append(p)

    with app.app_context():
        print("\n=== PLAN ===")
        print("- Will clean: orgs/directorates/departments/sections + templates(steps/parallel) + org_unit_manager")
        print("- Will keep: teams + workflow_routing_rules (disable only broken ones)")
        print(f"- Dry-run: {args.dry_run}")
        print(f"- Force: {args.force}")

        if not args.force:
            print("\n⚠️ لم يتم التنفيذ. أعد التشغيل مع --force")
            print("مثال: python reset_and_import_masterdata.py --dir seed_excels --force")
            return

        sqlite_toggle_fk(db, on=False)

        # 1) Clean
        cleanup_masterdata(db)

        # 2) Import
        print("\n=== Importing Excels ===")
        for screen in IMPORT_ORDER:
            files = by_screen.get(screen, [])
            if not files:
                continue

            print(f"\n--- Screen: {screen} ({len(files)} file(s)) ---")
            for f in files:
                sheet, rows = read_excel_rows(f)
                print(f"- {os.path.basename(f)} | sheet={sheet} | rows={len(rows)}")

                if screen == "organizations":
                    import_organizations(db, Organization, rows)
                elif screen == "directorates":
                    import_directorates(db, Organization, Directorate, rows)
                elif screen == "departments":
                    import_departments(db, Directorate, Department, rows)
                elif screen == "sections":
                    import_sections(db, Directorate, Department, Section, rows)
                elif screen == "workflow_templates":
                    import_workflow_templates(db, WorkflowTemplate, rows)

        sqlite_toggle_fk(db, on=True)

        if args.dry_run:
            print("\nDRY RUN: rollback (لن يتم حفظ أي شيء).")
            db.session.rollback()
        else:
            db.session.commit()
            print("\n✅ Done. Cleanup + Import committed.")


if __name__ == "__main__":
    main()
