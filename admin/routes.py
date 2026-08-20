from models import (
    FilePermission,
    InvWarehousePermission,
    PortalPermissionPreset,
    Role,
    RolePermission,
    StoreFilePermission,
    User,
    UserPermission,
)
from flask import (
    render_template, Blueprint,
    request, redirect, url_for, flash,
    send_file,
    current_app
)
from flask_login import login_required, current_user, logout_user
from utils.perms import perm_required
from permissions import roles_required, role_perm_required
from models import WorkflowRequest, SystemSetting, AuditLog, ArchivedFile, WorkflowRoutingRule, RequestType, Organization, Directorate, Department, WorkflowTemplate, RequestEscalation, OrgNode, OrgNodeType
from extensions import db
from sqlalchemy import func, or_
from datetime import datetime, timedelta
from filters.request_filters import apply_request_filters
from filters.request_filters import get_sla_days, get_escalation_days
from utils.ui_labels import ui_label
from sqlalchemy import case
from io import BytesIO

import os
import json
import hashlib
import shutil
import sqlite3
import zipfile
import tempfile
import uuid

from utils.excel import make_xlsx_bytes, make_xlsx_bytes_multi
from utils.importer import read_excel_rows, pick, to_str, to_int, to_bool, replace_all
from utils.org_dynamic import build_org_node_picker_tree
from portal.perm_defs import ALL_KEYS as PORTAL_ALL_KEYS

# =========================
# Blueprint
# =========================
admin_bp = Blueprint(
    "admin",
    __name__,
    url_prefix="/admin"
)

# Register sub-modules
from .evaluations import register_evaluation_routes
register_evaluation_routes(admin_bp)

# =========================
# Constants
# =========================
FINAL_STATUSES = ["APPROVED", "REJECTED"]

DASHBOARD_CACHE = {
    "data": None,
    "last_update": None
}

DASHBOARD_TTL_SECONDS = 30


ESCALATION_ROLE_MAP = {
    "dept_head": "secretary_general",
    "finance": "secretary_general"
}

SYSTEM_USER_ID = None  # system action


# =========================
# Helpers
# =========================
def get_sla_days():
    setting = SystemSetting.query.filter_by(key="SLA_DAYS").first()
    return int(setting.value) if setting else 3


def get_escalation_days():
    setting = SystemSetting.query.filter_by(key="ESCALATION_DAYS").first()
    return int(setting.value) if setting else 2


def get_trash_retention_days() -> int:
    """How many days deleted archive files remain in recycle bin before purge."""
    setting = SystemSetting.query.filter_by(key="TRASH_RETENTION_DAYS").first()
    try:
        return int(setting.value) if setting and setting.value is not None else 30
    except Exception:
        return 30


# =========================
# Update SLA
# =========================
@admin_bp.route("/update-sla", methods=["POST"])
@login_required
@perm_required("PORTAL_ADMIN_PERMISSIONS_MANAGE")  # SUPER/SUPERADMIN bypass exists in User.has_perm
def update_sla():

    sla_days = request.form.get("sla_days", type=int)

    if sla_days is None or sla_days <= 0:
        flash("قيمة مدة اتفاقية مستوى الخدمة غير صحيحة", "danger")
        return redirect(url_for("admin.dashboard"))

    setting = SystemSetting.query.filter_by(key="SLA_DAYS").first()

    if not setting:
        setting = SystemSetting(
            key="SLA_DAYS",
            value=str(sla_days)
        )
        db.session.add(setting)
    else:
        setting.value = str(sla_days)

    db.session.commit()

    flash(f"تم تحديث مدة اتفاقية مستوى الخدمة إلى {sla_days} يوم", "success")
    return redirect(url_for("admin.dashboard"))


# =========================
# Update Recycle Bin Retention
# =========================
@admin_bp.route("/update-trash-retention", methods=["POST"])
@login_required
@perm_required("PORTAL_ADMIN_PERMISSIONS_MANAGE")  # SUPER/SUPERADMIN bypass exists in User.has_perm
def update_trash_retention():
    days = request.form.get("trash_retention_days", type=int)
    if days is None or days < 1:
        flash("قيمة سياسة الاحتفاظ غير صحيحة", "danger")
        return redirect(url_for("admin.dashboard"))

    setting = SystemSetting.query.filter_by(key="TRASH_RETENTION_DAYS").first()
    if not setting:
        setting = SystemSetting(key="TRASH_RETENTION_DAYS", value=str(days))
        db.session.add(setting)
    else:
        setting.value = str(days)

    db.session.commit()
    flash(f"تم تحديث سياسة الاحتفاظ بسلة المحذوفات إلى {days} يوم", "success")
    return redirect(url_for("admin.dashboard"))


# =========================
# Admin Dashboard
# =========================
@admin_bp.route("/dashboard")
@login_required
@role_perm_required("VIEW_DASHBOARD")
def dashboard():
    now = datetime.utcnow()

    if (
        DASHBOARD_CACHE["data"]
        and DASHBOARD_CACHE["last_update"]
        and (now - DASHBOARD_CACHE["last_update"]).seconds < DASHBOARD_TTL_SECONDS
    ):
        return render_template(
            "admin/dashboard.html",
            **DASHBOARD_CACHE["data"]
        )

    stats = db.session.query(
        func.count(WorkflowRequest.id),
        func.sum(case((WorkflowRequest.status == "APPROVED", 1), else_=0)),
        func.sum(case((WorkflowRequest.status == "REJECTED", 1), else_=0)),
        func.sum(case((WorkflowRequest.status == "DRAFT", 1), else_=0)),
        func.sum(
            case(
                (WorkflowRequest.status.notin_(FINAL_STATUSES + ["DRAFT"]), 1),
                else_=0
            )
        )
    ).one()

    total, approved, rejected, drafts, in_progress = stats
    delegated = 0

    SLA_DAYS = get_sla_days()
    sla_threshold = now - timedelta(days=SLA_DAYS)

    aging_requests = (
        WorkflowRequest.query
        .filter(
            WorkflowRequest.status.notin_(FINAL_STATUSES),
            WorkflowRequest.created_at <= sla_threshold
        )
        .order_by(WorkflowRequest.created_at.asc())
        .limit(10)
        .all()
    )

    archive_total = ArchivedFile.query.count()
    archive_active = ArchivedFile.query.filter_by(is_deleted=False).count()
    archive_deleted = ArchivedFile.query.filter_by(is_deleted=True).count()

    context = {
        "counters": {
            "total": total,
            "approved": approved,
            "rejected": rejected,
            "drafts": drafts,
            "in_progress": in_progress,
            "delegated": delegated
        },
        "archive_counters": {
            "total": archive_total,
            "active": archive_active,
            "deleted": archive_deleted
        },
        "aging_requests": aging_requests,
        "sla_days": SLA_DAYS,
        "trash_retention_days": get_trash_retention_days(),
        "now": now
    }

    DASHBOARD_CACHE["data"] = context
    DASHBOARD_CACHE["last_update"] = now

    return render_template("admin/dashboard.html", **context)


@admin_bp.route("/dashboard/export.xlsx")
@login_required
@role_perm_required("VIEW_DASHBOARD")
def dashboard_export_excel():
    """Export dashboard counters + overdue list to Excel."""
    now = datetime.utcnow()
    SLA_DAYS = get_sla_days()
    sla_threshold = now - timedelta(days=SLA_DAYS)

    stats = db.session.query(
        func.count(WorkflowRequest.id),
        func.sum(case((WorkflowRequest.status == "APPROVED", 1), else_=0)),
        func.sum(case((WorkflowRequest.status == "REJECTED", 1), else_=0)),
        func.sum(case((WorkflowRequest.status == "DRAFT", 1), else_=0)),
        func.sum(
            case(
                (WorkflowRequest.status.notin_(FINAL_STATUSES + ["DRAFT"]), 1),
                else_=0
            )
        )
    ).one()

    total, approved, rejected, drafts, in_progress = stats
    archive_total = ArchivedFile.query.count()
    archive_active = ArchivedFile.query.filter_by(is_deleted=False).count()
    archive_deleted = ArchivedFile.query.filter_by(is_deleted=True).count()

    overdue = (
        WorkflowRequest.query
        .filter(
            WorkflowRequest.status.notin_(FINAL_STATUSES),
            WorkflowRequest.created_at <= sla_threshold
        )
        .order_by(WorkflowRequest.created_at.asc())
        .all()
    )

    headers = [
        "Request ID", "Title", "Status", "Created At", "Days Open",
        "Escalated", "Current Role"
        "PORTAL_VIEW",
        "HR_ATTENDANCE_IMPORT",
        "CORR_VIEW",
        "CORR_IN_CREATE",
        "CORR_OUT_CREATE",
        "CORR_MANAGE",
    ]

    rows = []
    for r in overdue:
        rows.append([
            r.id,
            r.title,
            ui_label(r.status),
            r.created_at.strftime("%Y-%m-%d %H:%M"),
            (now - r.created_at).days,
            "YES" if r.is_escalated else "NO",
            r.current_role,
        ])

    # Prepend summary rows (as plain rows)
    summary_headers = ["المؤشر", "القيمة"]
    summary_rows = [
        ("إجمالي الطلبات", total),
        ("موافق عليه", approved),
        ("مرفوض", rejected),
        ("المسودات", drafts),
        ("قيد التنفيذ", in_progress),
        ("أيام SLA", SLA_DAYS),
        ("إجمالي الأرشيف", archive_total),
        ("الأرشيف الفعال", archive_active),
        ("الأرشيف المحذوف", archive_deleted),
        ("أيام الاحتفاظ بسلة المحذوفات", get_trash_retention_days()),
        ("وقت التصدير", now.strftime("%Y-%m-%d %H:%M")),
    ]

    # Build workbook with two sheets
    try:
        # Create two sheets by generating bytes twice and merging is heavy;
        # Instead: build manually using openpyxl inside util function.
        import openpyxl
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Summary"[:31]
        ws1.append(list(summary_headers))
        for c in range(1, 3):
            cell = ws1.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for m, v in summary_rows:
            ws1.append([m, v])
        ws1.freeze_panes = "A2"
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 20

        ws2 = wb.create_sheet("Overdue")
        ws2.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws2.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for r in rows:
            ws2.append(r)
        ws2.freeze_panes = "A2"
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws2.column_dimensions[col_letter].width = 22

        bio = BytesIO()
        wb.save(bio)
        data = bio.getvalue()
    except Exception:
        # Fallback single-sheet export
        data = make_xlsx_bytes("Overdue", headers, rows)

    filename = f"admin_dashboard_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )



@admin_bp.route("/permissions", methods=["GET", "POST"])
@login_required
@perm_required("PORTAL_ADMIN_PERMISSIONS_MANAGE")  # SUPER/SUPERADMIN bypass exists in User.has_perm
def manage_permissions():
    """Manage permissions by ROLE (RolePermission table).

    Note: This is different from per-user CRUD permissions in masterdata/permissions.
    """
    # Prefer roles from master data table
    roles = Role.query.filter_by(is_active=True).order_by(Role.code.asc()).all()
    role_codes = [r.code for r in roles]

    # If roles table empty for some reason, fall back to roles in DB users
    if not role_codes:
        role_codes = [
            r for (r,) in db.session.query(User.role).distinct().order_by(User.role.asc()).all()
            if (r or "").strip()
        ]

    permissions = [
        "VIEW_DASHBOARD",
        "VIEW_ESCALATIONS",
        "CREATE_REQUEST",
        "APPROVE_REQUEST",
        "UPLOAD_ATTACHMENT",
        "SIGN_ARCHIVE",
        "DELETE_ARCHIVE",
        "VIEW_TIMELINE",
        "DELEGATION_MANAGE",
        "DELEGATION_SELF",
        "AUDIT_DASHBOARD_READ",
        "AUDIT_TIMELINE_READ",
        "WORKFLOW_NOTIFICATIONS_DASHBOARD_READ",
        "HR_EVALUATIONS_MANAGE",
        # Portal/HR keys (so role-based access works from this UI)
        "PORTAL_READ",
        "HR_SYSTEM_EVALUATION_VIEW",
    ]

    selected_role = (request.args.get("role") or "").strip()
    if request.method == "POST":
        selected_role = (request.form.get("role") or "").strip()
        perms = request.form.getlist("permissions")

        if not selected_role:
            flash("اختر دوراً وظيفياً.", "danger")
            return redirect(url_for("admin.manage_permissions"))

        known = set(permissions) | set(PORTAL_ALL_KEYS)

        # احذف فقط الصلاحيات المعروفة لتجنب مسح صلاحيات أخرى قد تكون أضيفت لاحقًا
        RolePermission.query.filter_by(role=selected_role).filter(RolePermission.permission.in_(known)).delete(synchronize_session=False)

        for p in perms:
            p = (p or "").strip()
            if p:
                db.session.add(RolePermission(role=selected_role, permission=p))

        db.session.commit()
        flash("تم تحديث صلاحيات الدور.", "success")
        return redirect(url_for("admin.manage_permissions", role=selected_role))

    # Pre-check existing permissions for selected role
    checked = set()
    if selected_role:
        rows = RolePermission.query.filter_by(role=selected_role).all()
        checked = { (r.permission or "").strip() for r in rows if r.permission }

    return render_template(
        "admin/permissions.html",
        role_codes=role_codes,
        permissions=permissions,
        selected_role=selected_role,
        checked=checked,
    )

@admin_bp.route("/requests")
@login_required
@perm_required("PORTAL_ADMIN_PERMISSIONS_MANAGE")  # SUPER/SUPERADMIN bypass exists in User.has_perm
def admin_requests():

    base_query = WorkflowRequest.query

    query = apply_request_filters(
        base_query,
        request.args
    )

    requests = query.order_by(
        WorkflowRequest.created_at.desc()
    ).all()

    return render_template(
        "admin/requests.html",
        requests=requests,
        is_admin=True
    )


@admin_bp.route("/requests/export.xlsx")
@login_required
@perm_required("PORTAL_ADMIN_PERMISSIONS_MANAGE")  # SUPER/SUPERADMIN bypass exists in User.has_perm
def admin_requests_export_excel():
    """Export admin requests list (with the same advanced filters) to Excel."""
    base_query = WorkflowRequest.query
    query = apply_request_filters(base_query, request.args)

    reqs = query.order_by(WorkflowRequest.created_at.desc()).all()

    headers = [
        "ID",
        "Title",
        "Status",
        "Created At",
        "Requester",
        "Current Role",
        "Request Type",
    ]

    rows = []
    for r in reqs:
        rt_label = ""
        try:
            if getattr(r, "request_type", None):
                rt_label = (r.request_type.name_ar or r.request_type.code or "")
        except Exception:
            rt_label = ""

        rows.append([
            r.id,
            r.title,
            r.status,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            (r.requester.email if getattr(r, "requester", None) else ""),
            r.current_role,
            rt_label,
        ])

    data = make_xlsx_bytes("Requests", headers, rows)
    filename = f"admin_requests_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@admin_bp.route("/escalations")
@login_required
@role_perm_required("VIEW_ESCALATIONS")
def escalations():

    now = datetime.utcnow()
    esc_deadline = now - timedelta(
        days=get_sla_days() + get_escalation_days()
    )

    # 1) SLA-overdue (legacy definition)
    escalated_requests = WorkflowRequest.query.filter(
        WorkflowRequest.status.notin_(["APPROVED", "REJECTED"]),
        WorkflowRequest.created_at < esc_deadline
    ).order_by(WorkflowRequest.created_at.asc()).all()

    # 2) Full escalation log (manual + system) stored in RequestEscalation
    manual_escalations = (
        RequestEscalation.query
        .order_by(RequestEscalation.created_at.desc(), RequestEscalation.id.desc())
        .limit(500)
        .all()
    )

    return render_template(
        "admin/escalations.html",
        requests=escalated_requests,
        escalations=manual_escalations,
    )


@admin_bp.route("/escalations/export.xlsx")
@login_required
@role_perm_required("VIEW_ESCALATIONS")
def escalations_export_excel():
    """Export escalations report to Excel.

    Includes two sheets:
      - Escalation_Log: all recorded escalations (manual + system) from RequestEscalation
      - SLA_Overdue: legacy SLA-overdue requests
    """
    now = datetime.utcnow()
    esc_deadline = now - timedelta(days=get_sla_days() + get_escalation_days())

    escalated_requests = WorkflowRequest.query.filter(
        WorkflowRequest.status.notin_(["APPROVED", "REJECTED"]),
        WorkflowRequest.created_at < esc_deadline
    ).order_by(WorkflowRequest.created_at.asc()).all()

    # Sheet 1: Escalation log
    limit = request.args.get("limit", type=int) or 10000
    escalation_log = (
        RequestEscalation.query
        .order_by(RequestEscalation.created_at.desc(), RequestEscalation.id.desc())
        .limit(limit)
        .all()
    )

    headers_log = [
        "Escalation ID",
        "Request ID",
        "الخطوة",
        "Category",
        "Created At",
        "From",
        "To (primary)",
        "Targets",
        "Description",
    ]

    rows_log = []
    for e in escalation_log:
        rows_log.append([
            e.id,
            e.request_id,
            getattr(e, "step_order", "") or "",
            e.category,
            e.created_at.strftime("%Y-%m-%d %H:%M") if e.created_at else "",
            (e.from_user.email if getattr(e, "from_user", None) else ""),
            (e.to_user.email if getattr(e, "to_user", None) else ""),
            getattr(e, "targets", "") or "",
            (e.description or "")[:2000],
        ])

    # Sheet 2: SLA-overdue requests (legacy)
    headers_overdue = [
        "ID",
        "Title",
        "Created At",
        "Status",
        "Current Role",
        "Days Open",
    ]

    rows_overdue = []
    for r in escalated_requests:
        days_open = (now - r.created_at).days if r.created_at else ""
        rows_overdue.append([
            r.id,
            r.title,
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
            r.status,
            r.current_role,
            days_open,
        ])

    data = make_xlsx_bytes_multi([
        ("Escalation_Log", headers_log, rows_log),
        ("SLA_Overdue", headers_overdue, rows_overdue),
    ])

    filename = f"escalations_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# =========================
# Workflow Routing Rules (Admin)
# =========================
@admin_bp.route("/workflow-routing")
@login_required
@perm_required("WORKFLOW_ROUTING_READ")
def workflow_routing_list():
    q = (request.args.get("q") or "").strip()
    sort = (request.args.get("sort") or "default").strip().lower()
    direction = (request.args.get("direction") or "asc").strip().lower()
    if direction not in ("asc", "desc"):
        direction = "asc"

    query = (
        WorkflowRoutingRule.query
        .outerjoin(RequestType, WorkflowRoutingRule.request_type_id == RequestType.id)
        .outerjoin(WorkflowTemplate, WorkflowRoutingRule.template_id == WorkflowTemplate.id)
        .outerjoin(Organization, WorkflowRoutingRule.organization_id == Organization.id)
        .outerjoin(Directorate, WorkflowRoutingRule.directorate_id == Directorate.id)
        .outerjoin(Department, WorkflowRoutingRule.department_id == Department.id)
        .outerjoin(OrgNode, WorkflowRoutingRule.org_node_id == OrgNode.id)
        .outerjoin(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            RequestType.code.ilike(like),
            RequestType.name_ar.ilike(like),
            RequestType.name_en.ilike(like),
            WorkflowTemplate.name.ilike(like),
            Organization.name_ar.ilike(like),
            Organization.name_en.ilike(like),
            Directorate.name_ar.ilike(like),
            Directorate.name_en.ilike(like),
            Department.name_ar.ilike(like),
            Department.name_en.ilike(like),
            OrgNode.name_ar.ilike(like),
            OrgNode.name_en.ilike(like),
            OrgNode.code.ilike(like),
            OrgNodeType.name_ar.ilike(like),
            OrgNodeType.name_en.ilike(like),
        ))

    sort_map = {
        "id": WorkflowRoutingRule.id,
        "request_type": func.coalesce(RequestType.name_ar, RequestType.name_en, RequestType.code, ""),
        "organization": func.coalesce(Organization.name_ar, Organization.name_en, ""),
        "directorate": func.coalesce(Directorate.name_ar, Directorate.name_en, ""),
        "department": func.coalesce(Department.name_ar, Department.name_en, ""),
        "org_node": func.coalesce(OrgNode.name_ar, OrgNode.name_en, OrgNode.code, ""),
        "template": func.coalesce(WorkflowTemplate.name, ""),
        "priority": WorkflowRoutingRule.priority,
        "is_active": WorkflowRoutingRule.is_active,
    }

    if sort in sort_map:
        sort_expr = sort_map[sort]
        order_expr = sort_expr.desc() if direction == "desc" else sort_expr.asc()
        order_by = [order_expr, WorkflowRoutingRule.id.desc()]
    else:
        sort = "default"
        order_by = [
            WorkflowRoutingRule.is_active.desc(),
            WorkflowRoutingRule.priority.asc(),
            WorkflowRoutingRule.id.desc(),
        ]

    rules = (
        query
        .order_by(*order_by)
        .all()
    )
    return render_template(
        "admin/workflow_routing/list.html",
        rules=rules,
        q=q,
        sort=sort,
        direction=direction,
    )


@admin_bp.route("/workflow-routing/org-node-tree")
@login_required
@perm_required("WORKFLOW_ROUTING_READ")
def workflow_routing_org_node_tree():
    """AJAX: return OrgNode picker tree for routing rules (mode=routes)."""
    mode = (request.args.get("mode") or "routes").strip().lower()
    sel = (request.args.get("selected") or "").strip()
    selected_id = int(sel) if sel.isdigit() else None
    tree = build_org_node_picker_tree(mode=mode)
    return render_template(
        "components/_org_node_picker_tree.html",
        tree=tree,
        selected_id=selected_id,
        mode=mode,
    )

@admin_bp.route("/workflow-routing/export.xlsx")
@login_required
@perm_required("WORKFLOW_ROUTING_READ")
def workflow_routing_export_excel():
    """Export workflow routing rules to .xlsx."""
    q = (request.args.get("q") or "").strip()

    query = (
        WorkflowRoutingRule.query
        .outerjoin(RequestType, WorkflowRoutingRule.request_type_id == RequestType.id)
        .outerjoin(WorkflowTemplate, WorkflowRoutingRule.template_id == WorkflowTemplate.id)
        .outerjoin(Organization, WorkflowRoutingRule.organization_id == Organization.id)
        .outerjoin(Directorate, WorkflowRoutingRule.directorate_id == Directorate.id)
        .outerjoin(Department, WorkflowRoutingRule.department_id == Department.id)
        .outerjoin(OrgNode, WorkflowRoutingRule.org_node_id == OrgNode.id)
        .outerjoin(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
    )

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            RequestType.code.ilike(like),
            RequestType.name_ar.ilike(like),
            RequestType.name_en.ilike(like),
            WorkflowTemplate.name.ilike(like),
            Organization.name_ar.ilike(like),
            Organization.name_en.ilike(like),
            Directorate.name_ar.ilike(like),
            Directorate.name_en.ilike(like),
            Department.name_ar.ilike(like),
            Department.name_en.ilike(like),
            OrgNode.name_ar.ilike(like),
            OrgNode.name_en.ilike(like),
            OrgNode.code.ilike(like),
            OrgNodeType.name_ar.ilike(like),
            OrgNodeType.name_en.ilike(like),
        ))

    rules = (
        query
        .order_by(
            WorkflowRoutingRule.is_active.desc(),
            WorkflowRoutingRule.priority.asc(),
            WorkflowRoutingRule.id.desc()
        )
        .all()
    )

    headers = [
        "ID",
        "RequestType Code",
        "RequestType AR",
        "RequestType EN",
        "Organization",
        "الإدارة",
        "الدائرة",
        "OrgNode",
        "MatchSubtree",
        "Template",
        "Priority",
        "Active",
    ]

    rows = []
    for r in rules:
        rows.append([
            r.id,
            r.request_type.code if r.request_type else None,
            r.request_type.name_ar if r.request_type else None,
            r.request_type.name_en if r.request_type else None,
            r.organization.name_ar if r.organization else None,
            r.directorate.name_ar if r.directorate else None,
            r.department.name_ar if r.department else None,
            (f"{(r.org_node.type.name_ar if r.org_node and r.org_node.type else r.org_node.type.code if r.org_node and r.org_node.type else '')} — {r.org_node.name_ar}".strip(" —") if r.org_node else None),
            ("Yes" if getattr(r, "match_subtree", False) else "No") if r.org_node_id is not None else None,
            r.template.name if r.template else None,
            r.priority,
            "Yes" if r.is_active else "No",
        ])

    xlsx = make_xlsx_bytes("RoutingRules", headers, rows)
    bio = BytesIO(xlsx)
    bio.seek(0)

    filename = f"routing_rules_{datetime.utcnow().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename)





@admin_bp.route("/workflow-routing/import-excel", methods=["POST"])
@login_required
@perm_required("WORKFLOW_ROUTING_UPDATE")
def workflow_routing_import_excel():
    """Import workflow routing rules from Excel.

    Modes:
      - safe: upsert by composite key (request_type + org + dir + dept)
      - replace: try delete-all then insert; if FK prevents deletion, soft-fallback by deactivating all then upsert

    Supported/tolerant columns:
      - request_type_code (or "RequestType Code") (required)
      - organization / organization_code (optional)
      - directorate / directorate_code (optional)
      - department / department_code (optional)
      - org_node / org_node_id / OrgNode (optional)
      - match_subtree / subtree / MatchSubtree (optional, default True when org_node provided)
      - template (name or id) (required)
      - priority (optional, default 100)
      - active / is_active (optional, default True)
    """
    mode = (request.form.get("mode") or "safe").strip().lower()
    file_storage = request.files.get("file")
    if not file_storage:
        flash("يرجى اختيار ملف Excel (.xlsx).", "danger")
        return redirect(url_for("admin.workflow_routing_list"))

    try:
        _title, rows, _headers = read_excel_rows(file_storage)
    except Exception as e:
        flash(f"تعذر قراءة ملف Excel: {e}", "danger")
        return redirect(url_for("admin.workflow_routing_list"))

    if not rows:
        flash("ملف Excel فارغ أو لا يحتوي صفوف بيانات.", "warning")
        return redirect(url_for("admin.workflow_routing_list"))

    def _clean_txt(v):
        s = to_str(v)
        return s.strip() if s else None

    def _resolve_obj(model, value, *, code_field: str = "code", name_fields=("name_ar", "name_en")):
        val = _clean_txt(value)
        if not val:
            return None
        # by id
        if val.isdigit():
            try:
                return model.query.get(int(val))
            except Exception:
                pass
        # by code
        if hasattr(model, code_field):
            col = getattr(model, code_field)
            obj = model.query.filter(col == val).first()
            if obj:
                return obj
            try:
                obj = model.query.filter(func.lower(col) == val.lower()).first()
                if obj:
                    return obj
            except Exception:
                pass
        # exact by names
        for nf in name_fields:
            if hasattr(model, nf):
                col = getattr(model, nf)
                obj = model.query.filter(col == val).first()
                if obj:
                    return obj
        # fallback ilike
        like = f"%{val}%"
        conds = []
        if hasattr(model, code_field):
            try:
                conds.append(getattr(model, code_field).ilike(like))
            except Exception:
                pass
        for nf in name_fields:
            if hasattr(model, nf):
                try:
                    conds.append(getattr(model, nf).ilike(like))
                except Exception:
                    pass
        if conds:
            obj = model.query.filter(or_(*conds)).first()
            if obj:
                return obj
        return None

    def _resolve_request_type(code_or_name):
        val = _clean_txt(code_or_name)
        if not val:
            return None
        code = val.upper()
        rt = RequestType.query.filter(RequestType.code == code).first()
        if rt:
            return rt
        # fallback by name
        rt = RequestType.query.filter(or_(RequestType.name_ar == val, RequestType.name_en == val)).first()
        if rt:
            return rt
        like = f"%{val}%"
        return RequestType.query.filter(or_(RequestType.code.ilike(like), RequestType.name_ar.ilike(like), RequestType.name_en.ilike(like))).first()

    def _resolve_template(name_or_id):
        val = _clean_txt(name_or_id)
        if not val:
            return None
        if val.isdigit():
            return WorkflowTemplate.query.get(int(val))
        # exact by name
        tpl = WorkflowTemplate.query.filter(WorkflowTemplate.name == val).first()
        if tpl:
            return tpl
        like = f"%{val}%"
        return WorkflowTemplate.query.filter(WorkflowTemplate.name.ilike(like)).first()

    def _parse_row(r):
        rt_val = pick(r, "request_type_code", "requesttypecode", "request type code", "RequestType Code", "rtype", "نوعالطلب", "نوع الطلب")
        org_val = pick(r, "organization", "org", "org_code", "organization_code", "organizationcode", "Organization", "منظمة")
        dir_val = pick(r, "directorate", "dir", "dir_code", "directorate_code", "directoratecode", "Directorate", "إدارة")
        dept_val = pick(r, "department", "dept", "dept_code", "department_code", "departmentcode", "Department", "دائرة")
        node_val = pick(r, "org_node", "orgnode", "org_node_id", "orgnodeid", "OrgNode", "Org Node", "node", "node_id", "عنصرهيكلي", "عنصر هيكلي", "عنصر")
        subtree_val = pick(r, "match_subtree", "subtree", "matchsubtree", "MatchSubtree", "Match Subtree", "شامل", "شامل الفروع")
        tpl_val = pick(r, "template", "template_id", "templateid", "Template", "workflow_template", "المسار")
        pr_val = pick(r, "priority", "Priority", "الأولوية")
        act_val = pick(r, "is_active", "active", "Active", "نشط", "فعال")

        rt = _resolve_request_type(rt_val)
        tpl = _resolve_template(tpl_val)
        org = _resolve_obj(Organization, org_val)
        direc = _resolve_obj(Directorate, dir_val)
        dept = _resolve_obj(Department, dept_val)
        node = _resolve_obj(OrgNode, node_val)

        match_subtree = to_bool(subtree_val, default=True)

        priority = to_int(pr_val, default=100) or 100
        is_active = to_bool(act_val, default=True)
        return rt, org, direc, dept, node, bool(match_subtree), tpl, int(priority), bool(is_active)

    skipped = 0
    created = 0
    updated = 0

    def _upsert_all():
        nonlocal created, updated, skipped
        created = updated = skipped = 0
        for rr in rows:
            rt, org, direc, dept, node, match_subtree, tpl, priority, is_active = _parse_row(rr)
            if not rt or not tpl:
                skipped += 1
                continue
            # If org_node is provided, we treat it as the unified hierarchy target
            # and clear legacy org/dir/dept constraints.
            node_id = node.id if node else None
            if node_id is not None:
                org_id = None
                dir_id = None
                dept_id = None
            else:
                org_id = org.id if org else None
                dir_id = direc.id if direc else None
                dept_id = dept.id if dept else None

            existing = WorkflowRoutingRule.query.filter_by(
                request_type_id=rt.id,
                organization_id=org_id,
                directorate_id=dir_id,
                department_id=dept_id,
                org_node_id=node_id,
            ).first()

            if existing:
                existing.template_id = tpl.id
                existing.priority = priority
                existing.is_active = is_active
                existing.org_node_id = node_id
                existing.match_subtree = bool(match_subtree) if node_id is not None else True
                updated += 1
            else:
                db.session.add(WorkflowRoutingRule(
                    request_type_id=rt.id,
                    organization_id=org_id,
                    directorate_id=dir_id,
                    department_id=dept_id,
                    org_node_id=node_id,
                    match_subtree=(bool(match_subtree) if node_id is not None else True),
                    template_id=tpl.id,
                    priority=priority,
                    is_active=is_active,
                ))
                created += 1
        return created, updated

    used_soft = False
    try:
        if mode == "replace":
            def _insert_fn():
                return _upsert_all()

            def _soft():
                WorkflowRoutingRule.query.update({WorkflowRoutingRule.is_active: False})
                db.session.flush()
                return _upsert_all()

            c, u, used_soft = replace_all(db.session, WorkflowRoutingRule.query, _insert_fn, soft_fallback=_soft)
            created, updated = c, u
        else:
            _upsert_all()

        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f"فشل استيراد قواعد التوجيه: {e}", "danger")
        return redirect(url_for("admin.workflow_routing_list"))

    msg = f"تم الاستيراد. تم إنشاء {created} وتحديث {updated}."
    if skipped:
        msg += f" (تم تخطي {skipped} صف/صفوف لغياب نوع الطلب أو المسار أو لعدم التطابق)"
    if mode == "replace" and used_soft:
        msg += " — تم استخدام Soft Replace (تعطيل الكل) بسبب قيود في قاعدة البيانات."
    flash(msg, "success")
    return redirect(url_for("admin.workflow_routing_list"))

@admin_bp.route("/workflow-routing/new", methods=["GET", "POST"])
@login_required
@perm_required("WORKFLOW_ROUTING_CREATE")
def workflow_routing_new():
    r = WorkflowRoutingRule()
    return _workflow_routing_form(r, is_new=True)


@admin_bp.route("/workflow-routing/<int:rule_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("WORKFLOW_ROUTING_UPDATE")
def workflow_routing_edit(rule_id):
    r = WorkflowRoutingRule.query.get_or_404(rule_id)
    return _workflow_routing_form(r, is_new=False)


def _workflow_routing_form(r: WorkflowRoutingRule, is_new: bool):
    request_types = RequestType.query.order_by(RequestType.name_ar.asc()).all()
    orgs = Organization.query.order_by(Organization.name_ar.asc()).all()
    dirs = Directorate.query.order_by(Directorate.name_ar.asc()).all()
    depts = Department.query.order_by(Department.name_ar.asc()).all()
    templates = WorkflowTemplate.query.order_by(WorkflowTemplate.name.asc()).all()

    if request.method == "POST":
        # required
        rt_id = (request.form.get("request_type_id") or "").strip()
        template_id = (request.form.get("template_id") or "").strip()

        if not rt_id.isdigit():
            flash("اختر نوع طلب.", "danger")
            return redirect(request.url)
        if not template_id.isdigit():
            flash("اختر مسار (Template).", "danger")
            return redirect(request.url)

        r.request_type_id = int(rt_id)
        r.template_id = int(template_id)

        # optional hierarchy
        org_id = (request.form.get("organization_id") or "").strip()
        dir_id = (request.form.get("directorate_id") or "").strip()
        dept_id = (request.form.get("department_id") or "").strip()
        node_id = (request.form.get("org_node_id") or "").strip()
        match_subtree = (request.form.get("match_subtree") == "1")

        r.org_node_id = int(node_id) if node_id.isdigit() else None
        r.match_subtree = bool(match_subtree) if r.org_node_id is not None else True

        # If unified OrgNode is set, clear legacy org/dir/dept constraints
        if r.org_node_id is not None:
            r.organization_id = None
            r.directorate_id = None
            r.department_id = None
        else:
            r.organization_id = int(org_id) if org_id.isdigit() else None
            r.directorate_id = int(dir_id) if dir_id.isdigit() else None
            r.department_id = int(dept_id) if dept_id.isdigit() else None

        # priority + active
        try:
            r.priority = int((request.form.get("priority") or "100").strip())
        except Exception:
            r.priority = 100

        r.is_active = (request.form.get("is_active") == "1")

        # validation: don't allow dept without dir, or dir without org (legacy mode only)
        if r.org_node_id is None:
            if r.department_id and not r.directorate_id:
                flash("لا يمكن تحديد دائرة بدون تحديد إدارة.", "danger")
                return redirect(request.url)
            if r.directorate_id and not r.organization_id:
                flash("لا يمكن تحديد إدارة بدون تحديد منظمة.", "danger")
                return redirect(request.url)

        if is_new:
            db.session.add(r)

        db.session.commit()
        flash("تم حفظ قاعدة التوجيه.", "success")
        return redirect(url_for("admin.workflow_routing_list"))

    return render_template(
        "admin/workflow_routing/form.html",
        r=r,
        is_new=is_new,
        request_types=request_types,
        orgs=orgs,
        dirs=dirs,
        depts=depts,
        templates=templates
    )


@admin_bp.route("/workflow-routing/<int:rule_id>/delete", methods=["POST"])
@login_required
@perm_required("WORKFLOW_ROUTING_DELETE")
def workflow_routing_delete(rule_id):
    r = WorkflowRoutingRule.query.get_or_404(rule_id)
    db.session.delete(r)
    db.session.commit()
    flash("تم حذف قاعدة التوجيه.", "warning")
    return redirect(url_for("admin.workflow_routing_list"))

# =========================
# Backup & Restore (Full System)
# =========================

BACKUP_FORMAT_VERSION = 2

PERMISSION_BACKUP_MODELS = [
    ("roles", Role),
    ("role_permissions", RolePermission),
    ("user_permissions", UserPermission),
    ("file_permissions", FilePermission),
    ("store_file_permissions", StoreFilePermission),
    ("inventory_warehouse_permissions", InvWarehousePermission),
    ("portal_permission_presets", PortalPermissionPreset),
]


def _get_db_path() -> str:
    return os.path.join(current_app.instance_path, "workflow.db")


def _get_project_root() -> str:
    # app.py lives in project root, and Flask root_path points to that directory
    return current_app.root_path


def _get_archive_storage_dir() -> str:
    return os.path.join(_get_project_root(), "storage", "archive")


def _get_portal_uploads_dir() -> str:
    """Portal/admin uploads live under instance/uploads/* (correspondence, store, HR, etc.)."""
    return os.path.join(current_app.instance_path, "uploads")



def _get_static_uploads_dir() -> str:
    """Static uploads live under static/uploads (e.g., user avatars/photos)."""
    return os.path.join(_get_project_root(), "static", "uploads")


def _sha256_file(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _sqlite_identifier(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _inspect_sqlite_database(db_path: str) -> dict:
    """Validate a SQLite file and return every table, column and row count."""
    if not os.path.isfile(db_path) or os.path.getsize(db_path) < 100:
        raise ValueError("SQLite database file is missing or empty")

    connection = sqlite3.connect(f"file:{os.path.abspath(db_path)}?mode=ro", uri=True)
    try:
        quick_check = [str(row[0]) for row in connection.execute("PRAGMA quick_check").fetchall()]
        if quick_check != ["ok"]:
            raise ValueError("SQLite quick_check failed: " + "; ".join(quick_check[:5]))

        table_rows = connection.execute(
            "SELECT name, sql FROM sqlite_master "
            "WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
        ).fetchall()
        tables = {}
        total_rows = 0
        for table_name, create_sql in table_rows:
            quoted_name = _sqlite_identifier(table_name)
            row_count = int(connection.execute(f"SELECT COUNT(*) FROM {quoted_name}").fetchone()[0])
            total_rows += row_count
            columns = [
                {
                    "name": column[1],
                    "type": column[2],
                    "not_null": bool(column[3]),
                    "default": column[4],
                    "primary_key": int(column[5] or 0),
                }
                for column in connection.execute(f"PRAGMA table_info({quoted_name})").fetchall()
            ]
            tables[table_name] = {
                "row_count": row_count,
                "columns": columns,
                "create_sql": create_sql,
            }

        if "users" not in tables:
            raise ValueError("Backup database does not contain the users table")

        return {
            "table_count": len(tables),
            "total_rows": total_rows,
            "tables": tables,
        }
    finally:
        connection.close()


def _database_module_summary(database_manifest: dict) -> dict:
    tables = database_manifest.get("tables") or {}

    def summarize(predicate) -> dict:
        selected = [payload for name, payload in tables.items() if predicate(name.lower())]
        return {
            "table_count": len(selected),
            "row_count": sum(int(payload.get("row_count") or 0) for payload in selected),
        }

    return {
        "support_tickets": summarize(lambda name: name.startswith("trouble_")),
        "transport": summarize(lambda name: name.startswith("transport_")),
        "inventory": summarize(lambda name: name.startswith("inv_") or name.startswith("store_")),
        "permissions": summarize(lambda name: "permission" in name or name in {"roles", "portal_permission_preset"}),
        "workflow": summarize(lambda name: name.startswith("workflow_") or name in {"requests", "approvals", "audit_logs", "notifications", "messages"}),
        "human_resources": summarize(lambda name: name.startswith("hr_") or name.startswith("employee_") or name.startswith("attendance_")),
    }


def _validate_sqlite_snapshot(snapshot_path: str, expected_manifest: dict | None = None) -> dict:
    actual = _inspect_sqlite_database(snapshot_path)
    expected_manifest = expected_manifest or {}

    expected_sha256 = (expected_manifest.get("sha256") or "").strip().lower()
    if expected_sha256 and _sha256_file(snapshot_path).lower() != expected_sha256:
        raise ValueError("Database checksum does not match backup manifest")

    expected_tables = expected_manifest.get("tables") or {}
    for table_name, expected in expected_tables.items():
        actual_table = actual["tables"].get(table_name)
        if actual_table is None:
            raise ValueError(f"Database table is missing: {table_name}")
        if int(actual_table.get("row_count") or 0) != int(expected.get("row_count") or 0):
            raise ValueError(f"Database row count mismatch: {table_name}")
        expected_columns = [column.get("name") for column in (expected.get("columns") or [])]
        actual_columns = [column.get("name") for column in (actual_table.get("columns") or [])]
        if expected_columns and actual_columns != expected_columns:
            raise ValueError(f"Database schema mismatch: {table_name}")

    return actual


def _persistent_path_is_excluded(scope: str, relative_path: str) -> bool:
    normalized = relative_path.replace("\\", "/").strip("/")
    parts = [part for part in normalized.split("/") if part]
    if not parts or any("_before_restore_" in part.lower() for part in parts):
        return True

    top = parts[0].lower()
    if scope == "instance":
        if top in {"backups", "tmp"} or top.startswith("uploads_before_restore_"):
            return True
        if len(parts) == 1 and (
            top == "automatic_backup_state.json"
            or top.endswith(".log")
            or top.endswith(".db")
            or top.endswith(".db-wal")
            or top.endswith(".db-shm")
        ):
            return True
    return False


def _iter_persistent_backup_files():
    source_specs = [
        ("instance", current_app.instance_path, "instance"),
        ("storage", os.path.join(_get_project_root(), "storage"), "storage"),
        ("static/uploads", _get_static_uploads_dir(), "static"),
    ]

    for archive_prefix, source_root, scope in source_specs:
        if not os.path.isdir(source_root):
            continue
        for root, directories, filenames in os.walk(source_root):
            relative_root = os.path.relpath(root, source_root)
            directories[:] = [
                directory
                for directory in directories
                if not _persistent_path_is_excluded(
                    scope,
                    directory if relative_root == "." else os.path.join(relative_root, directory),
                )
            ]
            for filename in filenames:
                full_path = os.path.join(root, filename)
                relative_path = os.path.relpath(full_path, source_root)
                if _persistent_path_is_excluded(scope, relative_path):
                    continue
                archive_path = f"{archive_prefix}/{relative_path.replace(os.sep, '/')}"
                if archive_prefix == "static/uploads":
                    restore_root = "static/uploads"
                else:
                    top_level = relative_path.replace("\\", "/").split("/", 1)[0]
                    restore_root = f"{archive_prefix}/{top_level}"
                yield full_path, archive_path, restore_root


def _discover_persistent_restore_roots() -> list[str]:
    roots = []
    for archive_prefix, source_root, scope in [
        ("instance", current_app.instance_path, "instance"),
        ("storage", os.path.join(_get_project_root(), "storage"), "storage"),
    ]:
        if not os.path.isdir(source_root):
            continue
        for name in os.listdir(source_root):
            source_path = os.path.join(source_root, name)
            if os.path.isdir(source_path) and not _persistent_path_is_excluded(scope, name):
                roots.append(f"{archive_prefix}/{name}")
    if os.path.isdir(_get_static_uploads_dir()):
        roots.append("static/uploads")
    return sorted(set(roots))


def _validate_files_manifest(extract_dir: str, files_manifest: dict) -> None:
    for entry in files_manifest.get("files") or []:
        archive_path = (entry.get("path") or "").replace("\\", "/").strip("/")
        if not archive_path:
            raise ValueError("Backup file manifest contains an empty path")
        extracted_path = os.path.abspath(os.path.join(extract_dir, *archive_path.split("/")))
        if os.path.commonpath([os.path.abspath(extract_dir), extracted_path]) != os.path.abspath(extract_dir):
            raise ValueError("Backup file manifest contains an unsafe path")
        if not os.path.isfile(extracted_path):
            raise ValueError(f"Backup file is missing: {archive_path}")
        if os.path.getsize(extracted_path) != int(entry.get("size") or 0):
            raise ValueError(f"Backup file size mismatch: {archive_path}")
        expected_hash = (entry.get("sha256") or "").strip().lower()
        if expected_hash and _sha256_file(extracted_path).lower() != expected_hash:
            raise ValueError(f"Backup file checksum mismatch: {archive_path}")


def _write_file_to_zip(archive: zipfile.ZipFile, source_path: str, archive_path: str) -> dict:
    digest = hashlib.sha256()
    total_size = 0
    with open(source_path, "rb") as source, archive.open(archive_path, "w") as destination:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            destination.write(chunk)
            digest.update(chunk)
            total_size += len(chunk)
    return {
        "path": archive_path,
        "size": total_size,
        "sha256": digest.hexdigest(),
    }


def _backup_json_value(value):
    if isinstance(value, datetime):
        return value.isoformat()
    return value


def _build_permissions_export() -> dict:
    """Export permission-related tables explicitly in addition to the DB snapshot."""
    data = {
        "version": 1,
        "created_at_utc": datetime.utcnow().isoformat() + "Z",
        "tables": {},
    }

    for export_name, model in PERMISSION_BACKUP_MODELS:
        columns = [c.name for c in model.__table__.columns]
        rows = []
        query = model.query
        if "id" in columns:
            query = query.order_by(model.id.asc())

        for row in query.all():
            rows.append({
                column: _backup_json_value(getattr(row, column))
                for column in columns
            })

        data["tables"][export_name] = {
            "table_name": model.__tablename__,
            "columns": columns,
            "rows": rows,
            "count": len(rows),
        }

    return data


def _permission_backup_counts(permissions_export: dict) -> dict:
    return {
        table_name: payload.get("count", 0)
        for table_name, payload in permissions_export.get("tables", {}).items()
    }


def _get_backups_dir() -> str:
    d = os.path.join(current_app.instance_path, "backups")
    os.makedirs(d, exist_ok=True)
    return d


def _get_runtime_tmp_dir() -> str:
    d = os.path.join(current_app.instance_path, "tmp")
    os.makedirs(d, exist_ok=True)
    return d


def _make_runtime_tmp_subdir(prefix: str) -> str:
    base = _get_runtime_tmp_dir()
    for _ in range(5):
        path = os.path.join(base, f"{prefix}_{uuid.uuid4().hex}")
        try:
            os.makedirs(path, exist_ok=False)
            return path
        except FileExistsError:
            continue
    path = os.path.join(base, f"{prefix}_{datetime.utcnow().strftime('%H%M%S%f')}")
    os.makedirs(path, exist_ok=False)
    return path


class _RuntimeTempDir:
    def __init__(self, path: str):
        self.name = path

    def cleanup(self) -> None:
        shutil.rmtree(self.name, ignore_errors=True)


def _make_runtime_tempdir(prefix: str) -> _RuntimeTempDir:
    return _RuntimeTempDir(_make_runtime_tmp_subdir(prefix))


def _make_restore_tempdir(prefix: str) -> _RuntimeTempDir:
    """Use the OS temp root so extracted backup paths stay below Windows MAX_PATH."""
    safe_prefix = "".join(ch for ch in (prefix or "restore") if ch.isalnum() or ch in {"-", "_"})
    return _RuntimeTempDir(tempfile.mkdtemp(prefix=f"{safe_prefix[:40]}_"))


def _safe_extract_backup_zip(archive: zipfile.ZipFile, extract_dir: str) -> None:
    """Reject traversal paths before extracting an uploaded backup archive."""
    root = os.path.abspath(extract_dir)
    for info in archive.infolist():
        normalized = (info.filename or "").replace("\\", "/")
        parts = [part for part in normalized.split("/") if part not in {"", "."}]
        if (
            not normalized
            or normalized.startswith("/")
            or "\x00" in normalized
            or any(part == ".." for part in parts)
            or (parts and ":" in parts[0])
        ):
            raise zipfile.BadZipFile(f"Unsafe backup member: {info.filename!r}")

        target = os.path.abspath(os.path.join(root, *parts))
        if os.path.commonpath([root, target]) != root:
            raise zipfile.BadZipFile(f"Unsafe backup member: {info.filename!r}")

    archive.extractall(root)


def _create_sqlite_snapshot(src_db_path: str, snapshot_path: str) -> None:
    """Create a consistent snapshot of a SQLite DB using the sqlite3 backup API."""
    os.makedirs(os.path.dirname(snapshot_path), exist_ok=True)

    if not os.path.exists(src_db_path):
        # If DB doesn't exist yet, create an empty snapshot.
        sqlite3.connect(snapshot_path).close()
        return

    src = sqlite3.connect(src_db_path)
    try:
        try:
            # Integrate WAL if enabled
            src.execute("PRAGMA wal_checkpoint(FULL);")
        except Exception:
            pass

        dst = sqlite3.connect(snapshot_path)
        try:
            src.backup(dst)
            dst.commit()
        finally:
            dst.close()
    finally:
        src.close()


def _build_backup_zip() -> str:
    """Build a verified full-data ZIP with every DB table and persistent file."""
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    backups_dir = _make_runtime_tmp_subdir(f"workflow_backup_{ts}")

    zip_path = os.path.join(backups_dir, f"workflow_backup_{ts}.zip")
    tmp_db_path = os.path.join(backups_dir, f"workflow_snapshot_{ts}.db")

    # DB snapshot
    _create_sqlite_snapshot(_get_db_path(), tmp_db_path)
    database_manifest = _inspect_sqlite_database(tmp_db_path)
    database_manifest["sha256"] = _sha256_file(tmp_db_path)
    database_manifest["modules"] = _database_module_summary(database_manifest)
    permissions_export = _build_permissions_export()
    persistent_files = list(_iter_persistent_backup_files())
    restore_roots = sorted(
        {restore_root for _source, _archive, restore_root in persistent_files}
        | set(_discover_persistent_restore_roots())
    )
    files_manifest = {
        "version": 1,
        "restore_roots": restore_roots,
        "files": [],
        "file_count": 0,
        "total_bytes": 0,
    }

    with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as z:
        z.writestr(
            "permissions/permissions_export.json",
            json.dumps(permissions_export, ensure_ascii=False, indent=2)
        )
        z.write(tmp_db_path, "db/workflow.db")
        for restore_root in restore_roots:
            z.writestr(restore_root.rstrip("/") + "/", b"")
        for source_path, archive_path, _restore_root in persistent_files:
            files_manifest["files"].append(_write_file_to_zip(z, source_path, archive_path))

        files_manifest["file_count"] = len(files_manifest["files"])
        files_manifest["total_bytes"] = sum(entry["size"] for entry in files_manifest["files"])
        z.writestr("database/database_manifest.json", json.dumps(database_manifest, ensure_ascii=False, indent=2))
        z.writestr("files/files_manifest.json", json.dumps(files_manifest, ensure_ascii=False, indent=2))

        meta = {
            "format_version": BACKUP_FORMAT_VERSION,
            "created_at_utc": datetime.utcnow().isoformat() + "Z",
            "project": "Workflow-PNCECS",
            "includes": [
                "all SQLite tables, rows, indexes and schema",
                "all persistent instance data",
                "all active storage data",
                "all static uploads",
                "explicit permissions export",
            ],
            "database": {
                "table_count": database_manifest["table_count"],
                "total_rows": database_manifest["total_rows"],
                "sha256": database_manifest["sha256"],
                "modules": database_manifest["modules"],
            },
            "files": {
                "file_count": files_manifest["file_count"],
                "total_bytes": files_manifest["total_bytes"],
                "restore_roots": restore_roots,
            },
            "permission_counts": _permission_backup_counts(permissions_export),
        }
        z.writestr("backup_meta.json", json.dumps(meta, ensure_ascii=False, indent=2))
    # cleanup temp snapshot
    try:
        os.remove(tmp_db_path)
    except Exception:
        pass

    return zip_path


def _restore_sqlite_from_snapshot(snapshot_db: str, dest_db: str) -> None:
    """Restore SQLite DB content from snapshot into the destination DB file."""
    os.makedirs(os.path.dirname(dest_db), exist_ok=True)

    # Make SQLAlchemy release connections before overwriting content
    try:
        db.session.remove()
    except Exception:
        pass
    try:
        db.engine.dispose()
    except Exception:
        pass

    src = sqlite3.connect(snapshot_db)
    try:
        dst = sqlite3.connect(dest_db)
        try:
            # Replace destination content
            src.backup(dst)
            dst.commit()
        finally:
            dst.close()
    finally:
        src.close()


def _restore_json_value(column, value):
    if value is None:
        return None

    try:
        if column.type.python_type is datetime and isinstance(value, str):
            return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except Exception:
        pass

    return value


def _restore_permissions_export(extract_dir: str) -> bool:
    """Restore explicit permission export if the backup contains it."""
    export_path = os.path.join(extract_dir, "permissions", "permissions_export.json")
    if not os.path.exists(export_path):
        return False

    with open(export_path, "r", encoding="utf-8") as f:
        permissions_export = json.load(f)

    tables = permissions_export.get("tables") or {}

    try:
        for _export_name, model in reversed(PERMISSION_BACKUP_MODELS):
            db.session.query(model).delete(synchronize_session=False)

        db.session.flush()

        for export_name, model in PERMISSION_BACKUP_MODELS:
            payload = tables.get(export_name) or {}
            columns = {column.name: column for column in model.__table__.columns}

            for item in payload.get("rows") or []:
                values = {
                    name: _restore_json_value(column, item[name])
                    for name, column in columns.items()
                    if name in item
                }
                db.session.add(model(**values))

        db.session.commit()
        return True
    except Exception:
        db.session.rollback()
        raise


def _copy_tree(src_dir: str, dst_dir: str) -> None:
    os.makedirs(dst_dir, exist_ok=True)
    for root, dirs, files in os.walk(src_dir):
        rel_root = os.path.relpath(root, src_dir)
        target_root = dst_dir if rel_root == "." else os.path.join(dst_dir, rel_root)
        os.makedirs(target_root, exist_ok=True)

        for d in dirs:
            os.makedirs(os.path.join(target_root, d), exist_ok=True)

        for f in files:
            s = os.path.join(root, f)
            t = os.path.join(target_root, f)
            shutil.copy2(s, t)


def _read_backup_json(extract_dir: str, relative_path: str) -> dict:
    path = os.path.join(extract_dir, *relative_path.split("/"))
    if not os.path.isfile(path):
        return {}
    with open(path, "r", encoding="utf-8") as source:
        payload = json.load(source)
    if not isinstance(payload, dict):
        raise ValueError(f"Backup manifest is invalid: {relative_path}")
    return payload


def _restore_destination_for_root(archive_root: str) -> str:
    normalized = (archive_root or "").replace("\\", "/").strip("/")
    parts = [part for part in normalized.split("/") if part]
    if any(part in {".", ".."} or ":" in part for part in parts):
        raise ValueError(f"Unsafe restore root: {archive_root}")

    if len(parts) == 2 and parts[0] == "instance":
        return os.path.join(current_app.instance_path, parts[1])
    if len(parts) == 2 and parts[0] == "storage":
        return os.path.join(_get_project_root(), "storage", parts[1])
    if parts == ["static", "uploads"]:
        return _get_static_uploads_dir()
    raise ValueError(f"Unsupported restore root: {archive_root}")


def _replace_persistent_path(source_path: str, destination_path: str, timestamp: str) -> dict:
    os.makedirs(os.path.dirname(destination_path), exist_ok=True)
    safety_path = f"{destination_path}_before_restore_{timestamp}_{uuid.uuid4().hex[:8]}"
    had_original = os.path.exists(destination_path)

    if had_original:
        shutil.move(destination_path, safety_path)

    try:
        if os.path.isdir(source_path):
            shutil.copytree(source_path, destination_path)
        else:
            shutil.copy2(source_path, destination_path)
    except Exception:
        if os.path.isdir(destination_path):
            shutil.rmtree(destination_path, ignore_errors=True)
        elif os.path.exists(destination_path):
            os.remove(destination_path)
        if had_original and os.path.exists(safety_path):
            shutil.move(safety_path, destination_path)
        raise

    return {
        "destination": destination_path,
        "safety": safety_path,
        "had_original": had_original,
    }


def _rollback_persistent_replacements(replacements: list[dict]) -> None:
    for replacement in reversed(replacements):
        destination = replacement["destination"]
        safety = replacement["safety"]
        if os.path.isdir(destination):
            shutil.rmtree(destination, ignore_errors=True)
        elif os.path.exists(destination):
            os.remove(destination)
        if replacement.get("had_original") and os.path.exists(safety):
            shutil.move(safety, destination)


def _restore_persistent_files(extract_dir: str, files_manifest: dict, timestamp: str) -> list[dict]:
    restore_roots = list(files_manifest.get("restore_roots") or [])
    if not restore_roots:
        restore_roots = [
            root
            for root in ("storage/archive", "instance/uploads", "static/uploads")
            if os.path.exists(os.path.join(extract_dir, *root.split("/")))
        ]

    replacements = []
    try:
        for archive_root in sorted(set(restore_roots)):
            source_path = os.path.join(extract_dir, *archive_root.replace("\\", "/").split("/"))
            if not os.path.exists(source_path):
                raise ValueError(f"Backup restore root is missing: {archive_root}")
            destination_path = _restore_destination_for_root(archive_root)
            replacements.append(_replace_persistent_path(source_path, destination_path, timestamp))
        return replacements
    except Exception:
        _rollback_persistent_replacements(replacements)
        raise


def _cleanup_restore_work_files(uploaded_zip: str, extract_tmp: _RuntimeTempDir) -> None:
    try:
        extract_tmp.cleanup()
    except Exception:
        pass
    try:
        if os.path.exists(uploaded_zip):
            os.remove(uploaded_zip)
    except Exception:
        pass


@admin_bp.route("/backup", methods=["GET"])
@login_required
@roles_required("ADMIN", "SUPER_ADMIN")
def backup_page():
    # show last few backups if exist
    backups_dir = _get_backups_dir()
    backups = []
    try:
        for fn in sorted(os.listdir(backups_dir), reverse=True):
            if fn.lower().endswith(".zip") and fn.startswith("workflow_backup_"):
                backups.append(fn)
            if len(backups) >= 10:
                break
    except Exception:
        backups = []

    backup_summary = {}
    try:
        backup_summary = _inspect_sqlite_database(_get_db_path())
        backup_summary["modules"] = _database_module_summary(backup_summary)
    except Exception:
        current_app.logger.exception("Could not inspect current database for backup page")

    return render_template("admin/backup.html", backups=backups, backup_summary=backup_summary)


@admin_bp.route("/backup/download", methods=["GET"])
@login_required
@roles_required("ADMIN", "SUPER_ADMIN")
def backup_download():
    zip_path = _build_backup_zip()
    filename = os.path.basename(zip_path)
    response = send_file(
        zip_path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/zip"
    )
    response.call_on_close(lambda: shutil.rmtree(os.path.dirname(zip_path), ignore_errors=True))
    return response


@admin_bp.route("/backup/download/<path:fname>", methods=["GET"])
@login_required
@roles_required("ADMIN", "SUPER_ADMIN")
def backup_download_existing(fname):
    # allow downloading previously generated backups (from backups dir only)
    backups_dir = _get_backups_dir()
    safe_name = os.path.basename(fname)
    path = os.path.join(backups_dir, safe_name)

    if not os.path.exists(path) or not safe_name.lower().endswith(".zip"):
        flash("ملف النسخة الاحتياطية غير موجود.", "danger")
        return redirect(url_for("admin.backup_page"))

    return send_file(
        path,
        as_attachment=True,
        download_name=safe_name,
        mimetype="application/zip"
    )


@admin_bp.route("/backup/restore", methods=["POST"])
@login_required
@roles_required("ADMIN", "SUPER_ADMIN")
def backup_restore():
    confirm = request.form.get("confirm_restore")
    if confirm != "1":
        flash("يجب تأكيد الاستيراد قبل المتابعة.", "danger")
        return redirect(url_for("admin.backup_page"))

    up = request.files.get("backup_file")
    if not up or up.filename == "":
        flash("الرجاء اختيار ملف Backup بصيغة ZIP.", "danger")
        return redirect(url_for("admin.backup_page"))

    if not up.filename.lower().endswith(".zip"):
        flash("صيغة الملف غير مدعومة. الرجاء رفع ملف ZIP.", "danger")
        return redirect(url_for("admin.backup_page"))

    backups_dir = _get_backups_dir()
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    uploaded_zip = os.path.join(_get_runtime_tmp_dir(), f"workflow_uploaded_restore_{ts}.zip")
    # The project can live under a long Windows path. Extracting under instance/tmp
    # then exceeds the legacy 260-character limit for nested backup members.
    extract_tmp = _make_restore_tempdir(f"wf_restore_{ts}")
    extract_dir = extract_tmp.name

    try:
        up.save(uploaded_zip)
        try:
            up.close()
        except Exception:
            pass

        with zipfile.ZipFile(uploaded_zip, "r") as z:
            damaged_member = z.testzip()
            if damaged_member:
                raise zipfile.BadZipFile(f"Damaged backup member: {damaged_member}")
            _safe_extract_backup_zip(z, extract_dir)
    except Exception:
        current_app.logger.exception("Backup archive validation or extraction failed")
        _cleanup_restore_work_files(uploaded_zip, extract_tmp)
        flash("ملف النسخة الاحتياطية غير صالح أو تالف.", "danger")
        return redirect(url_for("admin.backup_page"))

    # Locate DB inside backup
    db_candidates = [
        os.path.join(extract_dir, "db", "workflow.db"),
        os.path.join(extract_dir, "instance", "workflow.db"),
        os.path.join(extract_dir, "workflow.db"),
    ]
    snap_db = next((p for p in db_candidates if os.path.exists(p)), None)

    if not snap_db:
        _cleanup_restore_work_files(uploaded_zip, extract_tmp)
        flash("النسخة الاحتياطية لا تحتوي على قاعدة بيانات (workflow.db).", "danger")
        return redirect(url_for("admin.backup_page"))

    try:
        backup_meta = _read_backup_json(extract_dir, "backup_meta.json")
        if backup_meta.get("project") not in {None, "", "Workflow-PNCECS"}:
            raise ValueError("Backup belongs to a different project")
        format_version = int(backup_meta.get("format_version") or 1)
        database_manifest = _read_backup_json(extract_dir, "database/database_manifest.json")
        files_manifest = _read_backup_json(extract_dir, "files/files_manifest.json")
        if format_version >= BACKUP_FORMAT_VERSION and (not database_manifest or not files_manifest):
            raise ValueError("Full backup manifests are missing")
        validated_database = _validate_sqlite_snapshot(snap_db, database_manifest)
        if files_manifest:
            _validate_files_manifest(extract_dir, files_manifest)
    except Exception:
        current_app.logger.exception("Backup content validation failed before restore")
        _cleanup_restore_work_files(uploaded_zip, extract_tmp)
        flash("لم يتم الاستيراد: فشل فحص سلامة قاعدة البيانات أو الملفات داخل النسخة.", "danger")
        return redirect(url_for("admin.backup_page"))

    safety_db_path = os.path.join(backups_dir, f"workflow_before_restore_{ts}.db")
    try:
        _create_sqlite_snapshot(_get_db_path(), safety_db_path)
        _restore_sqlite_from_snapshot(snap_db, _get_db_path())
        _validate_sqlite_snapshot(_get_db_path(), database_manifest)
        db.create_all()

        if format_version < BACKUP_FORMAT_VERSION:
            _restore_permissions_export(extract_dir)

        replacements = _restore_persistent_files(extract_dir, files_manifest, ts)
        _validate_sqlite_snapshot(_get_db_path(), database_manifest)
    except Exception:
        current_app.logger.exception("Backup restore failed; rolling back current data")
        try:
            if "replacements" in locals():
                _rollback_persistent_replacements(replacements)
        except Exception:
            current_app.logger.exception("Could not roll back restored persistent files")
        try:
            if os.path.isfile(safety_db_path):
                _restore_sqlite_from_snapshot(safety_db_path, _get_db_path())
                db.create_all()
        except Exception:
            current_app.logger.exception("Could not roll back database after restore failure")
        _cleanup_restore_work_files(uploaded_zip, extract_tmp)
        flash("فشل الاستيراد وتمت إعادة البيانات السابقة تلقائيًا. لم تُعتمد النسخة المرفوعة.", "danger")
        return redirect(url_for("admin.backup_page"))

    # After restore, force re-login
    try:
        logout_user()
    except Exception:
        pass

    restored_file_count = int(files_manifest.get("file_count") or 0)
    flash(
        f"✅ تم فحص واستيراد النسخة بنجاح: {validated_database['table_count']} جدول، "
        f"{validated_database['total_rows']} سجل، {restored_file_count} ملف. يرجى تسجيل الدخول من جديد.",
        "success",
    )
    flash(f"تم حفظ قاعدة البيانات السابقة احتياطياً في: {os.path.basename(safety_db_path)}", "info")

    _cleanup_restore_work_files(uploaded_zip, extract_tmp)

    return redirect(url_for("login"))
