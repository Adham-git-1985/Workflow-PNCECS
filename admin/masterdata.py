# admin/masterdata.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, abort, send_file
from io import BytesIO
from datetime import datetime

from flask_login import login_required
from sqlalchemy import or_
from sqlalchemy.orm import selectinload
from extensions import db
from utils.perms import perm_required
from utils.excel import make_xlsx_bytes, make_xlsx_bytes_multi
from utils.importer import read_excel_rows, pick, to_str, to_int, to_bool, upsert_by_code, replace_all
from utils.org_dynamic import ensure_dynamic_org_seed, sync_legacy_now

from models import Organization, Directorate, Unit, Department, Section, Division, Role, User, UserPermission, RequestType, WorkflowRoutingRule, WorkflowRequest, Committee, CommitteeAssignee, WorkflowTemplateStep, WorkflowTemplateParallelAssignee, WorkflowInstanceStep, OrgNodeType, OrgNode, SystemSetting

masterdata_bp = Blueprint("masterdata", __name__, url_prefix="/admin/masterdata")

PERM_MODULES = [
    ("MASTERDATA", "Ø§Ù„Ø¨ÙŠØ§Ù†Ø§Øª Ø§Ù„Ø£Ø³Ø§Ø³ÙŠØ© (Ù…Ù†Ø¸Ù…Ø§Øª/Ø¥Ø¯Ø§Ø±Ø§Øª/Ø¯ÙˆØ§Ø¦Ø±)", "MASTERDATA_MANAGE"),
    ("WORKFLOW_TEMPLATES", "Ù…Ø³Ø§Ø±Ø§Øª Ø§Ù„Ø¹Ù…Ù„ (Templates)", "WORKFLOW_TEMPLATES_MANAGE"),
    ("REQUEST_TYPES", "Ø£Ù†ÙˆØ§Ø¹ Ø§Ù„Ø·Ù„Ø¨Ø§Øª (Request Types)", "REQUEST_TYPES_MANAGE"),
    ("WORKFLOW_ROUTING", "Ù‚ÙˆØ§Ø¹Ø¯ Ø§Ù„ØªÙˆØ¬ÙŠÙ‡ (Routing Rules)", "WORKFLOW_ROUTING_MANAGE"),
    ("USER_PERMISSIONS", "Ø¥Ø¯Ø§Ø±Ø© ØµÙ„Ø§Ø­ÙŠØ§Øª Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù…ÙŠÙ†", "USER_PERMISSIONS_MANAGE"),
]

PERM_EXTRA_KEYS = [
    ("VIEW_DASHBOARD", "Ø±Ø¤ÙŠØ© Ù„ÙˆØ­Ø© Dashboard"),
    ("VIEW_ESCALATIONS", "Ø±Ø¤ÙŠØ© ØµÙØ­Ø© ğŸš¨ Escalations"),
    ("HR_SYSTEM_EVALUATION_VIEW", "Ø¹Ø±Ø¶ Ø§Ù„ØªÙ‚ÙŠÙŠÙ… Ø§Ù„Ù†Ø¸Ø§Ù…ÙŠ Ù„Ù„Ù…ÙˆØ¸Ù (Ø´Ù‡Ø±ÙŠ/Ø³Ù†ÙˆÙŠ)"),
]

PERM_ACTIONS = [
    ("READ", "Ù‚Ø±Ø§Ø¡Ø©"),
    ("CREATE", "Ø¥Ù†Ø´Ø§Ø¡"),
    ("UPDATE", "ØªØ¹Ø¯ÙŠÙ„"),
    ("DELETE", "Ø­Ø°Ù"),
]


def _clean(s): return (s or "").strip()

# -------------------------
# Legacy org lock + sync helpers
# -------------------------

def _get_setting(key: str, default: str | None = None) -> str | None:
    row = SystemSetting.query.filter_by(key=key).first()
    return (row.value if row else default)


def _set_setting(key: str, value: str | None):
    row = SystemSetting.query.filter_by(key=key).first()
    if not row:
        row = SystemSetting(key=key, value=(value or ""))
        db.session.add(row)
    else:
        row.value = (value or "")


def _legacy_org_locked() -> bool:
    return str(_get_setting("ORG_LEGACY_LOCKED", "0") or "0").strip() == "1"


@masterdata_bp.before_request
def _legacy_org_lock_guard():
    # When legacy org is locked, block CRUD/import on legacy hierarchy tables.
    # Read/list/export remain available. Dynamic org pages are unaffected.
    if not _legacy_org_locked():
        return

    path = (request.path or "")

    legacy_prefixes = (
        "/admin/masterdata/organizations",
        "/admin/masterdata/directorates",
        "/admin/masterdata/units",
        "/admin/masterdata/departments",
        "/admin/masterdata/sections",
        "/admin/masterdata/divisions",
    )

    if not any(path.startswith(pfx) for pfx in legacy_prefixes):
        return

    # Allow list + export only
    if request.method == "GET":
        if path in legacy_prefixes or path.endswith("/export.xlsx"):
            return

    flash("Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ø«Ø§Ø¨ØªØ© Ù…Ù‚ÙÙ„Ø© (Ù‚Ø±Ø§Ø¡Ø© ÙÙ‚Ø·). Ø§Ø³ØªØ®Ø¯Ù… Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù…ÙˆØ­Ø¯Ø© (Dynamic).", "warning")

    for pfx in legacy_prefixes:
        if path.startswith(pfx):
            return redirect(pfx)

    return redirect(url_for("masterdata.org_node_types_list"))





def _apply_sort(query, model, allowed: list[str], default_sort: str = "name_ar"):
    sort = (request.args.get("sort") or default_sort).strip()
    direction = (request.args.get("dir") or "asc").strip().lower()

    if sort not in allowed:
        sort = default_sort
    if direction not in ("asc", "desc"):
        direction = "asc"

    col = getattr(model, sort, None)
    if col is None:
        sort = default_sort
        col = getattr(model, sort)

    query = query.order_by(col.desc() if direction == "desc" else col.asc())
    return query, sort, direction


def _is_super_admin_user(u: User) -> bool:
    return ((getattr(u, "role", "") or "").strip().upper() == "SUPER_ADMIN")


def _current_is_super_admin() -> bool:
    # keep it local to avoid import cycles
    from flask_login import current_user
    return ((getattr(current_user, "role", "") or "").strip().upper() == "SUPER_ADMIN")

# -------------------------
# Organizations
# -------------------------
@masterdata_bp.route("/organizations")
@login_required
@perm_required("MASTERDATA_READ")
def org_list():
    q = request.args.get("q", "").strip()
    query = Organization.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Organization.name_ar.ilike(like),
            Organization.name_en.ilike(like),
            Organization.code.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Organization,
        allowed=["id", "name_ar", "name_en", "code", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    return render_template("admin/masterdata/org_list.html", items=items, q=q, sort=sort, direction=direction)



@masterdata_bp.route("/organizations/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def org_export_excel():
    q = request.args.get("q", "").strip()
    query = Organization.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Organization.name_ar.ilike(like),
            Organization.name_en.ilike(like),
            Organization.code.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Organization,
        allowed=["id", "name_ar", "name_en", "code", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    headers = ["ID", "Ø§Ù„Ø§Ø³Ù… (AR)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Code", "Ù†Ø´Ø·"]
    rows = []
    for o in items:
        rows.append([
            o.id,
            o.name_ar,
            o.name_en,
            o.code,
            "Ù†Ø¹Ù…" if o.is_active else "Ù„Ø§",
        ])

    data = make_xlsx_bytes("Organizations", headers, rows)
    filename = f"organizations_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@masterdata_bp.route("/organizations/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def org_new():
    if request.method == "POST":
        name_ar = _clean(request.form.get("name_ar"))
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ù…Ù†Ø¸Ù…Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        o = Organization(
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")),
            code=_clean(request.form.get("code")) or None,
            is_active=(request.form.get("is_active") == "1"),
        )
        db.session.add(o)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ù…Ù†Ø¸Ù…Ø©.", "success")
        return redirect(url_for("masterdata.org_list"))

    return render_template("admin/masterdata/org_form.html", o=None)

@masterdata_bp.route("/organizations/<int:org_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def org_edit(org_id):
    o = Organization.query.get_or_404(org_id)
    if request.method == "POST":
        name_ar = _clean(request.form.get("name_ar"))
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ù…Ù†Ø¸Ù…Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        o.name_ar = name_ar
        o.name_en = _clean(request.form.get("name_en"))
        o.code = _clean(request.form.get("code")) or None
        o.is_active = (request.form.get("is_active") == "1")

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.org_list"))

    return render_template("admin/masterdata/org_form.html", o=o)


@masterdata_bp.route("/organizations/<int:org_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def org_delete(org_id):
    o = Organization.query.get_or_404(org_id)
    try:
        has_children = Directorate.query.filter_by(organization_id=o.id).first() is not None
        if has_children:
            o.is_active = False
            db.session.commit()
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ù…Ù†Ø¸Ù…Ø© Ù„ÙˆØ¬ÙˆØ¯ Ø¥Ø¯Ø§Ø±Ø§Øª Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡Ø§ØŒ ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡Ø§ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
        else:
            db.session.delete(o)
            db.session.commit()
            flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ù…Ù†Ø¸Ù…Ø©.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„Ù…Ù†Ø¸Ù…Ø©: {e}", "danger")
    return redirect(url_for("masterdata.org_list"))


# -------------------------
# Directorates
# -------------------------
@masterdata_bp.route("/directorates")
@login_required
@perm_required("MASTERDATA_READ")
def dir_list():
    q = request.args.get("q", "").strip()
    query = Directorate.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Directorate.name_ar.ilike(like),
            Directorate.name_en.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Directorate,
        allowed=["id", "name_ar", "name_en", "code", "organization_id", "is_active"],
        default_sort="name_ar"
    )
    items = query.all()

    orgs = Organization.query.order_by(Organization.id.asc()).all()
    org_map = {o.id: o for o in orgs}
    return render_template(
        "admin/masterdata/dir_list.html",
        items=items,
        q=q,
        org_map=org_map,
        sort=sort,
        direction=direction,
    )



@masterdata_bp.route("/directorates/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def dir_export_excel():
    q = request.args.get("q", "").strip()
    query = Directorate.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Directorate.name_ar.ilike(like),
            Directorate.name_en.ilike(like),
            Directorate.code.ilike(like),
        ))

    query, sort, direction = _apply_sort(
        query,
        Directorate,
        allowed=["id", "name_ar", "name_en", "code", "organization_id", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    orgs = Organization.query.order_by(Organization.id.asc()).all()
    org_map = {o.id: o for o in orgs}

    headers = ["ID", "Ø§Ù„Ù…Ù†Ø¸Ù…Ø©", "Ø§Ù„Ø§Ø³Ù… (AR)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Code", "Ù†Ø´Ø·"]
    rows = []
    for d in items:
        rows.append([
            d.id,
            (org_map.get(d.organization_id).name_ar if org_map.get(d.organization_id) else d.organization_id),
            d.name_ar,
            d.name_en,
            d.code,
            "Ù†Ø¹Ù…" if d.is_active else "Ù„Ø§",
        ])

    data = make_xlsx_bytes("Directorates", headers, rows)
    filename = f"directorates_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    from io import BytesIO
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@masterdata_bp.route("/directorates/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def dir_new():
    orgs = Organization.query.order_by(Organization.name_ar.asc()).all()

    if request.method == "POST":
        org_id = request.form.get("organization_id")
        name_ar = _clean(request.form.get("name_ar"))

        if not (org_id or "").isdigit():
            flash("Ø§Ø®ØªØ± Ù…Ù†Ø¸Ù…Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        d = Directorate(
            organization_id=int(org_id),
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")),
            code=_clean(request.form.get("code")) or None,
            is_active=(request.form.get("is_active") == "1"),
        )
        db.session.add(d)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©.", "success")
        return redirect(url_for("masterdata.dir_list"))

    return render_template("admin/masterdata/dir_form.html", d=None, orgs=orgs)

@masterdata_bp.route("/directorates/<int:dir_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def dir_edit(dir_id):
    d = Directorate.query.get_or_404(dir_id)
    orgs = Organization.query.order_by(Organization.name_ar.asc()).all()

    if request.method == "POST":
        org_id = request.form.get("organization_id")
        name_ar = _clean(request.form.get("name_ar"))

        if not (org_id or "").isdigit():
            flash("Ø§Ø®ØªØ± Ù…Ù†Ø¸Ù…Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        d.organization_id = int(org_id)
        d.name_ar = name_ar
        d.name_en = _clean(request.form.get("name_en"))
        d.code = _clean(request.form.get("code")) or None
        d.is_active = (request.form.get("is_active") == "1")

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.dir_list"))

    return render_template("admin/masterdata/dir_form.html", d=d, orgs=orgs)


@masterdata_bp.route("/directorates/<int:dir_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def dir_delete(dir_id):
    d = Directorate.query.get_or_404(dir_id)
    try:
        has_children = Department.query.filter_by(directorate_id=d.id).first() is not None
        if has_children:
            d.is_active = False
            db.session.commit()
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ù„ÙˆØ¬ÙˆØ¯ Ø¯ÙˆØ§Ø¦Ø± Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡Ø§ØŒ ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡Ø§ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
        else:
            db.session.delete(d)
            db.session.commit()
            flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©: {e}", "danger")
    return redirect(url_for("masterdata.dir_list"))


# -------------------------
# Units / Ø§Ù„ÙˆØ­Ø¯Ø§Øª
# -------------------------
# -------------------------
# Units / Ø§Ù„ÙˆØ­Ø¯Ø§Øª
# -------------------------
@masterdata_bp.route("/units")
@login_required
@perm_required("MASTERDATA_READ")
def units_list():
    q = request.args.get("q", "").strip()
    query = Unit.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Unit.name_ar.ilike(like),
            Unit.name_en.ilike(like),
            Unit.code.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Unit,
        allowed=["id", "name_ar", "name_en", "code", "organization_id", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    orgs = Organization.query.order_by(Organization.id.asc()).all()
    org_map = {o.id: o for o in orgs}
    return render_template(
        "admin/masterdata/unit_list.html",
        items=items,
        q=q,
        sort=sort,
        direction=direction,
        org_map=org_map,
    )


@masterdata_bp.route("/units/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def units_export_excel():
    q = request.args.get("q", "").strip()
    query = Unit.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Unit.name_ar.ilike(like),
            Unit.name_en.ilike(like),
            Unit.code.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Unit,
        allowed=["id", "name_ar", "name_en", "code", "organization_id", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    orgs = Organization.query.order_by(Organization.id.asc()).all()
    org_map = {o.id: o for o in orgs}

    headers = ["ID", "Ø§Ù„Ù…Ù†Ø¸Ù…Ø©", "Ø§Ù„Ø§Ø³Ù… (AR)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Code", "Ù†Ø´Ø·"]
    rows = []
    for u in items:
        rows.append([
            u.id,
            (org_map.get(u.organization_id).name_ar if org_map.get(u.organization_id) else u.organization_id),
            u.name_ar,
            u.name_en,
            u.code,
            "Ù†Ø¹Ù…" if u.is_active else "Ù„Ø§",
        ])

    data = make_xlsx_bytes("Units", headers, rows)
    filename = f"units_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


@masterdata_bp.route("/units/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def unit_new():
    orgs = Organization.query.order_by(Organization.name_ar.asc()).all()
    if request.method == "POST":
        organization_id = request.form.get("organization_id")
        name_ar = _clean(request.form.get("name_ar"))

        if not (organization_id or "").isdigit():
            flash("Ø§Ø®ØªØ± Ù…Ù†Ø¸Ù…Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„ÙˆØ­Ø¯Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        u = Unit(
            organization_id=int(organization_id),
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")),
            code=_clean(request.form.get("code")) or None,
            is_active=(request.form.get("is_active") == "1"),
        )
        db.session.add(u)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„ÙˆØ­Ø¯Ø©.", "success")
        return redirect(url_for("masterdata.units_list"))

    return render_template("admin/masterdata/unit_form.html", u=None, orgs=orgs)


@masterdata_bp.route("/units/<int:unit_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def unit_edit(unit_id):
    u = Unit.query.get_or_404(unit_id)
    orgs = Organization.query.order_by(Organization.name_ar.asc()).all()

    if request.method == "POST":
        organization_id = request.form.get("organization_id")
        name_ar = _clean(request.form.get("name_ar"))

        if not (organization_id or "").isdigit():
            flash("Ø§Ø®ØªØ± Ù…Ù†Ø¸Ù…Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„ÙˆØ­Ø¯Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        u.organization_id = int(organization_id)
        u.name_ar = name_ar
        u.name_en = _clean(request.form.get("name_en"))
        u.code = _clean(request.form.get("code")) or None
        u.is_active = (request.form.get("is_active") == "1")

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.units_list"))

    return render_template("admin/masterdata/unit_form.html", u=u, orgs=orgs)


@masterdata_bp.route("/units/<int:unit_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def unit_delete(unit_id):
    u = Unit.query.get_or_404(unit_id)
    try:
        has_children = Department.query.filter_by(unit_id=u.id).first() is not None
        if has_children:
            u.is_active = False
            db.session.commit()
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„ÙˆØ­Ø¯Ø© Ù„ÙˆØ¬ÙˆØ¯ Ø¯ÙˆØ§Ø¦Ø± Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡Ø§ØŒ ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡Ø§ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
        else:
            db.session.delete(u)
            db.session.commit()
            flash("ØªÙ… Ø­Ø°Ù Ø§Ù„ÙˆØ­Ø¯Ø©.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„ÙˆØ­Ø¯Ø©: {e}", "danger")
    return redirect(url_for("masterdata.units_list"))


# -------------------------
# Departments
# -------------------------
@masterdata_bp.route("/departments")
@login_required
@perm_required("MASTERDATA_READ")
def dept_list():
    q = request.args.get("q", "").strip()
    query = Department.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Department.name_ar.ilike(like),
            Department.name_en.ilike(like)
        ))

    query, sort, direction = _apply_sort(
        query,
        Department,
        allowed=["id", "name_ar", "name_en", "code", "directorate_id", "unit_id", "is_active"],
        default_sort="name_ar"
    )
    items = query.all()

    dirs = Directorate.query.order_by(Directorate.id.asc()).all()
    units = Unit.query.order_by(Unit.id.asc()).all()
    dir_map = {d.id: d for d in dirs}
    unit_map = {u.id: u for u in units}
    return render_template(
        "admin/masterdata/dept_list.html",
        items=items,
        q=q,
        dir_map=dir_map,
        unit_map=unit_map,
        sort=sort,
        direction=direction,
    )



@masterdata_bp.route("/departments/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def dept_export_excel():
    q = request.args.get("q", "").strip()
    query = Department.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Department.name_ar.ilike(like),
            Department.name_en.ilike(like),
            Department.code.ilike(like),
        ))

    query, sort, direction = _apply_sort(
        query,
        Department,
        allowed=["id", "name_ar", "name_en", "code", "directorate_id", "unit_id", "is_active"],
        default_sort="name_ar"
    )

    items = query.all()
    dirs = Directorate.query.order_by(Directorate.id.asc()).all()
    units = Unit.query.order_by(Unit.id.asc()).all()
    dir_map = {d.id: d for d in dirs}
    unit_map = {u.id: u for u in units}

    headers = ["ID", "Ø§Ù„ØªØ¨Ø¹ÙŠØ©", "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (AR)", "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (EN)", "Code", "Ù†Ø´Ø·"]
    rows = []
    for dep in items:
        parent_label = "-"
        if dep.directorate_id:
            dir_obj = dir_map.get(dep.directorate_id)
            parent_label = (dir_obj.name_ar if dir_obj else str(dep.directorate_id))
        elif dep.unit_id:
            u_obj = unit_map.get(dep.unit_id)
            parent_label = (f"ÙˆØ­Ø¯Ø©: {u_obj.name_ar}" if u_obj else f"ÙˆØ­Ø¯Ø©: {dep.unit_id}")
        rows.append([
            dep.id,
            parent_label,
            dep.name_ar,
            dep.name_en,
            dep.code,
            "Ù†Ø¹Ù…" if dep.is_active else "Ù„Ø§",
        ])

    data = make_xlsx_bytes("Departments", headers, rows)
    filename = f"departments_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@masterdata_bp.route("/departments/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def dept_new():
    dirs = Directorate.query.order_by(Directorate.name_ar.asc()).all()
    units = Unit.query.order_by(Unit.name_ar.asc()).all()

    if request.method == "POST":
        dir_id = (request.form.get("directorate_id") or "").strip()
        unit_id = (request.form.get("unit_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))

        # XOR: one parent only
        picked = [x for x in [dir_id if dir_id.isdigit() else None, unit_id if unit_id.isdigit() else None] if x]
        if len(picked) != 1:
            flash("Ø§Ø®ØªØ±: Ø¥Ø¯Ø§Ø±Ø© Ø£Ùˆ ÙˆØ­Ø¯Ø© (ÙˆØ§Ø­Ø¯ ÙÙ‚Ø·).", "danger")
            return redirect(request.url)

        if dir_id.isdigit() and not Directorate.query.get(int(dir_id)):
            flash("Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if unit_id.isdigit() and not Unit.query.get(int(unit_id)):
            flash("Ø§Ù„ÙˆØ­Ø¯Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        dep = Department(
            directorate_id=(int(dir_id) if dir_id.isdigit() else None),
            unit_id=(int(unit_id) if unit_id.isdigit() else None),
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")),
            code=_clean(request.form.get("code")) or None,
            is_active=(request.form.get("is_active") == "1"),
        )
        db.session.add(dep)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©.", "success")
        return redirect(url_for("masterdata.dept_list"))

    return render_template("admin/masterdata/dept_form.html", dep=None, dirs=dirs, units=units)

@masterdata_bp.route("/departments/<int:dept_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def dept_edit(dept_id):
    dep = Department.query.get_or_404(dept_id)
    dirs = Directorate.query.order_by(Directorate.name_ar.asc()).all()
    units = Unit.query.order_by(Unit.name_ar.asc()).all()

    if request.method == "POST":
        dir_id = (request.form.get("directorate_id") or "").strip()
        unit_id = (request.form.get("unit_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))

        # XOR: one parent only
        picked = [x for x in [dir_id if dir_id.isdigit() else None, unit_id if unit_id.isdigit() else None] if x]
        if len(picked) != 1:
            flash("Ø§Ø®ØªØ±: Ø¥Ø¯Ø§Ø±Ø© Ø£Ùˆ ÙˆØ­Ø¯Ø© (ÙˆØ§Ø­Ø¯ ÙÙ‚Ø·).", "danger")
            return redirect(request.url)

        if dir_id.isdigit() and not Directorate.query.get(int(dir_id)):
            flash("Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if unit_id.isdigit() and not Unit.query.get(int(unit_id)):
            flash("Ø§Ù„ÙˆØ­Ø¯Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        dep.directorate_id = int(dir_id) if dir_id.isdigit() else None
        dep.unit_id = int(unit_id) if unit_id.isdigit() else None
        dep.name_ar = name_ar
        dep.name_en = _clean(request.form.get("name_en"))
        dep.code = _clean(request.form.get("code")) or None
        dep.is_active = (request.form.get("is_active") == "1")

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.dept_list"))

    return render_template("admin/masterdata/dept_form.html", dep=dep, dirs=dirs, units=units)


@masterdata_bp.route("/departments/<int:dept_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def dept_delete(dept_id):
    dep = Department.query.get_or_404(dept_id)
    try:
        has_children = Section.query.filter_by(department_id=dep.id).first() is not None
        if has_children:
            dep.is_active = False
            db.session.commit()
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© Ù„ÙˆØ¬ÙˆØ¯ Ø£Ù‚Ø³Ø§Ù… Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡Ø§ØŒ ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡Ø§ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
        else:
            db.session.delete(dep)
            db.session.commit()
            flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©.", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©: {e}", "danger")
    return redirect(url_for("masterdata.dept_list"))


# -------------------------
# Sections (Ø£Ù‚Ø³Ø§Ù…)
# -------------------------
@masterdata_bp.route("/sections")
@login_required
@perm_required("MASTERDATA_READ")
def sections_list():
    q = request.args.get("q", "").strip()

    query = Section.query
    if q:
        like = f"%{q}%"
        query = (
            query
            .outerjoin(Department, Section.department_id == Department.id)
            .outerjoin(Unit, Section.unit_id == Unit.id)
            .outerjoin(
                Directorate,
                or_(
                    Section.directorate_id == Directorate.id,
                    Department.directorate_id == Directorate.id,
                )
            )
            .filter(or_(
                Section.name_ar.ilike(like),
                Section.name_en.ilike(like),
                Section.code.ilike(like),
                Department.name_ar.ilike(like),
                Department.name_en.ilike(like),
                Unit.name_ar.ilike(like),
                Unit.name_en.ilike(like),
                Unit.code.ilike(like),
                Directorate.name_ar.ilike(like),
                Directorate.name_en.ilike(like),
            ))
        )

    query, sort, direction = _apply_sort(
        query,
        Section,
        allowed=["id", "name_ar", "name_en", "code", "department_id", "directorate_id", "unit_id", "is_active"],
        default_sort="name_ar",
    )

    items = query.all()
    dept_map = {d.id: d for d in Department.query.order_by(Department.id.asc()).all()}
    dir_map = {d.id: d for d in Directorate.query.order_by(Directorate.id.asc()).all()}
    unit_map = {u.id: u for u in Unit.query.order_by(Unit.id.asc()).all()}
    org_map = {o.id: o for o in Organization.query.order_by(Organization.id.asc()).all()}

    return render_template(
        "admin/masterdata/sections_list.html",
        items=items,
        q=q,
        dept_map=dept_map,
        dir_map=dir_map,
        unit_map=unit_map,
        org_map=org_map,
        sort=sort,
        direction=direction,
    )




@masterdata_bp.route("/sections/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def sections_export_excel():
    # Export sections, allowing Section to belong to either:
    # - Department (which may belong to a Directorate or a Unit)
    # - directly to a Directorate
    # - directly to a Unit
    sections = Section.query.order_by(Section.id.asc()).all()
    dept_map = {d.id: d for d in Department.query.all()}
    dir_map = {d.id: d for d in Directorate.query.all()}
    unit_map = {u.id: u for u in Unit.query.all()}
    org_map = {o.id: o for o in Organization.query.all()}

    headers = ["ID", "Ø§Ù„Ù…Ù†Ø¸Ù…Ø©/Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©", "Ø§Ø³Ù… Ø§Ù„Ù‚Ø³Ù… (AR)", "Ø§Ø³Ù… Ø§Ù„Ù‚Ø³Ù… (EN)", "Code", "Ù†Ø´Ø·"]
    rows = []
    for s in sections:
        dep = dept_map.get(s.department_id) if s.department_id else None
        unit_obj = unit_map.get(s.unit_id) if s.unit_id else None

        dir_obj = None
        if dep and dep.directorate_id:
            dir_obj = dir_map.get(dep.directorate_id)
        elif s.directorate_id:
            dir_obj = dir_map.get(s.directorate_id)

        top_label = "-"
        if dir_obj:
            top_label = dir_obj.name_ar
        elif unit_obj and unit_obj.organization_id:
            top_label = (org_map.get(unit_obj.organization_id).name_ar if org_map.get(unit_obj.organization_id) else "-")

        rows.append([
            s.id,
            top_label,
            (dep.name_ar if dep else (f"ÙˆØ­Ø¯Ø©: {unit_obj.name_ar}" if unit_obj else "-")),
            s.name_ar,
            s.name_en,
            s.code,
            "Ù†Ø¹Ù…" if s.is_active else "Ù„Ø§",
        ])

    data = make_xlsx_bytes("Sections", headers, rows)
    filename = f"sections_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(data),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )

@masterdata_bp.route("/sections/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def sections_new():
    departments = Department.query.order_by(Department.name_ar.asc()).all()
    directorates = Directorate.query.order_by(Directorate.name_ar.asc()).all()
    units = Unit.query.order_by(Unit.name_ar.asc()).all()

    if request.method == "POST":
        dept_id = (request.form.get("department_id") or "").strip()
        dir_id = (request.form.get("directorate_id") or "").strip()
        unit_id = (request.form.get("unit_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        code = _clean(request.form.get("code")) or None
        is_active = (request.form.get("is_active") == "1")

        # XOR: exactly one parent
        picked = [x for x in [dept_id if dept_id.isdigit() else None, dir_id if dir_id.isdigit() else None, unit_id if unit_id.isdigit() else None] if x]
        if len(picked) != 1:
            flash("Ø§Ø®ØªØ±: Ø¯Ø§Ø¦Ø±Ø© Ø£Ùˆ Ø¥Ø¯Ø§Ø±Ø© Ø£Ùˆ ÙˆØ­Ø¯Ø© (ÙˆØ§Ø­Ø¯ ÙÙ‚Ø·).", "danger")
            return redirect(request.url)

        department = Department.query.get(int(dept_id)) if dept_id.isdigit() else None
        directorate = Directorate.query.get(int(dir_id)) if dir_id.isdigit() else None
        unit = Unit.query.get(int(unit_id)) if unit_id.isdigit() else None

        if dept_id.isdigit() and not department:
            flash("Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if dir_id.isdigit() and not directorate:
            flash("Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if unit_id.isdigit() and not unit:
            flash("Ø§Ù„ÙˆØ­Ø¯Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ù‚Ø³Ù… (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        s = Section(
            department_id=department.id if department else None,
            directorate_id=directorate.id if directorate else None,
            unit_id=unit.id if unit else None,
            name_ar=name_ar,
            name_en=name_en,
            code=code,
            is_active=is_active,
        )
        db.session.add(s)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ù‚Ø³Ù….", "success")
        return redirect(url_for("masterdata.sections_list"))

    return render_template(
        "admin/masterdata/sections_form.html",
        s=None,
        departments=departments,
        directorates=directorates,
        units=units,
    )


@masterdata_bp.route("/sections/<int:section_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def sections_edit(section_id):
    s = Section.query.get_or_404(section_id)
    departments = Department.query.order_by(Department.name_ar.asc()).all()
    directorates = Directorate.query.order_by(Directorate.name_ar.asc()).all()
    units = Unit.query.order_by(Unit.name_ar.asc()).all()

    if request.method == "POST":
        dept_id = (request.form.get("department_id") or "").strip()
        dir_id = (request.form.get("directorate_id") or "").strip()
        unit_id = (request.form.get("unit_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        code = _clean(request.form.get("code")) or None
        is_active = (request.form.get("is_active") == "1")

        # XOR: exactly one parent
        picked = [x for x in [dept_id if dept_id.isdigit() else None, dir_id if dir_id.isdigit() else None, unit_id if unit_id.isdigit() else None] if x]
        if len(picked) != 1:
            flash("Ø§Ø®ØªØ±: Ø¯Ø§Ø¦Ø±Ø© Ø£Ùˆ Ø¥Ø¯Ø§Ø±Ø© Ø£Ùˆ ÙˆØ­Ø¯Ø© (ÙˆØ§Ø­Ø¯ ÙÙ‚Ø·).", "danger")
            return redirect(request.url)

        department = Department.query.get(int(dept_id)) if dept_id.isdigit() else None
        directorate = Directorate.query.get(int(dir_id)) if dir_id.isdigit() else None
        unit = Unit.query.get(int(unit_id)) if unit_id.isdigit() else None

        if dept_id.isdigit() and not department:
            flash("Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if dir_id.isdigit() and not directorate:
            flash("Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if unit_id.isdigit() and not unit:
            flash("Ø§Ù„ÙˆØ­Ø¯Ø© Ø§Ù„Ù…Ø®ØªØ§Ø±Ø© ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯Ø©.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ø³Ù… Ø§Ù„Ù‚Ø³Ù… (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        s.department_id = department.id if department else None
        s.directorate_id = directorate.id if directorate else None
        s.unit_id = unit.id if unit else None
        s.name_ar = name_ar
        s.name_en = name_en
        s.code = code
        s.is_active = is_active

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.sections_list"))

    return render_template(
        "admin/masterdata/sections_form.html",
        s=s,
        departments=departments,
        directorates=directorates,
        units=units,
    )


@masterdata_bp.route("/sections/<int:section_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def sections_delete(section_id):
    s = Section.query.get_or_404(section_id)
    try:
        db.session.delete(s)
        db.session.commit()
        flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ù‚Ø³Ù….", "warning")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„Ù‚Ø³Ù…: {e}", "danger")
    return redirect(url_for("masterdata.sections_list"))

# -------------------------
# Divisions (Ø´ÙØ¹Ø¨)
# -------------------------
@masterdata_bp.route("/divisions")
@login_required
@perm_required("MASTERDATA_READ")
def divisions_list():
    q = request.args.get("q", "").strip()
    query = Division.query
    if q:
        like = f"%{q}%"
        # Search in division fields + parent names
        query = (query
                 .outerjoin(Section, Division.section_id == Section.id)
                 .outerjoin(Department, Division.department_id == Department.id)
                 .filter(or_(
                     Division.name_ar.ilike(like),
                     Division.name_en.ilike(like),
                     Division.code.ilike(like),
                     Section.name_ar.ilike(like),
                     Section.name_en.ilike(like),
                     Department.name_ar.ilike(like),
                     Department.name_en.ilike(like),
                     Department.code.ilike(like),
                 )))

    allowed = ["id", "department_id", "section_id", "name_ar", "name_en", "code", "is_active", "created_at"]
    query, sort, direction = _apply_sort(query, Division, allowed, default_sort="id")

    items = query.all()
    sections = Section.query.order_by(Section.name_ar.asc()).all()
    departments = Department.query.order_by(Department.name_ar.asc()).all()
    directorates = Directorate.query.order_by(Directorate.name_ar.asc()).all()

    sec_map = {s.id: s for s in sections}
    dept_map = {d.id: d for d in departments}
    dir_map = {d.id: d for d in directorates}

    return render_template(
        "admin/masterdata/divisions_list.html",
        items=items,
        q=q,
        sort=sort,
        direction=direction,
        dir_map=dir_map,
        sec_map=sec_map,
        dept_map=dept_map,
    )

@masterdata_bp.route("/divisions/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def divisions_export_excel():
    """Export divisions list as Excel (honors current search/sort)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    q = request.args.get("q", "").strip()
    query = Division.query
    if q:
        like = f"%{q}%"
        query = (query
                 .outerjoin(Section, Division.section_id == Section.id)
                 .outerjoin(Department, Division.department_id == Department.id)
                 .filter(or_(
                     Division.name_ar.ilike(like),
                     Division.name_en.ilike(like),
                     Division.code.ilike(like),
                     Section.name_ar.ilike(like),
                     Section.name_en.ilike(like),
                     Department.name_ar.ilike(like),
                     Department.name_en.ilike(like),
                     Department.code.ilike(like),
                 )))
    allowed = ["id", "department_id", "section_id", "name_ar", "name_en", "code", "is_active", "created_at"]
    query, _, _ = _apply_sort(query, Division, allowed, default_sort="id")
    rows = query.all()

    # Build maps for display columns
    sections = {s.id: s for s in Section.query.all()}
    departments = {d.id: d for d in Department.query.all()}
    directorates = {d.id: d for d in Directorate.query.all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Divisions"
    headers = ["ID", "Directorate", "Department/Office", "Section", "Division (AR)", "Division (EN)", "Code", "Active", "Created"]
    ws.append(headers)
    for r in rows:
        sec = sections.get(r.section_id) if r.section_id else None
        dept = departments.get(r.department_id) if r.department_id else None
        dir_obj = None
        if dept and getattr(dept, "directorate_id", None):
            dir_obj = directorates.get(dept.directorate_id)
        elif sec and getattr(sec, "department_id", None):
            dep2 = departments.get(sec.department_id)
            if dep2 and getattr(dep2, "directorate_id", None):
                dir_obj = directorates.get(dep2.directorate_id)
        ws.append([
            r.id,
            (dir_obj.name_ar if dir_obj else ""),
            (dept.name_ar if dept else ""),
            (sec.name_ar if sec else ""),
            r.name_ar,
            (r.name_en or ""),
            (r.code or ""),
            "Yes" if r.is_active else "No",
            r.created_at.strftime("%Y-%m-%d") if r.created_at else "",
        ])
    # Auto column widths (simple)
    for col in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col)]:
            v = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(v))
        ws.column_dimensions[get_column_letter(col)].width = min(max(12, max_len + 2), 50)

    import io
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = "divisions.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
@masterdata_bp.route("/divisions/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def divisions_new():
    sections = Section.query.order_by(Section.name_ar.asc()).all()
    departments = Department.query.order_by(Department.name_ar.asc()).all()

    if request.method == "POST":
        code = _clean(request.form.get("code")).upper()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        is_active = (request.form.get("is_active") == "1")

        section_id = request.form.get("section_id") or None
        department_id = request.form.get("department_id") or None

        if not name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        if section_id and section_id.isdigit():
            section_id = int(section_id)
        else:
            section_id = None

        if department_id and department_id.isdigit():
            department_id = int(department_id)
        else:
            department_id = None

        if not section_id and not department_id:
            flash("ÙŠØ¬Ø¨ Ø§Ø®ØªÙŠØ§Ø± Ù‚Ø³Ù… Ø£Ùˆ Ø¯Ø§Ø¦Ø±Ø© (ÙˆØ§Ø­Ø¯ Ø¹Ù„Ù‰ Ø§Ù„Ø£Ù‚Ù„).", "danger")
            return redirect(request.url)

        if code and Division.query.filter_by(code=code).first():
            flash("Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ Ù…Ø³ØªØ®Ø¯Ù… Ù…Ø³Ø¨Ù‚Ø§Ù‹.", "danger")
            return redirect(request.url)

        dv = Division(
            code=code or None,
            name_ar=name_ar,
            name_en=name_en,
            is_active=is_active,
            section_id=section_id,
            department_id=department_id,
        )
        db.session.add(dv)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ø´Ø¹Ø¨Ø©.", "success")
        return redirect(url_for("masterdata.divisions_list"))

    return render_template(
        "admin/masterdata/divisions_form.html",
        dv=None,
        division=None,
        sections=sections,
        departments=departments,
    )


@masterdata_bp.route("/divisions/<int:division_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def divisions_edit(division_id):
    dv = Division.query.get_or_404(division_id)
    sections = Section.query.order_by(Section.name_ar.asc()).all()
    departments = Department.query.order_by(Department.name_ar.asc()).all()

    if request.method == "POST":
        code = _clean(request.form.get("code")).upper()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        is_active = (request.form.get("is_active") == "1")

        section_id = request.form.get("section_id") or None
        department_id = request.form.get("department_id") or None

        if not name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        if section_id and section_id.isdigit():
            section_id = int(section_id)
        else:
            section_id = None

        if department_id and department_id.isdigit():
            department_id = int(department_id)
        else:
            department_id = None

        if not section_id and not department_id:
            flash("ÙŠØ¬Ø¨ Ø§Ø®ØªÙŠØ§Ø± Ù‚Ø³Ù… Ø£Ùˆ Ø¯Ø§Ø¦Ø±Ø© (ÙˆØ§Ø­Ø¯ Ø¹Ù„Ù‰ Ø§Ù„Ø£Ù‚Ù„).", "danger")
            return redirect(request.url)

        if code:
            other = Division.query.filter(Division.code == code, Division.id != dv.id).first()
            if other:
                flash("Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ Ù…Ø³ØªØ®Ø¯Ù… Ù…Ø³Ø¨Ù‚Ø§Ù‹.", "danger")
                return redirect(request.url)

        dv.code = code or None
        dv.name_ar = name_ar
        dv.name_en = name_en
        dv.is_active = is_active
        dv.section_id = section_id
        dv.department_id = department_id

        db.session.commit()
        flash("ØªÙ… ØªØ­Ø¯ÙŠØ« Ø§Ù„Ø´Ø¹Ø¨Ø©.", "success")
        return redirect(url_for("masterdata.divisions_list"))

    return render_template(
        "admin/masterdata/divisions_form.html",
        dv=dv,
        division=dv,
        sections=sections,
        departments=departments,
    )


@masterdata_bp.route("/divisions/<int:division_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def divisions_delete(division_id):
    dv = Division.query.get_or_404(division_id)
    try:
        db.session.delete(dv)
        db.session.commit()
        flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø´Ø¹Ø¨Ø©.", "warning")
    except Exception as e:
        db.session.rollback()
        flash(f"ØªØ¹Ø°Ø± Ø­Ø°Ù Ø§Ù„Ø´Ø¹Ø¨Ø©: {e}", "danger")
    return redirect(url_for("masterdata.divisions_list"))

# -------------------------
# Request Types
# -------------------------
@masterdata_bp.route("/request-types")
@login_required
@perm_required("REQUEST_TYPES_READ")
def request_types_list():
    q = request.args.get("q", "").strip()
    query = RequestType.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            RequestType.name_ar.ilike(like),
            RequestType.name_en.ilike(like),
            RequestType.code.ilike(like)
        ))

    items = query.order_by(RequestType.id.asc()).all()
    return render_template("admin/masterdata/request_types_list.html", items=items, q=q)

@masterdata_bp.route("/request-types/new", methods=["GET", "POST"])
@login_required
@perm_required("REQUEST_TYPES_CREATE")
def request_types_new():
    if request.method == "POST":
        code = _clean(request.form.get("code")).upper()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        is_active = (request.form.get("is_active") == "1")

        if not code:
            flash("Ø§Ù„ÙƒÙˆØ¯ (CODE) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        exists = RequestType.query.filter_by(code=code).first()
        if exists:
            flash("Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ Ù…Ø³ØªØ®Ø¯Ù… Ù…Ø³Ø¨Ù‚Ø§Ù‹.", "danger")
            return redirect(request.url)

        rt = RequestType(code=code, name_ar=name_ar, name_en=name_en, is_active=is_active)
        db.session.add(rt)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ù†ÙˆØ¹ Ø§Ù„Ø·Ù„Ø¨.", "success")
        return redirect(url_for("masterdata.request_types_list"))

    return render_template("admin/masterdata/request_types_form.html", rt=None)

@masterdata_bp.route("/request-types/<int:rt_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("REQUEST_TYPES_UPDATE")
def request_types_edit(rt_id):
    rt = RequestType.query.get_or_404(rt_id)

    if request.method == "POST":
        code = _clean(request.form.get("code")).upper()
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en")) or None
        is_active = (request.form.get("is_active") == "1")

        if not code:
            flash("Ø§Ù„ÙƒÙˆØ¯ (CODE) Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)
        if not name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        other = RequestType.query.filter(RequestType.code == code, RequestType.id != rt.id).first()
        if other:
            flash("Ù‡Ø°Ø§ Ø§Ù„ÙƒÙˆØ¯ Ù…Ø³ØªØ®Ø¯Ù… Ù…Ø³Ø¨Ù‚Ø§Ù‹.", "danger")
            return redirect(request.url)

        rt.code = code
        rt.name_ar = name_ar
        rt.name_en = name_en
        rt.is_active = is_active

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.request_types_list"))

    return render_template("admin/masterdata/request_types_form.html", rt=rt)

@masterdata_bp.route("/request-types/<int:rt_id>/delete", methods=["POST"])
@login_required
@perm_required("REQUEST_TYPES_DELETE")
def request_types_delete(rt_id):
    rt = RequestType.query.get_or_404(rt_id)

    # prevent delete if used
    used_in_req = WorkflowRequest.query.filter(WorkflowRequest.request_type_id == rt.id).count()
    used_in_rules = WorkflowRoutingRule.query.filter(WorkflowRoutingRule.request_type_id == rt.id).count()

    if used_in_req or used_in_rules:
        # safer: deactivate
        rt.is_active = False
        db.session.commit()
        flash("Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ Ù…Ø³ØªØ®Ø¯Ù… Ø¨Ø§Ù„ÙØ¹Ù„. ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø§Ù„Ø­Ø°Ù.", "warning")
        return redirect(url_for("masterdata.request_types_list"))

    db.session.delete(rt)
    db.session.commit()
    flash("ØªÙ… Ø­Ø°Ù Ù†ÙˆØ¹ Ø§Ù„Ø·Ù„Ø¨.", "warning")
    return redirect(url_for("masterdata.request_types_list"))

# -------------------------
# Permissions management
# -------------------------
# -------------------------
# Permissions management
# -------------------------
@masterdata_bp.route("/permissions", methods=["GET", "POST"])
@login_required
@perm_required("USER_PERMISSIONS_UPDATE")
def permissions_manage():
    users = User.query.order_by(User.id.asc()).all()
    if not _current_is_super_admin():
        users = [u for u in users if not _is_super_admin_user(u)]

    selected_user_id = request.args.get("user_id")

    action_codes = [a for a, _ in PERM_ACTIONS]  # e.g. ["CREATE","READ","UPDATE","DELETE"]

    # Build managed keys list for cleanup (CRUD + legacy *_MANAGE + extra)
    managed_keys = set()
    for prefix, _label, legacy in PERM_MODULES:
        for act in action_codes:
            managed_keys.add(f"{prefix}_{act}")
        if legacy:
            managed_keys.add(legacy)
    for k, _lbl in PERM_EXTRA_KEYS:
        managed_keys.add(k)

    if request.method == "POST":
        uid = (request.form.get("user_id") or "").strip()
        if not uid.isdigit():
            flash("ÙŠØ±Ø¬Ù‰ Ø§Ø®ØªÙŠØ§Ø± Ù…Ø³ØªØ®Ø¯Ù… ØµØ­ÙŠØ­.", "danger")
            return redirect(url_for("masterdata.permissions_manage"))

        uid = int(uid)
        target = User.query.get(uid)
        if not target:
            flash("Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯.", "danger")
            return redirect(url_for("masterdata.permissions_manage"))

        # HARD BLOCK: only SUPER_ADMIN can edit SUPER_ADMIN permissions
        if _is_super_admin_user(target) and not _current_is_super_admin():
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ†Ùƒ ØªØ¹Ø¯ÙŠÙ„ ØµÙ„Ø§Ø­ÙŠØ§Øª Ø­Ø³Ø§Ø¨ SUPER_ADMIN.", "danger")
            return redirect(url_for("masterdata.permissions_manage"))

        # Remove old perms for managed keys
        UserPermission.query.filter(
            UserPermission.user_id == uid,
            UserPermission.key.in_(list(managed_keys))
        ).delete(synchronize_session=False)

        # Add selected CRUD perms
        for prefix, _label, _legacy in PERM_MODULES:
            for act in action_codes:
                if request.form.get(f"perm_{prefix}_{act}") == "1":
                    k = f"{prefix}_{act}"
                    db.session.add(UserPermission(user_id=uid, key=k, is_allowed=True))

        # Add selected EXTRA perms
        for k, _lbl in PERM_EXTRA_KEYS:
            if request.form.get(f"perm_extra_{k}") == "1":
                db.session.add(UserPermission(user_id=uid, key=k, is_allowed=True))

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ ØµÙ„Ø§Ø­ÙŠØ§Øª Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù….", "success")
        return redirect(url_for("masterdata.permissions_manage", user_id=uid))

    # GET view
    selected = None
    checked_keys = set()
    current_keys = set()

    if selected_user_id and str(selected_user_id).isdigit():
        uid = int(selected_user_id)
        selected = User.query.get(uid)
        if selected:
            if _is_super_admin_user(selected) and not _current_is_super_admin():
                flash("Ù„Ø§ ÙŠÙ…ÙƒÙ†Ùƒ Ø¹Ø±Ø¶/ØªØ¹Ø¯ÙŠÙ„ ØµÙ„Ø§Ø­ÙŠØ§Øª Ø­Ø³Ø§Ø¨ SUPER_ADMIN.", "warning")
                return redirect(url_for("masterdata.permissions_manage"))

            rows = UserPermission.query.filter_by(user_id=uid).all()
            current_keys = {
                (p.key or "").strip().upper()
                for p in rows
                if p.is_allowed
            }

            # CRUD keys (+ legacy mapping)
            for prefix, _label, legacy in PERM_MODULES:
                # legacy -> all CRUD enabled
                if legacy and legacy.upper() in current_keys:
                    for act in action_codes:
                        checked_keys.add(f"{prefix}_{act}")
                    checked_keys.add(legacy)
                    continue

                for act in action_codes:
                    k = f"{prefix}_{act}".upper()
                    if k in current_keys:
                        checked_keys.add(f"{prefix}_{act}")

                if legacy and legacy.upper() in current_keys:
                    checked_keys.add(legacy)

            # Extra keys
            for k, _lbl in PERM_EXTRA_KEYS:
                if k.upper() in current_keys:
                    checked_keys.add(k)

    return render_template(
        "admin/masterdata/permissions.html",
        users=users,
        selected=selected,
        perm_modules=PERM_MODULES,
        perm_actions=PERM_ACTIONS,
        checked_keys=checked_keys,
        current_keys=current_keys,
        extra_keys=PERM_EXTRA_KEYS,
    )
@masterdata_bp.route("/roles")
@login_required
@perm_required("MASTERDATA_READ")
def roles_list():
    q = request.args.get("q", "").strip()
    query = Role.query

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Role.code.ilike(like),
            Role.name_ar.ilike(like),
            Role.name_en.ilike(like)
        ))

    items = query.order_by(Role.id.asc()).all()
    return render_template("admin/masterdata/roles_list.html", items=items, q=q)


@masterdata_bp.route("/roles/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def roles_new():
    if request.method == "POST":
        code = _clean(request.form.get("code"))
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en"))
        is_active = (request.form.get("is_active") == "1")

        if not code:
            flash("ÙƒÙˆØ¯ Ø§Ù„Ø¯ÙˆØ± Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        code = code.replace(" ", "_")
        if Role.query.filter_by(code=code).first():
            flash("Ù‡Ø°Ø§ Ø§Ù„Ø¯ÙˆØ± Ù…ÙˆØ¬ÙˆØ¯ Ù…Ø³Ø¨Ù‚Ù‹Ø§.", "danger")
            return redirect(request.url)

        r = Role(code=code, name_ar=name_ar or None, name_en=name_en or None, is_active=is_active)
        db.session.add(r)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ø¯ÙˆØ±.", "success")
        return redirect(url_for("masterdata.roles_list"))

    return render_template("admin/masterdata/roles_form.html", r=None)


@masterdata_bp.route("/roles/<int:role_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def roles_edit(role_id):
    r = Role.query.get_or_404(role_id)

    if request.method == "POST":
        code = _clean(request.form.get("code"))
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en"))
        is_active = (request.form.get("is_active") == "1")

        if not code:
            flash("ÙƒÙˆØ¯ Ø§Ù„Ø¯ÙˆØ± Ù…Ø·Ù„ÙˆØ¨.", "danger")
            return redirect(request.url)

        code = code.replace(" ", "_")
        exists = Role.query.filter(Role.code == code, Role.id != r.id).first()
        if exists:
            flash("ÙŠÙˆØ¬Ø¯ Ø¯ÙˆØ± Ø¢Ø®Ø± Ø¨Ù†ÙØ³ Ø§Ù„ÙƒÙˆØ¯.", "danger")
            return redirect(request.url)

        r.code = code
        r.name_ar = name_ar or None
        r.name_en = name_en or None
        r.is_active = is_active

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„Ø¯ÙˆØ±.", "success")
        return redirect(url_for("masterdata.roles_list"))

    return render_template("admin/masterdata/roles_form.html", r=r)


@masterdata_bp.route("/roles/<int:role_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def roles_delete(role_id):
    r = Role.query.get_or_404(role_id)
    try:
        used = User.query.filter(User.role.isnot(None)).filter(User.role.ilike(r.code)).count() > 0
    except Exception:
        used = False

    if used:
        r.is_active = False
        db.session.commit()
        flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ø¯ÙˆØ± Ù„Ø£Ù†Ù‡ Ù…Ø³ØªØ®Ø¯Ù… Ù„Ø¯Ù‰ Ù…Ø³ØªØ®Ø¯Ù…ÙŠÙ†. ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
    else:
        db.session.delete(r)
        db.session.commit()
        flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø¯ÙˆØ±.", "warning")
    return redirect(url_for("masterdata.roles_list"))


# -------------------------------------------------
# Committees (Lajna)
# -------------------------------------------------

@masterdata_bp.route("/committees")
@login_required
@perm_required("MASTERDATA_READ")
def committees_list():
    q = _clean(request.args.get("q"))

    qry = Committee.query.options(selectinload(Committee.assignees))
    if q:
        like = f"%{q}%"
        qry = qry.filter(
            or_(
                Committee.name_ar.ilike(like),
                Committee.name_en.ilike(like),
                Committee.code.ilike(like),
            )
        )

    committees = qry.order_by(Committee.is_active.desc(), Committee.name_ar.asc()).all()

    return render_template(
        "admin/masterdata/committees_list.html",
        committees=committees,
        q=q,
    )


@masterdata_bp.route("/committees/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def committees_new():
    if request.method == "POST":
        name_ar = _clean(request.form.get("name_ar"))
        name_en = _clean(request.form.get("name_en"))
        code = _clean(request.form.get("code"))
        notes = _clean(request.form.get("notes"))
        is_active = bool(request.form.get("is_active"))

        if not name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨", "danger")
            return redirect(url_for("masterdata.committees_new"))

        c = Committee(
            name_ar=name_ar,
            name_en=name_en,
            code=code,
            notes=notes,
            is_active=is_active,
        )
        db.session.add(c)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø§Ù„Ù„Ø¬Ù†Ø©", "success")
        return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

    return render_template(
        "admin/masterdata/committees_form.html",
        c=None,
        members=[],
        users=User.query.order_by(User.email.asc()).all(),
        roles=Role.query.order_by(Role.code.asc()).all(),
    )


@masterdata_bp.route("/committees/<int:committee_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def committees_edit(committee_id):
    c = Committee.query.get_or_404(committee_id)

    if request.method == "POST":
        c.name_ar = _clean(request.form.get("name_ar"))
        c.name_en = _clean(request.form.get("name_en"))
        c.code = _clean(request.form.get("code"))
        c.notes = _clean(request.form.get("notes"))
        c.is_active = bool(request.form.get("is_active"))

        if not c.name_ar:
            flash("Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ Ù…Ø·Ù„ÙˆØ¨", "danger")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„Ù„Ø¬Ù†Ø©", "success")
        return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

    members = (
        CommitteeAssignee.query.filter_by(committee_id=c.id)
        .order_by(CommitteeAssignee.is_active.desc(), CommitteeAssignee.member_role.asc())
        .all()
    )

    return render_template(
        "admin/masterdata/committees_form.html",
        c=c,
        members=members,
        users=User.query.order_by(User.email.asc()).all(),
        roles=Role.query.order_by(Role.code.asc()).all(),
    )


@masterdata_bp.route("/committees/<int:committee_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def committees_delete(committee_id):
    c = Committee.query.get_or_404(committee_id)

    used = (
        WorkflowTemplateStep.query.filter_by(approver_committee_id=c.id).first()
        or WorkflowTemplateParallelAssignee.query.filter_by(approver_committee_id=c.id).first()
        or WorkflowInstanceStep.query.filter_by(approver_committee_id=c.id).first()
    )

    if used:
        c.is_active = False
        db.session.commit()
        flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ù„Ø¬Ù†Ø© Ù…Ø³ØªØ®Ø¯Ù…Ø© ÙÙŠ Ù…Ø³Ø§Ø±/Ø·Ù„Ø¨. ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡Ø§ (Inactive).", "warning")
        return redirect(url_for("masterdata.committees_list"))

    CommitteeAssignee.query.filter_by(committee_id=c.id).delete()
    db.session.delete(c)
    db.session.commit()
    flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ù„Ø¬Ù†Ø©", "success")
    return redirect(url_for("masterdata.committees_list"))


@masterdata_bp.route("/committees/<int:committee_id>/members/add", methods=["POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def committee_member_add(committee_id):
    c = Committee.query.get_or_404(committee_id)

    kind = _clean(request.form.get("kind")).upper() or "USER"
    member_role = _clean(request.form.get("member_role")).upper() or "MEMBER"
    is_active = bool(request.form.get("is_active"))

    user_id = request.form.get("user_id")
    role_code = _clean(request.form.get("role"))

    if kind not in ("USER", "ROLE"):
        flash("Ù†ÙˆØ¹ Ø§Ù„Ø¹Ø¶Ùˆ ØºÙŠØ± ØµØ§Ù„Ø­", "danger")
        return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

    if member_role not in ("CHAIR", "SECRETARY", "MEMBER"):
        member_role = "MEMBER"

    # Ensure single chair/secretary (active)
    if member_role in ("CHAIR", "SECRETARY") and is_active:
        exists = CommitteeAssignee.query.filter_by(
            committee_id=c.id,
            member_role=member_role,
            is_active=True,
        ).first()
        if exists:
            flash(f"ÙŠÙˆØ¬Ø¯ Ø¨Ø§Ù„ÙØ¹Ù„ {member_role} Ù†Ø´Ø· ÙÙŠ Ù‡Ø°Ù‡ Ø§Ù„Ù„Ø¬Ù†Ø©. Ø¹Ø·Ù‘Ù„Ù‡ Ø£ÙˆÙ„Ø§Ù‹ Ø£Ùˆ ØºÙŠÙ‘Ø± Ø§Ù„Ø¯ÙˆØ±.", "warning")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

    if kind == "USER":
        try:
            user_id_int = int(user_id)
        except Exception:
            user_id_int = None
        if not user_id_int:
            flash("Ø§Ø®ØªØ± Ù…Ø³ØªØ®Ø¯Ù…", "danger")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        # prevent duplicates
        dup = CommitteeAssignee.query.filter_by(committee_id=c.id, kind="USER", user_id=user_id_int).first()
        if dup:
            flash("Ù‡Ø°Ø§ Ø§Ù„Ù…Ø³ØªØ®Ø¯Ù… Ù…ÙˆØ¬ÙˆØ¯ Ø¨Ø§Ù„ÙØ¹Ù„ Ø¶Ù…Ù† Ø§Ù„Ù„Ø¬Ù†Ø©", "warning")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        m = CommitteeAssignee(
            committee_id=c.id,
            kind="USER",
            user_id=user_id_int,
            role=None,
            member_role=member_role,
            is_active=is_active,
        )

    else:
        if not role_code:
            flash("Ø§Ø®ØªØ± Role", "danger")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        dup = CommitteeAssignee.query.filter_by(committee_id=c.id, kind="ROLE", role=role_code).first()
        if dup:
            flash("Ù‡Ø°Ø§ Ø§Ù„Ø¯ÙˆØ± Ù…ÙˆØ¬ÙˆØ¯ Ø¨Ø§Ù„ÙØ¹Ù„ Ø¶Ù…Ù† Ø§Ù„Ù„Ø¬Ù†Ø©", "warning")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        m = CommitteeAssignee(
            committee_id=c.id,
            kind="ROLE",
            user_id=None,
            role=role_code,
            member_role=member_role,
            is_active=is_active,
        )

    db.session.add(m)
    db.session.commit()
    flash("ØªÙ… Ø¥Ø¶Ø§ÙØ© Ø§Ù„Ø¹Ø¶Ùˆ", "success")
    return redirect(url_for("masterdata.committees_edit", committee_id=c.id))


@masterdata_bp.route("/committees/members/<int:member_id>/update", methods=["POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def committee_member_update(member_id):
    m = CommitteeAssignee.query.get_or_404(member_id)

    member_role = _clean(request.form.get("member_role")).upper() or "MEMBER"
    is_active = bool(request.form.get("is_active"))

    if member_role not in ("CHAIR", "SECRETARY", "MEMBER"):
        member_role = "MEMBER"

    # Ensure single chair/secretary (active)
    if member_role in ("CHAIR", "SECRETARY") and is_active:
        exists = CommitteeAssignee.query.filter(
            CommitteeAssignee.committee_id == m.committee_id,
            CommitteeAssignee.member_role == member_role,
            CommitteeAssignee.is_active == True,
            CommitteeAssignee.id != m.id,
        ).first()
        if exists:
            flash(f"ÙŠÙˆØ¬Ø¯ Ø¨Ø§Ù„ÙØ¹Ù„ {member_role} Ù†Ø´Ø· ÙÙŠ Ù‡Ø°Ù‡ Ø§Ù„Ù„Ø¬Ù†Ø©.", "warning")
            return redirect(url_for("masterdata.committees_edit", committee_id=m.committee_id))

    m.member_role = member_role
    m.is_active = is_active
    db.session.commit()
    flash("ØªÙ… ØªØ­Ø¯ÙŠØ« Ø§Ù„Ø¹Ø¶Ùˆ", "success")
    return redirect(url_for("masterdata.committees_edit", committee_id=m.committee_id))


@masterdata_bp.route("/committees/members/<int:member_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def committee_member_delete(member_id):
    m = CommitteeAssignee.query.get_or_404(member_id)
    committee_id = m.committee_id
    db.session.delete(m)
    db.session.commit()
    flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø¹Ø¶Ùˆ", "success")
    return redirect(url_for("masterdata.committees_edit", committee_id=committee_id))


# =========================
# Excel exports + Excel imports (Masterdata)
# =========================

@masterdata_bp.route("/roles/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def roles_export_excel():
    q = (request.args.get("q") or "").strip()
    query = Role.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Role.code.ilike(like),
            Role.name_ar.ilike(like),
            Role.name_en.ilike(like),
        ))

    items = query.order_by(Role.code.asc()).all()

    headers = ["ID", "Code", "Ø§Ù„Ø§Ø³Ù… (AR)", "Name (EN)", "Ù†Ø´Ø·"]
    rows = [(r.id, r.code, r.name_ar or "", r.name_en or "", "Ù†Ø¹Ù…" if r.is_active else "Ù„Ø§") for r in items]

    xbytes = make_xlsx_bytes("Roles", headers, rows)
    filename = f"roles_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(BytesIO(xbytes), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)


@masterdata_bp.route("/request-types/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def request_types_export_excel():
    q = (request.args.get("q") or "").strip()
    query = RequestType.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            RequestType.code.ilike(like),
            RequestType.name_ar.ilike(like),
            RequestType.name_en.ilike(like),
        ))

    items = query.order_by(RequestType.id.asc()).all()

    headers = ["ID", "CODE", "Ø§Ù„Ø§Ø³Ù… (AR)", "Name (EN)", "Ù†Ø´Ø·"]
    rows = [(rt.id, rt.code, rt.name_ar or "", rt.name_en or "", "Ù†Ø¹Ù…" if rt.is_active else "Ù„Ø§") for rt in items]

    xbytes = make_xlsx_bytes("RequestTypes", headers, rows)
    filename = f"request_types_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(BytesIO(xbytes), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)




@masterdata_bp.route("/request-types/import-excel", methods=["POST"])
@login_required
@perm_required("REQUEST_TYPES_UPDATE")
def request_types_import_excel():
    """Import Request Types from Excel.

    Modes:
      - safe: upsert by code, do not delete
      - replace: try delete-all then insert; if FK prevents deletion, soft-fallback by deactivating all then upsert

    Expected columns (tolerant):
      - code (required)
      - name_ar (required)
      - name_en (optional)
      - is_active / active / Ù†Ø´Ø· (optional)
    """
    mode = (request.form.get("mode") or "safe").strip().lower()
    file_storage = request.files.get("file")
    if not file_storage:
        flash("ÙŠØ±Ø¬Ù‰ Ø§Ø®ØªÙŠØ§Ø± Ù…Ù„Ù Excel (.xlsx).", "danger")
        return redirect(url_for("masterdata.request_types_list"))

    try:
        _title, rows, headers = read_excel_rows(file_storage)
    except Exception as e:
        flash(f"ØªØ¹Ø°Ø± Ù‚Ø±Ø§Ø¡Ø© Ù…Ù„Ù Excel: {e}", "danger")
        return redirect(url_for("masterdata.request_types_list"))

    if not rows:
        flash("Ù…Ù„Ù Excel ÙØ§Ø±Øº Ø£Ùˆ Ù„Ø§ ÙŠØ­ØªÙˆÙŠ ØµÙÙˆÙ Ø¨ÙŠØ§Ù†Ø§Øª.", "warning")
        return redirect(url_for("masterdata.request_types_list"))

    def _code(r):
        v = pick(r, "code", "CODE", "requesttypecode", "request_type_code", "Ø§Ù„ÙƒÙˆØ¯", "ÙƒÙˆØ¯")
        s = to_str(v)
        return (s or "").upper()

    def _vals(r):
        name_ar = to_str(pick(r, "name_ar", "namear", "Ø§Ù„Ø§Ø³Ù…(ar)", "Ø§Ù„Ø§Ø³Ù… (ar)", "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ", "Ø§Ù„Ø§Ø³Ù…"))
        name_en = to_str(pick(r, "name_en", "nameen", "name(en)", "Ø§Ù„Ø§Ø³Ù…(en)", "Ø§Ù„Ø§Ø³Ù… (en)", "English", "EN"))
        is_active = to_bool(pick(r, "is_active", "active", "Ù†Ø´Ø·", "ÙØ¹Ø§Ù„"), default=True)
        out = {"name_ar": name_ar or "", "name_en": name_en, "is_active": bool(is_active)}
        return out

    # Validate at least one row has required fields
    any_ok = False
    for rr in rows[:50]:
        if _code(rr) and to_str(pick(rr, "name_ar", "namear", "Ø§Ù„Ø§Ø³Ù…(ar)", "Ø§Ù„Ø§Ø³Ù… Ø§Ù„Ø¹Ø±Ø¨ÙŠ", "Ø§Ù„Ø§Ø³Ù…")):
            any_ok = True
            break
    if not any_ok:
        flash("Ø£Ø¹Ù…Ø¯Ø© Ù…Ù„Ù Excel ØºÙŠØ± ØµØ­ÙŠØ­Ø©. Ø§Ù„Ù…Ø·Ù„ÙˆØ¨: code + name_ar Ø¹Ù„Ù‰ Ø§Ù„Ø£Ù‚Ù„.", "danger")
        return redirect(url_for("masterdata.request_types_list"))

    def _insert():
        created, updated = upsert_by_code(
            db.session,
            RequestType,
            rows,
            code_getter=_code,
            values_getter=_vals,
            normalize_code=lambda s: (s or "").strip().upper(),
        )
        return created, updated

    created = updated = 0
    used_soft = False

    try:
        if mode == "replace":
            def _soft():
                RequestType.query.update({RequestType.is_active: False}, synchronize_session=False)
                db.session.flush()
                return _insert()

            created, updated, used_soft = replace_all(
                db.session,
                RequestType.query,
                _insert,
                soft_fallback=_soft,
            )
        else:
            created, updated = _insert()

        db.session.commit()

        msg = f"ØªÙ… Ø§Ø³ØªÙŠØ±Ø§Ø¯ Ø£Ù†ÙˆØ§Ø¹ Ø§Ù„Ø·Ù„Ø¨Ø§Øª: Ø¥Ø¶Ø§ÙØ© {created}ØŒ ØªØ­Ø¯ÙŠØ« {updated}."
        if mode == "replace" and used_soft:
            msg += " (ØªÙ… Ø§Ø³ØªØ®Ø¯Ø§Ù… Soft Replace Ù„Ø£Ù† Ø§Ù„Ø­Ø°Ù Ø§Ù„ÙƒØ§Ù…Ù„ ØºÙŠØ± Ù…Ù…ÙƒÙ† Ø¨Ø³Ø¨Ø¨ Ø§Ø±ØªØ¨Ø§Ø·Ø§Øª.)"
        flash(msg, "success")

    except Exception as e:
        db.session.rollback()
        flash(f"ÙØ´Ù„ Ø§Ù„Ø§Ø³ØªÙŠØ±Ø§Ø¯: {e}", "danger")

    return redirect(url_for("masterdata.request_types_list"))
@masterdata_bp.route("/committees/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def committees_export_excel():
    q = (request.args.get("q") or "").strip()
    query = Committee.query.options(selectinload(Committee.assignees))

    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            Committee.name_ar.ilike(like),
            Committee.name_en.ilike(like),
            Committee.code.ilike(like),
        ))

    committees = query.order_by(Committee.id.desc()).all()

    # Round-trip friendly export (can be imported back with chair/members)
    headers = [
        "ID",
        "Code",
        "Ø§Ù„Ø§Ø³Ù… (AR)",
        "Name (EN)",
        "Active",
        "Notes",
        "Chair Email",
        "Secretary Email",
        "Members Emails",
        "Roles Codes",
        "Members (Active)",
        "Members (Total)",
    ]

    rows = []
    for c in committees:
        assignees = list(c.assignees or [])
        active_assignees = [m for m in assignees if m.is_active]

        def _role(m):
            return (m.member_role or "").strip().upper()

        chair_email = ""
        secretary_email = ""
        for m in active_assignees:
            if m.kind == "USER" and m.user and m.user.email:
                if _role(m) == "CHAIR" and not chair_email:
                    chair_email = m.user.email
                elif _role(m) == "SECRETARY" and not secretary_email:
                    secretary_email = m.user.email

        members_emails = []
        for m in active_assignees:
            if m.kind == "USER" and m.user and m.user.email:
                if _role(m) in ("CHAIR", "SECRETARY"):
                    continue
                members_emails.append(m.user.email)

        roles_codes = [m.role for m in active_assignees if m.kind == "ROLE" and m.role]

        total = len(assignees)
        active_cnt = len(active_assignees)

        rows.append((
            c.id,
            c.code or "",
            c.name_ar or "",
            c.name_en or "",
            "Ù†Ø¹Ù…" if c.is_active else "Ù„Ø§",
            (c.notes or ""),
            chair_email,
            secretary_email,
            "; ".join(members_emails),
            "; ".join(roles_codes),
            active_cnt,
            total,
        ))

    xbytes = make_xlsx_bytes("Committees", headers, rows)
    filename = f"committees_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        BytesIO(xbytes),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )



@masterdata_bp.route("/committees/<int:committee_id>/members/export.xlsx")
@login_required
@perm_required("MASTERDATA_READ")
def committee_members_export_excel(committee_id):
    c = Committee.query.options(selectinload(Committee.assignees)).get_or_404(committee_id)

    members = (
        CommitteeAssignee.query
        .filter_by(committee_id=c.id)
        .order_by(CommitteeAssignee.is_active.desc(), CommitteeAssignee.member_role.asc(), CommitteeAssignee.id.asc())
        .all()
    )

    headers = [
        "Member ID",
        "Committee ID",
        "Committee Code",
        "Committee Name (AR)",
        "Kind",
        "User Email",
        "Role Code",
        "Member Role",
        "Active",
    ]

    rows = []
    for m in members:
        user_email = m.user.email if (m.kind == 'USER' and m.user) else ''
        rows.append((
            m.id,
            c.id,
            c.code or "",
            c.name_ar or "",
            (m.kind or ""),
            user_email,
            m.role or "",
            m.member_role or "",
            "Ù†Ø¹Ù…" if m.is_active else "Ù„Ø§",
        ))

    xbytes = make_xlsx_bytes("CommitteeMembers", headers, rows)
    filename = f"committee_{c.id}_members_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(BytesIO(xbytes), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=filename)


# -------------------------
# Excel Imports (UI upload)
# -------------------------

def _import_to_bool(val):
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("Ù†Ø¹Ù…", "yes", "true", "1", "y", "on"):
        return True
    if s in ("Ù„Ø§", "no", "false", "0", "n", "off"):
        return False
    return None


def _import_norm(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def _row_has_any_key(row, *names):
    for n in names:
        if n in row:
            return True
    return False


def _split_list(val):
    """Split a cell containing a list into values (comma/semicolon/newline + Arabic comma)."""
    if val is None:
        return []
    s = str(val)
    s = s.replace("ØŒ", ",")
    # Normalize Windows newlines inside cells
    s = s.replace("\r", "\n")

    parts = []
    for chunk in s.split("\n"):
        chunk = chunk.replace(";", ",")
        for p in chunk.split(","):
            pp = str(p).strip()
            if pp:
                parts.append(pp)

    # de-dup (case-insensitive) while preserving order
    seen = set()
    out = []
    for p in parts:
        k = p.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(p)
    return out


def _read_excel_rows_from_filestorage(file_storage):
    # Import here to avoid hard dependency when not used
    from openpyxl import load_workbook

    wb = load_workbook(file_storage, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip() if cell.value is not None else "")

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or str(v).strip() == "" for v in r):
            continue
        row = {}
        for h, v in zip(headers, r):
            if h:
                row[h] = v
        rows.append(row)

    return ws.title, rows


def _hdr(row, *names):
    for n in names:
        if n in row:
            return row.get(n)
    return None


def _upsert(model, where: dict, values: dict):
    q = db.session.query(model)
    for k, v in where.items():
        q = q.filter(getattr(model, k) == v)
    obj = q.first()
    if not obj:
        obj = model(**{**where, **values})
        db.session.add(obj)
    else:
        for k, v in values.items():
            setattr(obj, k, v)
    return obj


def _import_organizations(rows):
    created = updated = 0
    for r in rows:
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (AR)", "Name (AR)", "name_ar"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (EN)", "Name (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (code or name_ar):
            continue
        where = {"code": code} if code else {"name_ar": name_ar}
        values = {
            "name_ar": name_ar,
            "name_en": name_en,
        }
        if code is not None:
            values["code"] = code
        if is_active is not None:
            values["is_active"] = is_active

        obj = Organization.query.filter_by(**where).first()
        if obj:
            for k, v in values.items():
                setattr(obj, k, v)
            updated += 1
        else:
            db.session.add(Organization(**{**where, **values}))
            created += 1
    return created, updated


def _import_directorates(rows):
    created = updated = 0
    for r in rows:
        org_name = _import_norm(_hdr(r, "Ø§Ù„Ù…Ù†Ø¸Ù…Ø©", "Organization", "organization"))
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (AR)", "Name (AR)", "name_ar"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (EN)", "Name (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (org_name and name_ar):
            continue

        org = Organization.query.filter_by(name_ar=org_name).first() or Organization.query.filter_by(code=org_name).first()
        if not org:
            # create org automatically if not exists
            org = Organization(name_ar=org_name, name_en=None, code=None, is_active=True)
            db.session.add(org)
            db.session.flush()

        where = {"organization_id": org.id, "name_ar": name_ar}
        obj = Directorate.query.filter_by(**where).first()
        if obj:
            obj.name_en = name_en
            obj.code = code
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            db.session.add(Directorate(
                organization_id=org.id,
                name_ar=name_ar,
                name_en=name_en,
                code=code,
                is_active=True if is_active is None else is_active,
            ))
            created += 1
    return created, updated


def _import_departments(rows):
    created = updated = 0
    for r in rows:
        dir_name = _import_norm(_hdr(r, "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "Directorate", "directorate"))
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (AR)", "Ø§Ù„Ø§Ø³Ù… (AR)", "Department (AR)", "name_ar"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (EN)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Department (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (dir_name and name_ar):
            continue

        d = Directorate.query.filter_by(name_ar=dir_name).first() or Directorate.query.filter_by(code=dir_name).first()
        if not d:
            continue

        where = {"directorate_id": d.id, "name_ar": name_ar}
        obj = Department.query.filter_by(**where).first()
        if obj:
            obj.name_en = name_en
            obj.code = code
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            db.session.add(Department(
                directorate_id=d.id,
                name_ar=name_ar,
                name_en=name_en,
                code=code,
                is_active=True if is_active is None else is_active,
            ))
            created += 1
    return created, updated


def _import_sections(rows):
    created = updated = 0
    for r in rows:
        dir_name = _import_norm(_hdr(r, "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "Directorate", "directorate"))
        dep_name = _import_norm(_hdr(r, "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©", "Department", "department"))
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ù‚Ø³Ù… (AR)", "Ø§Ù„Ø§Ø³Ù… (AR)", "Section (AR)", "name_ar"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ù‚Ø³Ù… (EN)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Section (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (dir_name and dep_name and name_ar):
            continue

        d = Directorate.query.filter_by(name_ar=dir_name).first() or Directorate.query.filter_by(code=dir_name).first()
        if not d:
            continue
        dep = Department.query.filter_by(name_ar=dep_name, directorate_id=d.id).first() or Department.query.filter_by(code=dep_name, directorate_id=d.id).first()
        if not dep:
            continue

        where = {"department_id": dep.id, "name_ar": name_ar}
        obj = Section.query.filter_by(**where).first()
        if obj:
            obj.name_en = name_en
            obj.code = code
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            db.session.add(Section(
                department_id=dep.id,
                name_ar=name_ar,
                name_en=name_en,
                code=code,
                is_active=True if is_active is None else is_active,
            ))
            created += 1
    return created, updated





def _import_divisions(rows):
    created = updated = 0

    def _find_directorate(val):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        if s.isdigit():
            return Directorate.query.get(int(s))
        return (Directorate.query.filter_by(code=s).first()
                or Directorate.query.filter_by(name_ar=s).first()
                or Directorate.query.filter_by(name_en=s).first())

    def _find_department(val, directorate=None):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        if s.isdigit():
            dep = Department.query.get(int(s))
            if dep and directorate and getattr(dep, "directorate_id", None) != directorate.id:
                return None
            return dep

        q = Department.query
        if directorate:
            q = q.filter(Department.directorate_id == directorate.id)
        return (q.filter(Department.code == s).first()
                or q.filter(Department.name_ar == s).first()
                or q.filter(Department.name_en == s).first())

    def _find_section(val, department=None, directorate=None):
        if not val:
            return None
        s = str(val).strip()
        if not s:
            return None
        if s.isdigit():
            sec = Section.query.get(int(s))
            if sec and department and getattr(sec, "department_id", None) != department.id:
                return None
            if sec and directorate and getattr(sec, "directorate_id", None) not in (None, directorate.id):
                # if section is directly under directorate, enforce match
                if getattr(sec, "directorate_id", None) != directorate.id:
                    return None
            return sec

        q = Section.query
        if department:
            q = q.filter(Section.department_id == department.id)
        elif directorate:
            # sections may be directly under a directorate in this system
            q = q.filter(Section.directorate_id == directorate.id)

        return (q.filter(Section.code == s).first()
                or q.filter(Section.name_ar == s).first()
                or q.filter(Section.name_en == s).first())

    for r in rows:
        # Accept both Arabic and English headers (also matches the export headers)
        dir_val = _import_norm(_hdr(r,
            "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "Directorate", "directorate",
            "Directorate Code", "directorate_code",
            "Directorate ID", "directorate_id"
        ))
        dep_val = _import_norm(_hdr(r,
            "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©", "Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©/Ø§Ù„Ù…ÙƒØªØ¨", "Department", "department",
            "Department/Office",
            "Department Code", "department_code",
            "Department ID", "department_id"
        ))
        sec_val = _import_norm(_hdr(r,
            "Ø§Ù„Ù‚Ø³Ù…", "Section", "section",
            "Section Code", "section_code",
            "Section ID", "section_id"
        ))

        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø´Ø¹Ø¨Ø© (AR)", "Division (AR)", "Ø§Ù„Ø§Ø³Ù… (AR)", "Name (AR)", "name_ar"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ø´Ø¹Ø¨Ø© (EN)", "Division (EN)", "Ø§Ù„Ø§Ø³Ù… (EN)", "Name (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if code:
            code = str(code).strip().upper()
        if not (name_ar or code):
            continue

        directorate = _find_directorate(dir_val)

        # Resolve parent (either section OR department)
        section = None
        department = None

        if sec_val:
            # try resolve section with department/directorate context if provided
            # if department is provided, resolve it first for better matching
            if dep_val:
                department = _find_department(dep_val, directorate=directorate) or _find_department(dep_val, directorate=None)
            section = _find_section(sec_val, department=department, directorate=directorate) or _find_section(sec_val, department=None, directorate=None)

        if not section:
            # fallback to department-only parent
            if dep_val:
                department = _find_department(dep_val, directorate=directorate) or _find_department(dep_val, directorate=None)

        if not section and not department:
            # parent is required by DB constraint
            continue

        # prefer section parent if found
        section_id = section.id if section else None
        department_id = None if section_id else (department.id if department else None)

        if code:
            obj = Division.query.filter_by(code=code).first()
        else:
            if section_id:
                obj = Division.query.filter_by(section_id=section_id, name_ar=name_ar).first()
            else:
                obj = Division.query.filter_by(department_id=department_id, name_ar=name_ar).first()

        if obj:
            # update
            if name_ar:
                obj.name_ar = name_ar
            obj.name_en = name_en
            obj.code = code or obj.code
            if is_active is not None:
                obj.is_active = is_active
            obj.section_id = section_id
            obj.department_id = department_id
            updated += 1
        else:
            db.session.add(Division(
                section_id=section_id,
                department_id=department_id,
                name_ar=name_ar or (code or ""),
                name_en=name_en,
                code=code,
                is_active=True if is_active is None else is_active,
            ))
            created += 1

    return created, updated


def _import_roles(rows):
    created = updated = 0
    for r in rows:
        code = _import_norm(_hdr(r, "Code", "code", "CODE"))
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (AR)", "name_ar", "Name (AR)", "Ø§Ù„Ø§Ø³Ù… (Ø¹Ø±Ø¨ÙŠ)"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (EN)", "name_en", "Name (EN)"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not code:
            continue

        obj = Role.query.filter_by(code=code).first()
        if obj:
            obj.name_ar = name_ar
            obj.name_en = name_en
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            db.session.add(Role(
                code=code,
                name_ar=name_ar,
                name_en=name_en,
                is_active=True if is_active is None else is_active,
            ))
            created += 1
    return created, updated


def _import_committees(rows):
    created = updated = 0

    # Cache lookups for membership resolution (by email/role code)
    users_by_email = {u.email.lower(): u for u in User.query.all() if u.email}
    roles_by_code = {r.code.lower(): r for r in Role.query.all() if r.code}

    CHAIR_HDRS = (
        "Chair Email", "Chair (Email)", "Chair", "Ø±Ø¦ÙŠØ³ Ø§Ù„Ù„Ø¬Ù†Ø©", "Ø±Ø¦ÙŠØ³ (Email)", "Ø±Ø¦ÙŠØ³ Ø§Ù„Ù„Ø¬Ù†Ø© (Email)",
    )
    SEC_HDRS = (
        "Secretary Email", "Secretary (Email)", "Secretary", "Ù…Ù‚Ø±Ø± Ø§Ù„Ù„Ø¬Ù†Ø©", "Ù…Ù‚Ø±Ø± (Email)", "Ù…Ù‚Ø±Ø± Ø§Ù„Ù„Ø¬Ù†Ø© (Email)",
    )
    MEM_HDRS = (
        "Members Emails", "Members (Emails)", "Members", "Ø§Ù„Ø£Ø¹Ø¶Ø§Ø¡", "Ø£Ø¹Ø¶Ø§Ø¡ Ø§Ù„Ù„Ø¬Ù†Ø©",
    )
    ROLE_HDRS = (
        "Roles Codes", "Roles (Codes)", "Roles", "Ø§Ù„Ø£Ø¯ÙˆØ§Ø±", "Ø£Ø¯ÙˆØ§Ø±",
    )

    def _sync_members(c: Committee, row: dict):
        # Only touch membership if the sheet contains membership columns
        has_members_cols = _row_has_any_key(row, *(CHAIR_HDRS + SEC_HDRS + MEM_HDRS + ROLE_HDRS))
        if not has_members_cols:
            return

        chair_vals = _split_list(_hdr(row, *CHAIR_HDRS))
        sec_vals = _split_list(_hdr(row, *SEC_HDRS))
        mem_vals = _split_list(_hdr(row, *MEM_HDRS))
        role_vals = _split_list(_hdr(row, *ROLE_HDRS))

        desired = {}  # (kind, id/code) -> member_role

        def add_user(email: str, member_role: str):
            if not email:
                return
            u = users_by_email.get(email.lower())
            if not u:
                return
            desired[("USER", u.id)] = member_role

        def add_role(code: str, member_role: str):
            if not code:
                return
            rr = roles_by_code.get(code.lower())
            if not rr:
                return
            desired[("ROLE", rr.code)] = member_role

        chair_email = chair_vals[0] if chair_vals else None
        if chair_email:
            add_user(chair_email, "CHAIR")
        for extra in chair_vals[1:]:
            add_user(extra, "MEMBER")

        sec_email = sec_vals[0] if sec_vals else None
        if sec_email:
            add_user(sec_email, "SECRETARY")
        for extra in sec_vals[1:]:
            add_user(extra, "MEMBER")

        skip = {e.lower() for e in (chair_email, sec_email) if e}
        for email in mem_vals:
            if email.lower() in skip:
                continue
            add_user(email, "MEMBER")

        for rc in role_vals:
            add_role(rc, "MEMBER")

        existing = CommitteeAssignee.query.filter_by(committee_id=c.id).all()
        ex_map = {}
        for m in existing:
            if m.kind == "USER":
                ex_map[("USER", m.user_id)] = m
            else:
                ex_map[("ROLE", (m.role or ""))] = m

        # Deactivate removed, upsert desired
        for k, m in ex_map.items():
            if k in desired:
                m.is_active = True
                m.member_role = desired[k]
            else:
                m.is_active = False

        for k, role in desired.items():
            if k in ex_map:
                continue
            if k[0] == "USER":
                _, uid = k
                db.session.add(CommitteeAssignee(
                    committee_id=c.id,
                    kind="USER",
                    user_id=uid,
                    role=None,
                    member_role=role,
                    is_active=True,
                ))
            else:
                _, rcode = k
                db.session.add(CommitteeAssignee(
                    committee_id=c.id,
                    kind="ROLE",
                    user_id=None,
                    role=rcode,
                    member_role=role,
                    is_active=True,
                ))

    for r in rows:
        name_ar = _import_norm(_hdr(r, "Ø§Ù„Ø§Ø³Ù… (AR)", "name_ar", "Name (AR)"))
        name_en = _import_norm(_hdr(r, "Name (EN)", "Ø§Ù„Ø§Ø³Ù… (EN)", "name_en"))
        code = _import_norm(_hdr(r, "Code", "code"))
        notes = _import_norm(_hdr(r, "Ù…Ù„Ø§Ø­Ø¸Ø§Øª", "Notes", "notes"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (code or name_ar):
            continue

        if code:
            obj = Committee.query.filter_by(code=code).first()
        else:
            obj = Committee.query.filter_by(name_ar=name_ar).first()

        if obj:
            obj.name_ar = name_ar or obj.name_ar
            obj.name_en = name_en
            obj.code = code
            obj.notes = notes
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            obj = Committee(
                name_ar=name_ar,
                name_en=name_en,
                code=code,
                notes=notes,
                is_active=True if is_active is None else is_active,
            )
            db.session.add(obj)
            db.session.flush()  # ensure obj.id for membership upserts
            created += 1

        # Optional: import chair/members when columns exist
        _sync_members(obj, r)

    return created, updated



def _import_units(rows):
    created = 0
    updated = 0

    # Map organizations by name/code
    orgs = Organization.query.all()
    org_by_name = {(o.name_ar or '').strip(): o for o in orgs if o.name_ar}
    org_by_code = {(o.code or '').strip(): o for o in orgs if o.code}

    for r in rows:
        org_name = _import_norm(_hdr(r, "Ø§Ù„Ù…Ù†Ø¸Ù…Ø©", "Organization", "organization", "org"))
        # Backward compatibility: some old sheets used "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©"
        if not org_name:
            org_name = _import_norm(_hdr(r, "Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©", "Directorate", "directorate"))

        name_ar = _import_norm(_hdr(r, "Ø§Ù„ÙˆØ­Ø¯Ø© (AR)", "Ø§Ù„ÙˆØ­Ø¯Ø©", "Unit (AR)", "name_ar", "Ø§Ù„Ø§Ø³Ù… (AR)"))
        name_en = _import_norm(_hdr(r, "Ø§Ù„ÙˆØ­Ø¯Ø© (EN)", "Unit (EN)", "name_en", "Ø§Ù„Ø§Ø³Ù… (EN)"))
        code = _import_norm(_hdr(r, "Code", "code"))
        is_active = _import_to_bool(_hdr(r, "Ù†Ø´Ø·", "Active", "is_active"))

        if not (org_name and name_ar):
            continue

        o = org_by_name.get(org_name) or org_by_code.get(org_name)
        if not o:
            # skip unknown organization
            continue

        obj = None
        if code:
            obj = Unit.query.filter_by(code=code).first()
        if not obj:
            obj = Unit.query.filter_by(organization_id=o.id, name_ar=name_ar).first()

        if obj:
            obj.organization_id = o.id
            obj.name_ar = name_ar or obj.name_ar
            obj.name_en = name_en
            obj.code = code
            if is_active is not None:
                obj.is_active = is_active
            updated += 1
        else:
            db.session.add(Unit(
                organization_id=o.id,
                name_ar=name_ar,
                name_en=name_en,
                code=code,
                is_active=True if is_active is None else is_active,
            ))
            created += 1
    return created, updated

IMPORT_SCREENS = {
    "organizations": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ù…Ù†Ø¸Ù…Ø§Øª",
        "back": "masterdata.org_list",
        "importer": _import_organizations,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ø§Ø³Ù… (AR), Ø§Ù„Ø§Ø³Ù… (EN), Code, Ù†Ø´Ø·",
    },
    "directorates": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ø¥Ø¯Ø§Ø±Ø§Øª",
        "back": "masterdata.dir_list",
        "importer": _import_directorates,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ù…Ù†Ø¸Ù…Ø©, Ø§Ù„Ø§Ø³Ù… (AR), Ø§Ù„Ø§Ø³Ù… (EN), Code, Ù†Ø´Ø·",
    },
    "units": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„ÙˆØ­Ø¯Ø§Øª",
        "back": "masterdata.units_list",
        "importer": _import_units,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ù…Ù†Ø¸Ù…Ø©, Ø§Ù„ÙˆØ­Ø¯Ø© (AR), Ø§Ù„ÙˆØ­Ø¯Ø© (EN), Code, Ù†Ø´Ø·",
    },
    "departments": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ø¯ÙˆØ§Ø¦Ø±",
        "back": "masterdata.dept_list",
        "importer": _import_departments,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©, Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (AR), Ø§Ù„Ø¯Ø§Ø¦Ø±Ø© (EN), Code, Ù†Ø´Ø·",
    },
    "sections": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ø£Ù‚Ø³Ø§Ù…",
        "back": "masterdata.sections_list",
        "importer": _import_sections,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ø¥Ø¯Ø§Ø±Ø©, Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©, Ø§Ù„Ù‚Ø³Ù… (AR), Ø§Ù„Ù‚Ø³Ù… (EN), Code, Ù†Ø´Ø·",
    },
"divisions": {
    "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ø´ÙØ¹Ø¨",
    "back": "masterdata.divisions_list",
    "importer": _import_divisions,
    "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ø¥Ø¯Ø§Ø±Ø© (Ø§Ø®ØªÙŠØ§Ø±ÙŠ), Ø§Ù„Ø¯Ø§Ø¦Ø±Ø©/Ø§Ù„Ù…ÙƒØªØ¨ Ø£Ùˆ Ø§Ù„Ù‚Ø³Ù… (Ø£Ø­Ø¯Ù‡Ù…Ø§ Ù…Ø·Ù„ÙˆØ¨), Ø§Ù„Ø´Ø¹Ø¨Ø© (AR), Ø§Ù„Ø´Ø¹Ø¨Ø© (EN), Code, Ù†Ø´Ø·",
},
    "roles": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ø£Ø¯ÙˆØ§Ø±",
        "back": "masterdata.roles_list",
        "importer": _import_roles,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Code, Ø§Ù„Ø§Ø³Ù… (AR), Ø§Ù„Ø§Ø³Ù… (EN), Ù†Ø´Ø·",
    },
    "committees": {
        "title": "Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø§Ù„Ù„Ø¬Ø§Ù†",
        "back": "masterdata.committees_list",
        "importer": _import_committees,
        "hint": "Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Ø§Ù„Ø§Ø³Ù… (AR), Name (EN), Code, Ù†Ø´Ø·, Ù…Ù„Ø§Ø­Ø¸Ø§Øª (+ Ø§Ø®ØªÙŠØ§Ø±ÙŠ: Chair Email, Secretary Email, Members Emails, Roles Codes)",
    },
}


@masterdata_bp.route("/import-excel/<string:screen>", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def import_excel(screen: str):
    screen = (screen or "").strip().lower()
    cfg = IMPORT_SCREENS.get(screen)
    if not cfg:
        abort(404)

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not getattr(f, "filename", ""):
            flash("Ø§Ø®ØªØ± Ù…Ù„Ù Excel (.xlsx)", "danger")
            return redirect(url_for("masterdata.import_excel", screen=screen))

        try:
            _sheet, rows = _read_excel_rows_from_filestorage(f)
            created, updated = cfg["importer"](rows)
            db.session.commit()
            flash(f"âœ… ØªÙ… Ø§Ù„Ø§Ø³ØªÙŠØ±Ø§Ø¯ Ø¨Ù†Ø¬Ø§Ø­. (Ø¬Ø¯ÙŠØ¯: {created} / ØªØ­Ø¯ÙŠØ«: {updated})", "success")
            return redirect(url_for(cfg["back"]))
        except Exception as e:
            db.session.rollback()
            flash(f"Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ù„Ø§Ø³ØªÙŠØ±Ø§Ø¯: {e}", "danger")
            return redirect(url_for("masterdata.import_excel", screen=screen))

    return render_template(
        "admin/masterdata/import_excel.html",
        title=cfg["title"],
        hint=cfg.get("hint"),
        back_url=url_for(cfg["back"]),
        screen=screen,
    )


@masterdata_bp.route("/committees/<int:committee_id>/members/import", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def committee_members_import(committee_id: int):
    c = Committee.query.get_or_404(committee_id)

    def _member_role_from(val):
        if val is None:
            return "MEMBER"
        s = str(val).strip().upper()
        if s in ("CHAIR", "Ø±Ø¦ÙŠØ³", "Ø±Ø¦ÙŠØ³ Ø§Ù„Ù„Ø¬Ù†Ø©"):
            return "CHAIR"
        if s in ("SECRETARY", "Ù…Ù‚Ø±Ø±", "Ù…Ù‚Ø±Ø± Ø§Ù„Ù„Ø¬Ù†Ø©"):
            return "SECRETARY"
        return "MEMBER"

    if request.method == "POST":
        f = request.files.get("file")
        if not f or not getattr(f, "filename", ""):
            flash("Ø§Ø®ØªØ± Ù…Ù„Ù Excel (.xlsx)", "danger")
            return redirect(url_for("masterdata.committee_members_import", committee_id=c.id))

        try:
            _sheet, rows = _read_excel_rows_from_filestorage(f)

            created = updated = skipped = 0

            # Cache lookups
            users_by_email = {u.email.lower(): u for u in User.query.all() if u.email}
            roles_by_code = {r.code.lower(): r for r in Role.query.all() if r.code}

            for r in rows:
                kind = _import_norm(_hdr(r, "Kind", "kind"))
                kind = (kind or "USER").strip().upper()

                member_role = _member_role_from(_hdr(r, "Member Role", "member_role", "Role Ø¯Ø§Ø®Ù„ Ø§Ù„Ù„Ø¬Ù†Ø©"))
                is_active = _import_to_bool(_hdr(r, "Active", "Ù†Ø´Ø·", "is_active"))
                if is_active is None:
                    is_active = True

                if kind not in ("USER", "ROLE"):
                    skipped += 1
                    continue

                # Ensure single chair/secretary if active
                if member_role in ("CHAIR", "SECRETARY") and is_active:
                    exists = CommitteeAssignee.query.filter_by(
                        committee_id=c.id,
                        member_role=member_role,
                        is_active=True,
                    ).first()
                    if exists:
                        # keep existing, skip this row
                        skipped += 1
                        continue

                if kind == "USER":
                    email = _import_norm(_hdr(r, "User Email", "Email", "user_email", "Ø§Ù„Ø¹Ø¶Ùˆ"))
                    if not email:
                        skipped += 1
                        continue
                    u = users_by_email.get(email.lower())
                    if not u:
                        skipped += 1
                        continue

                    obj = CommitteeAssignee.query.filter_by(committee_id=c.id, kind="USER", user_id=u.id).first()
                    if obj:
                        obj.member_role = member_role
                        obj.is_active = is_active
                        updated += 1
                    else:
                        db.session.add(CommitteeAssignee(
                            committee_id=c.id,
                            kind="USER",
                            user_id=u.id,
                            role=None,
                            member_role=member_role,
                            is_active=is_active,
                        ))
                        created += 1

                else:
                    role_code = _import_norm(_hdr(r, "Role Code", "role", "Role", "Ø§Ù„Ø¯ÙˆØ±"))
                    if not role_code:
                        skipped += 1
                        continue
                    rr = roles_by_code.get(role_code.lower())
                    if not rr:
                        skipped += 1
                        continue

                    obj = CommitteeAssignee.query.filter_by(committee_id=c.id, kind="ROLE", role=rr.code).first()
                    if obj:
                        obj.member_role = member_role
                        obj.is_active = is_active
                        updated += 1
                    else:
                        db.session.add(CommitteeAssignee(
                            committee_id=c.id,
                            kind="ROLE",
                            user_id=None,
                            role=rr.code,
                            member_role=member_role,
                            is_active=is_active,
                        ))
                        created += 1

            db.session.commit()
            flash(f"âœ… ØªÙ… Ø§Ø³ØªÙŠØ±Ø§Ø¯ Ø£Ø¹Ø¶Ø§Ø¡ Ø§Ù„Ù„Ø¬Ù†Ø©. (Ø¬Ø¯ÙŠØ¯: {created} / ØªØ­Ø¯ÙŠØ«: {updated} / ØªÙ… ØªØ¬Ø§Ù‡Ù„: {skipped})", "success")
            return redirect(url_for("masterdata.committees_edit", committee_id=c.id))

        except Exception as e:
            db.session.rollback()
            flash(f"Ø®Ø·Ø£ Ø£Ø«Ù†Ø§Ø¡ Ø§Ø³ØªÙŠØ±Ø§Ø¯ Ø£Ø¹Ø¶Ø§Ø¡ Ø§Ù„Ù„Ø¬Ù†Ø©: {e}", "danger")
            return redirect(url_for("masterdata.committee_members_import", committee_id=c.id))

    return render_template(
        "admin/masterdata/import_excel.html",
        title=f"Ø§Ø³ØªÙŠØ±Ø§Ø¯ Excel: Ø£Ø¹Ø¶Ø§Ø¡ Ø§Ù„Ù„Ø¬Ù†Ø© â€” {c.name_ar}",
        hint="Ø§Ù„Ø£Ø¹Ù…Ø¯Ø© Ø§Ù„Ù…ØªÙˆÙ‚Ø¹Ø©: Kind (USER/ROLE), User Email Ø£Ùˆ Role Code, Member Role (CHAIR/SECRETARY/MEMBER), Active",
        back_url=url_for("masterdata.committees_edit", committee_id=c.id),
        screen=f"committee_members_{c.id}",
    )





@masterdata_bp.route("/org-dynamic/toggle-legacy-lock", methods=["POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def org_dynamic_toggle_legacy_lock():
    locked = (request.form.get("locked") or "0").strip() in ("1", "true", "yes")
    _set_setting("ORG_LEGACY_LOCKED", "1" if locked else "0")
    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        flash("ØªØ¹Ø°Ø± Ø­ÙØ¸ Ø§Ù„Ø¥Ø¹Ø¯Ø§Ø¯.", "danger")
        return redirect(url_for("masterdata.org_node_types_list"))

    flash("âœ… ØªÙ… Ù‚ÙÙ„ Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ø«Ø§Ø¨ØªØ©: Ø£ØµØ¨Ø­Øª Ù‚Ø±Ø§Ø¡Ø© ÙÙ‚Ø· (ØªÙ… ØªØ¹Ø·ÙŠÙ„ Ø§Ù„Ø¥Ø¶Ø§ÙØ©/Ø§Ù„ØªØ¹Ø¯ÙŠÙ„/Ø§Ù„Ø­Ø°Ù ÙÙŠ Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù‚Ø¯ÙŠÙ…Ø©). Ø§Ø³ØªØ®Ø¯Ù… Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù…ÙˆØ­Ø¯Ø© (Dynamic) Ù„Ù„ØªØ¹ÙŠÙŠÙ†Ø§Øª ÙˆØ§Ù„Ù…Ø³Ø§Ø±Ø§Øª." if locked else "âœ… ØªÙ… ÙØªØ­ Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ø«Ø§Ø¨ØªØ©: Ø£ØµØ¨Ø­ Ø¨Ø¥Ù…ÙƒØ§Ù†Ùƒ ØªØ¹Ø¯ÙŠÙ„ Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù‚Ø¯ÙŠÙ…Ø© Ù…Ø¬Ø¯Ø¯Ù‹Ø§ (ÙŠÙÙØ¶Ù‘Ù„ Ø§Ù„Ø§Ø³ØªÙ…Ø±Ø§Ø± Ø¨Ø§Ø³ØªØ®Ø¯Ø§Ù… Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù…ÙˆØ­Ø¯Ø© Ù„ØªØ¬Ù†Ù‘Ø¨ Ø§Ù„ØªØ¹Ø§Ø±Ø¶).", "success" if not locked else "warning")
    return redirect(url_for("masterdata.org_node_types_list"))


@masterdata_bp.route("/org-dynamic/sync-legacy", methods=["POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def org_dynamic_sync_legacy():
    # Seed/sync Dynamic OrgNodes from legacy org tables.
    # - Creates dynamic types if missing (ensure_dynamic_org_seed)
    # - Upserts legacy-mapped nodes (does not delete custom dynamic nodes)
    try:
        ensure_dynamic_org_seed()
        sync_legacy_now()
        _set_setting("ORG_NODE_LAST_SYNC", datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"))
        db.session.commit()
        flash("ØªÙ…Øª Ù…Ø²Ø§Ù…Ù†Ø© Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ù…ÙˆØ­Ø¯Ø© Ù…Ù† Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ© Ø§Ù„Ø«Ø§Ø¨ØªØ© Ø¨Ù†Ø¬Ø§Ø­.", "success")
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass
        flash("ØªØ¹Ø°Ø± ØªÙ†ÙÙŠØ° Ø§Ù„Ù…Ø²Ø§Ù…Ù†Ø©. ØªØ­Ù‚Ù‚ Ù…Ù† Ø§Ù„Ø³Ø¬Ù„Ø§Øª.", "danger")

    return redirect(url_for("masterdata.org_node_types_list"))

# ======================
# Dynamic Org Structure (OrgNodeType / OrgNode)
# ======================

@masterdata_bp.route("/org-node-types")
@login_required
@perm_required("MASTERDATA_READ")
def org_node_types_list():
    q = (request.args.get("q") or "").strip()
    query = OrgNodeType.query
    if q:
        like = f"%{q}%"
        query = query.filter(or_(
            OrgNodeType.code.ilike(like),
            OrgNodeType.name_ar.ilike(like),
            OrgNodeType.name_en.ilike(like),
        ))
    query = query.order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc())
    items = query.all()
    return render_template("admin/masterdata/org_node_type_list.html", items=items, q=q, legacy_locked=_legacy_org_locked(), last_sync=_get_setting("ORG_NODE_LAST_SYNC", None))


@masterdata_bp.route("/org-node-types/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def org_node_types_new():
    types = OrgNodeType.query.order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc()).all()
    if request.method == "POST":
        code = _clean(request.form.get("code")).upper().replace(" ", "_")
        name_ar = _clean(request.form.get("name_ar"))
        if not code or not name_ar:
            flash("Ø§Ù„ÙƒÙˆØ¯ ÙˆØ§Ø³Ù… Ø§Ù„Ù†ÙˆØ¹ (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨Ø§Ù†.", "danger")
            return redirect(request.url)

        if OrgNodeType.query.filter_by(code=code).first():
            flash("Ø§Ù„ÙƒÙˆØ¯ Ù…ÙˆØ¬ÙˆØ¯ Ù…Ø³Ø¨Ù‚Ø§Ù‹.", "danger")
            return redirect(request.url)

        t = OrgNodeType(
            code=code,
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")) or None,
            sort_order=int(request.form.get("sort_order") or 0),
            allow_in_approvals=(request.form.get("allow_in_approvals") == "1"),
            show_in_chart=(request.form.get("show_in_chart") == "1"),
            show_in_routes=(request.form.get("show_in_routes") == "1"),
            is_active=(request.form.get("is_active") == "1"),
            created_at=datetime.utcnow(),
        )

        parent_ids = request.form.getlist("allowed_parent_type_ids")
        t.set_allowed_parent_type_ids(parent_ids)

        db.session.add(t)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ù†ÙˆØ¹ Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ©.", "success")
        return redirect(url_for("masterdata.org_node_types_list"))

    return render_template("admin/masterdata/org_node_type_form.html", t=None, types=types)


@masterdata_bp.route("/org-node-types/<int:type_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def org_node_types_edit(type_id: int):
    t = OrgNodeType.query.get_or_404(type_id)
    types = OrgNodeType.query.order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc()).all()

    if request.method == "POST":
        code = _clean(request.form.get("code")).upper().replace(" ", "_")
        name_ar = _clean(request.form.get("name_ar"))
        if not code or not name_ar:
            flash("Ø§Ù„ÙƒÙˆØ¯ ÙˆØ§Ø³Ù… Ø§Ù„Ù†ÙˆØ¹ (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨Ø§Ù†.", "danger")
            return redirect(request.url)

        other = OrgNodeType.query.filter(OrgNodeType.code == code, OrgNodeType.id != t.id).first()
        if other:
            flash("Ø§Ù„ÙƒÙˆØ¯ Ù…Ø³ØªØ®Ø¯Ù… Ù„Ù†ÙˆØ¹ Ø¢Ø®Ø±.", "danger")
            return redirect(request.url)

        t.code = code
        t.name_ar = name_ar
        t.name_en = _clean(request.form.get("name_en")) or None
        t.sort_order = int(request.form.get("sort_order") or 0)

        t.allow_in_approvals = (request.form.get("allow_in_approvals") == "1")
        t.show_in_chart = (request.form.get("show_in_chart") == "1")
        t.show_in_routes = (request.form.get("show_in_routes") == "1")
        t.is_active = (request.form.get("is_active") == "1")

        parent_ids = request.form.getlist("allowed_parent_type_ids")
        t.set_allowed_parent_type_ids(parent_ids)

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.org_node_types_list"))

    return render_template("admin/masterdata/org_node_type_form.html", t=t, types=types)


@masterdata_bp.route("/org-node-types/<int:type_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def org_node_types_delete(type_id: int):
    t = OrgNodeType.query.get_or_404(type_id)
    used = OrgNode.query.filter_by(type_id=t.id).first() is not None
    if used:
        t.is_active = False
        db.session.commit()
        flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ù†ÙˆØ¹ Ù„ÙˆØ¬ÙˆØ¯ Ø¹Ù†Ø§ØµØ± Ù…Ø±ØªØ¨Ø·Ø© Ø¨Ù‡. ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
    else:
        db.session.delete(t)
        db.session.commit()
        flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ù†ÙˆØ¹.", "success")
    return redirect(url_for("masterdata.org_node_types_list"))


# -------- Org Nodes --------

def _build_node_tree_for_picker(selected_type: OrgNodeType | None, exclude_node_id: int | None = None):
    # fetch all active nodes with types
    nodes = (
        OrgNode.query
        .options(selectinload(OrgNode.type))
        .filter(OrgNode.is_active == True)
        .order_by(OrgNode.parent_id.asc().nullslast(), OrgNode.type_id.asc(), OrgNode.sort_order.asc(), OrgNode.name_ar.asc())
        .all()
    )

    # build children map
    children_map: dict[int | None, list[OrgNode]] = {}
    by_id: dict[int, OrgNode] = {}
    for n in nodes:
        by_id[n.id] = n
        children_map.setdefault(n.parent_id, []).append(n)

    # compute disabled ids (exclude node + descendants)
    disabled_ids: set[int] = set()
    if exclude_node_id and exclude_node_id in by_id:
        stack = [exclude_node_id]
        while stack:
            cur = stack.pop()
            if cur in disabled_ids:
                continue
            disabled_ids.add(cur)
            for ch in children_map.get(cur, []):
                stack.append(ch.id)

    allowed_type_ids = set()
    root_allowed = True
    if selected_type is not None:
        allowed = selected_type.allowed_parent_type_ids()
        allowed_type_ids = set(int(x) for x in allowed if int(x) > 0)
        root_allowed = (len(allowed_type_ids) == 0)

    def to_dict(node: OrgNode) -> dict:
        t = node.type
        label = f"{(t.name_ar if t else '')} â€” {node.name_ar}".strip(' â€”')
        elig = (node.type_id in allowed_type_ids) if allowed_type_ids else False
        return {
            "id": node.id,
            "label": label,
            "name_ar": node.name_ar,
            "code": node.code,
            "type_name": (t.name_ar if t else ''),
            "type_code": (t.code if t else ''),
            "eligible": bool(elig) and (node.id not in disabled_ids),
            "disabled": (node.id in disabled_ids),
            "children": [to_dict(ch) for ch in children_map.get(node.id, [])],
        }

    roots = [to_dict(n) for n in children_map.get(None, [])]
    return roots, allowed_type_ids, disabled_ids, root_allowed


@masterdata_bp.route("/org-nodes")
@login_required
@perm_required("MASTERDATA_READ")
def org_nodes_list():
    q = (request.args.get("q") or "").strip()
    type_id = (request.args.get("type_id") or "").strip()

    types = OrgNodeType.query.order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc()).all()

    query = OrgNode.query.options(selectinload(OrgNode.type), selectinload(OrgNode.parent))
    if type_id.isdigit():
        query = query.filter(OrgNode.type_id == int(type_id))
    if q:
        like = f"%{q}%"
        query = query.filter(or_(OrgNode.name_ar.ilike(like), OrgNode.name_en.ilike(like), OrgNode.code.ilike(like)))

    query = query.order_by(OrgNode.type_id.asc(), OrgNode.parent_id.asc().nullslast(), OrgNode.sort_order.asc(), OrgNode.name_ar.asc())
    items = query.all()

    return render_template("admin/masterdata/org_node_list.html", items=items, q=q, types=types, type_id=type_id, legacy_locked=_legacy_org_locked(), last_sync=_get_setting("ORG_NODE_LAST_SYNC", None))


@masterdata_bp.route("/org-nodes/new", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_CREATE")
def org_nodes_new():
    types = OrgNodeType.query.filter_by(is_active=True).order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc()).all()

    if request.method == "POST":
        type_id = (request.form.get("type_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))
        if not type_id.isdigit() or not name_ar:
            flash("Ù†ÙˆØ¹ Ø§Ù„Ø¹Ù†ØµØ± ÙˆØ§Ø³Ù… (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨Ø§Ù†.", "danger")
            return redirect(request.url)

        t = OrgNodeType.query.get(int(type_id))
        if not t:
            flash("Ù†ÙˆØ¹ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯.", "danger")
            return redirect(request.url)

        parent_id_raw = (request.form.get("parent_id") or "").strip()
        parent_id = int(parent_id_raw) if parent_id_raw.isdigit() else None

        allowed = set(t.allowed_parent_type_ids() or [])
        if allowed:
            if not parent_id:
                flash("Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ ÙŠØ­ØªØ§Ø¬ Ø¥Ù„Ù‰ Ø£Ø¨ (Ø­Ø³Ø¨ ØªØ¹Ø±ÙŠÙ Ø§Ù„ØªØ¨Ø¹ÙŠØ©).", "danger")
                return redirect(request.url)
            parent = OrgNode.query.get(parent_id)
            if not parent or parent.type_id not in allowed:
                flash("Ø§Ù„Ø£Ø¨ Ø§Ù„Ù…Ø®ØªØ§Ø± ØºÙŠØ± Ù…Ø³Ù…ÙˆØ­ Ù„Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹.", "danger")
                return redirect(request.url)
        else:
            # root-only
            if parent_id:
                flash("Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ Ù…ÙØ¹Ø±Ù‘Ù ÙƒØ¬Ø°Ø±ÙŠ (Ù„Ø§ ÙŠØ³Ù…Ø­ Ø¨Ø§Ø®ØªÙŠØ§Ø± Ø£Ø¨).", "warning")
                parent_id = None

        n = OrgNode(
            type_id=t.id,
            parent_id=parent_id,
            name_ar=name_ar,
            name_en=_clean(request.form.get("name_en")) or None,
            code=_clean(request.form.get("code")) or None,
            sort_order=int(request.form.get("sort_order") or 0),
            is_active=(request.form.get("is_active") == "1"),
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow(),
        )
        db.session.add(n)
        db.session.commit()
        flash("ØªÙ… Ø¥Ù†Ø´Ø§Ø¡ Ø¹Ù†ØµØ± Ø§Ù„Ù‡ÙŠÙƒÙ„ÙŠØ©.", "success")
        return redirect(url_for("masterdata.org_nodes_list"))

    # default type
    default_type_id = request.args.get('type_id') or ''
    selected_type = OrgNodeType.query.get(int(default_type_id)) if str(default_type_id).isdigit() else None
    tree, allowed_type_ids, disabled_ids, root_allowed = _build_node_tree_for_picker(selected_type, exclude_node_id=None)

    return render_template(
        "admin/masterdata/org_node_form.html",
        n=None,
        types=types,
        selected_type=selected_type,
        tree=tree,
        allowed_type_ids=list(allowed_type_ids),
        root_allowed=root_allowed,
    )


@masterdata_bp.route("/org-nodes/<int:node_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("MASTERDATA_UPDATE")
def org_nodes_edit(node_id: int):
    n = OrgNode.query.get_or_404(node_id)
    types = OrgNodeType.query.filter_by(is_active=True).order_by(OrgNodeType.sort_order.asc(), OrgNodeType.name_ar.asc()).all()

    if request.method == "POST":
        type_id = (request.form.get("type_id") or "").strip()
        name_ar = _clean(request.form.get("name_ar"))
        if not type_id.isdigit() or not name_ar:
            flash("Ù†ÙˆØ¹ Ø§Ù„Ø¹Ù†ØµØ± ÙˆØ§Ø³Ù… (Ø¹Ø±Ø¨ÙŠ) Ù…Ø·Ù„ÙˆØ¨Ø§Ù†.", "danger")
            return redirect(request.url)

        t = OrgNodeType.query.get(int(type_id))
        if not t:
            flash("Ù†ÙˆØ¹ ØºÙŠØ± Ù…ÙˆØ¬ÙˆØ¯.", "danger")
            return redirect(request.url)

        parent_id_raw = (request.form.get("parent_id") or "").strip()
        parent_id = int(parent_id_raw) if parent_id_raw.isdigit() else None

        # prevent self/descendant parenting
        selected_type = t
        tree, allowed_type_ids, disabled_ids, root_allowed = _build_node_tree_for_picker(selected_type, exclude_node_id=n.id)
        if parent_id and parent_id in disabled_ids:
            flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø§Ø®ØªÙŠØ§Ø± Ø§Ù„Ø£Ø¨ Ù…Ù† Ù†ÙØ³ Ø§Ù„Ø¹Ù†ØµØ± Ø£Ùˆ Ø£Ø­Ø¯ Ø§Ù„Ø£Ø¨Ù†Ø§Ø¡.", "danger")
            return redirect(request.url)

        allowed = set(t.allowed_parent_type_ids() or [])
        if allowed:
            if not parent_id:
                flash("Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ ÙŠØ­ØªØ§Ø¬ Ø¥Ù„Ù‰ Ø£Ø¨ (Ø­Ø³Ø¨ ØªØ¹Ø±ÙŠÙ Ø§Ù„ØªØ¨Ø¹ÙŠØ©).", "danger")
                return redirect(request.url)
            parent = OrgNode.query.get(parent_id)
            if not parent or parent.type_id not in allowed:
                flash("Ø§Ù„Ø£Ø¨ Ø§Ù„Ù…Ø®ØªØ§Ø± ØºÙŠØ± Ù…Ø³Ù…ÙˆØ­ Ù„Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹.", "danger")
                return redirect(request.url)
        else:
            if parent_id:
                flash("Ù‡Ø°Ø§ Ø§Ù„Ù†ÙˆØ¹ Ù…ÙØ¹Ø±Ù‘Ù ÙƒØ¬Ø°Ø±ÙŠ (Ù„Ø§ ÙŠØ³Ù…Ø­ Ø¨Ø§Ø®ØªÙŠØ§Ø± Ø£Ø¨).", "warning")
                parent_id = None

        n.type_id = t.id
        n.parent_id = parent_id
        n.name_ar = name_ar
        n.name_en = _clean(request.form.get("name_en")) or None
        n.code = _clean(request.form.get("code")) or None
        n.sort_order = int(request.form.get("sort_order") or 0)
        n.is_active = (request.form.get("is_active") == "1")
        n.updated_at = datetime.utcnow()

        db.session.commit()
        flash("ØªÙ… Ø­ÙØ¸ Ø§Ù„ØªØ¹Ø¯ÙŠÙ„Ø§Øª.", "success")
        return redirect(url_for("masterdata.org_nodes_list"))

    selected_type = OrgNodeType.query.get(n.type_id)
    tree, allowed_type_ids, disabled_ids, root_allowed = _build_node_tree_for_picker(selected_type, exclude_node_id=n.id)

    return render_template(
        "admin/masterdata/org_node_form.html",
        n=n,
        types=types,
        selected_type=selected_type,
        tree=tree,
        allowed_type_ids=list(allowed_type_ids),
        root_allowed=root_allowed,
        disabled_ids=list(disabled_ids),
    )


@masterdata_bp.route("/org-nodes/<int:node_id>/delete", methods=["POST"])
@login_required
@perm_required("MASTERDATA_DELETE")
def org_nodes_delete(node_id: int):
    n = OrgNode.query.get_or_404(node_id)
    has_children = OrgNode.query.filter_by(parent_id=n.id).first() is not None
    if has_children:
        n.is_active = False
        db.session.commit()
        flash("Ù„Ø§ ÙŠÙ…ÙƒÙ† Ø­Ø°Ù Ø§Ù„Ø¹Ù†ØµØ± Ù„ÙˆØ¬ÙˆØ¯ Ø£Ø¨Ù†Ø§Ø¡ ØªØ­ØªÙ‡. ØªÙ… ØªØ¹Ø·ÙŠÙ„Ù‡ Ø¨Ø¯Ù„Ø§Ù‹ Ù…Ù† Ø°Ù„Ùƒ.", "warning")
    else:
        db.session.delete(n)
        db.session.commit()
        flash("ØªÙ… Ø­Ø°Ù Ø§Ù„Ø¹Ù†ØµØ±.", "success")
    return redirect(url_for("masterdata.org_nodes_list"))


@masterdata_bp.route("/org-nodes/parent-tree")
@login_required
@perm_required("MASTERDATA_READ")
def org_nodes_parent_tree():
    type_id = (request.args.get('type_id') or '').strip()
    exclude_id = (request.args.get('exclude_id') or '').strip()
    selected_parent = (request.args.get('selected_parent') or '').strip()

    selected_type = OrgNodeType.query.get(int(type_id)) if type_id.isdigit() else None
    exclude_node_id = int(exclude_id) if exclude_id.isdigit() else None
    selected_parent_id = int(selected_parent) if selected_parent.isdigit() else None

    tree, allowed_type_ids, disabled_ids, root_allowed = _build_node_tree_for_picker(selected_type, exclude_node_id=exclude_node_id)

    return render_template(
        'admin/masterdata/_org_parent_tree.html',
        tree=tree,
        allowed_type_ids=list(allowed_type_ids),
        disabled_ids=list(disabled_ids),
        selected_parent_id=selected_parent_id,
        root_allowed=root_allowed,
        selected_type=selected_type,
    )
