from pathlib import Path
import re

path = Path('/mnt/data/work_units_v2/portal/routes.py')
text = path.read_text(encoding='utf-8')

start_pat = r"def _import_org_structure_excel\(kind: str, file_storage\):"
end_pat = r"\n\n\n@portal_bp\.route\(\"/admin/delegations\""
start_m = re.search(start_pat, text)
end_m = re.search(end_pat, text)
if not start_m or not end_m:
    raise SystemExit('Could not locate _import_org_structure_excel block')

replacement = r'''def _import_org_structure_excel(kind: str, file_storage):
    """Import org structure from an Excel sheet (single sheet).

    Supported kinds:
      - orgs: code, name_ar, name_en, is_active
      - dirs: code, organization_code, name_ar, name_en, is_active
      - units: code, directorate_code, name_ar, name_en, is_active
      - depts: code, directorate_code OR unit_code, name_ar, name_en, is_active
      - secs: code, department_code OR directorate_code OR unit_code, name_ar, name_en, is_active
      - teams: code, section_code, name_ar, name_en, is_active

    Upserts by code.
    """
    kind = (kind or "").strip().lower()
    if kind not in ("orgs", "dirs", "units", "depts", "secs", "teams"):
        return False, "حدد نوع البيانات للاستيراد."

    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_storage, data_only=True)
        ws = wb.active

        # headers
        headers = []
        for cell in ws[1]:
            headers.append((str(cell.value).strip().lower() if cell.value is not None else ""))
        hidx = {h: i for i, h in enumerate(headers)}

        def cell(row, name):
            i = hidx.get(name)
            if i is None:
                return None
            return row[i].value

        def to_code(v):
            if v is None:
                return None
            s = str(v).strip()
            if s == "":
                return None
            # normalize numeric-like values (e.g. 1.0)
            try:
                if isinstance(v, float) and v.is_integer():
                    return str(int(v))
                if s.endswith('.0') and s.replace('.', '', 1).isdigit():
                    return str(int(float(s)))
            except Exception:
                pass
            return s

        def to_bool(v):
            if v is None:
                return True
            s = str(v).strip().lower()
            if s in ("0", "false", "no", "n", "غير", "لا"):
                return False
            return True

        inserted = 0
        updated = 0
        skipped = 0
        errors = []

        for row in ws.iter_rows(min_row=2):
            code = to_code(cell(row, "code"))
            if not code:
                skipped += 1
                continue

            name_ar = (cell(row, "name_ar") or "")
            name_en = (cell(row, "name_en") or "")
            is_active = to_bool(cell(row, "is_active"))

            # parent codes (optional by kind)
            org_code = to_code(cell(row, "organization_code"))
            dir_code = to_code(cell(row, "directorate_code"))
            unit_code = to_code(cell(row, "unit_code"))
            dept_code = to_code(cell(row, "department_code"))
            sec_code = to_code(cell(row, "section_code"))

            try:
                if kind == "orgs":
                    obj = Organization.query.filter_by(code=code).first() or Organization(code=code)
                    is_new = obj.id is None
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

                elif kind == "dirs":
                    org = Organization.query.filter_by(code=org_code).first() if org_code else None
                    if not org:
                        skipped += 1
                        errors.append(f"directorate code={code}: organization_code not found")
                        continue
                    obj = Directorate.query.filter_by(code=code).first() or Directorate(code=code)
                    is_new = obj.id is None
                    obj.organization_id = org.id
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

                elif kind == "units":
                    di = Directorate.query.filter_by(code=dir_code).first() if dir_code else None
                    if not di:
                        skipped += 1
                        errors.append(f"unit code={code}: directorate_code not found")
                        continue
                    obj = Unit.query.filter_by(code=code).first() or Unit(code=code)
                    is_new = obj.id is None
                    obj.directorate_id = di.id
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

                elif kind == "depts":
                    # parent: directorate OR unit
                    parent_dir = Directorate.query.filter_by(code=dir_code).first() if dir_code else None
                    parent_unit = Unit.query.filter_by(code=unit_code).first() if unit_code else None
                    if bool(parent_dir) == bool(parent_unit):
                        skipped += 1
                        errors.append(f"department code={code}: provide directorate_code OR unit_code")
                        continue
                    obj = Department.query.filter_by(code=code).first() or Department(code=code)
                    is_new = obj.id is None
                    if parent_unit:
                        obj.unit_id = parent_unit.id
                        obj.directorate_id = None
                    else:
                        obj.directorate_id = parent_dir.id
                        obj.unit_id = None
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

                elif kind == "secs":
                    parent_dept = Department.query.filter_by(code=dept_code).first() if dept_code else None
                    parent_dir = Directorate.query.filter_by(code=dir_code).first() if dir_code else None
                    parent_unit = Unit.query.filter_by(code=unit_code).first() if unit_code else None
                    # exactly one
                    parents = [p for p in (parent_dept, parent_dir, parent_unit) if p]
                    if len(parents) != 1:
                        skipped += 1
                        errors.append(f"section code={code}: provide exactly one of department_code/directorate_code/unit_code")
                        continue
                    obj = Section.query.filter_by(code=code).first() or Section(code=code)
                    is_new = obj.id is None
                    if parent_dept:
                        obj.department_id = parent_dept.id
                        obj.directorate_id = None
                        obj.unit_id = None
                    elif parent_unit:
                        obj.unit_id = parent_unit.id
                        obj.department_id = None
                        obj.directorate_id = None
                    else:
                        obj.directorate_id = parent_dir.id
                        obj.department_id = None
                        obj.unit_id = None
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

                elif kind == "teams":
                    sec = Section.query.filter_by(code=sec_code).first() if sec_code else None
                    if not sec:
                        skipped += 1
                        errors.append(f"team code={code}: section_code not found")
                        continue
                    obj = Team.query.filter_by(code=code).first() or Team(code=code)
                    is_new = obj.id is None
                    obj.section_id = sec.id
                    obj.name_ar = str(name_ar).strip() or None
                    obj.name_en = str(name_en).strip() or None
                    obj.is_active = bool(is_active)
                    db.session.add(obj)
                    inserted += int(is_new)
                    updated += int(not is_new)

            except Exception as e:
                skipped += 1
                errors.append(f"row code={code}: {e}")

        db.session.commit()
        msg = f"تم الاستيراد: {inserted} إضافة، {updated} تحديث، {skipped} تجاوز." 
        if errors:
            msg += " (ملاحظات: " + "; ".join(errors[:6]) + (")" if len(errors) > 6 else ")")
        return True, msg
    except Exception:
        db.session.rollback()
        return False, "تعذر قراءة ملف Excel. تأكد أنه .xlsx وبالأعمدة المطلوبة." 
'''

text_new = text[:start_m.start()] + replacement + text[end_m.start():]
path.write_text(text_new, encoding='utf-8')
print('portal _import_org_structure_excel updated')
