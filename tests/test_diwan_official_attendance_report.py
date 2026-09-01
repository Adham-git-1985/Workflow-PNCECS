import unittest
from datetime import date, datetime
from pathlib import Path

from flask import Flask

from extensions import db
from models import AttendanceDailySummary, SystemSetting, User, WorkAssignment, WorkPolicy, WorkSchedule
from portal.routes import _diwan_official_month_rows, _export_diwan_official_attendance_xlsx


PROJECT_ROOT = Path(__file__).resolve().parents[1]


class DiwanOfficialAttendanceReportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.routes = (PROJECT_ROOT / 'portal' / 'routes.py').read_text(encoding='utf-8')
        cls.template = (
            PROJECT_ROOT / 'templates' / 'portal' / 'hr' / 'reports_diwan.html'
        ).read_text(encoding='utf-8')

    def test_official_sheet_has_daily_symbols_and_symbol_key(self):
        self.assertIn('def _diwan_official_month_rows', self.routes)
        self.assertIn('def _export_diwan_official_attendance_xlsx', self.routes)
        self.assertIn("sheet.title = 'كشف الدوام'", self.routes)
        self.assertIn("key_sheet = workbook.create_sheet('الرموز')", self.routes)
        self.assertIn("download_name=f'official_diwan_attendance_{year}_{month:02d}.xlsx'", self.routes)
        for symbol in ('L', 'W', 'M', 'S', 'F', 'V', '+', 'X', 'E', '*'):
            with self.subTest(symbol=symbol):
                self.assertIn(f"('{symbol}',", self.routes)
        self.assertIn("cell['symbol'] = '*'", self.routes)
        self.assertIn("HYBRID_WEEKLY_QUOTA", self.routes)

    def test_departures_are_preserved_in_the_notes_column(self):
        self.assertIn("private_minutes", self.routes)
        self.assertIn("official_minutes", self.routes)
        self.assertIn("'شخصية {private_minutes} د'", self.routes)
        self.assertIn("'رسمية {official_minutes} د'", self.routes)
        self.assertIn("export='official_xlsx'", self.template)
        self.assertIn("scope_start = start - timedelta(days=start.weekday())", self.routes)


class DiwanEmergencyDutySymbolTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = Flask(__name__)
        cls.app.config.update(
            TESTING=True,
            SECRET_KEY='diwan-emergency-duty-test',
            SQLALCHEMY_DATABASE_URI='sqlite:///:memory:',
            SQLALCHEMY_TRACK_MODIFICATIONS=False,
        )
        db.init_app(cls.app)
        cls.context = cls.app.app_context()
        cls.context.push()

    @classmethod
    def tearDownClass(cls):
        db.session.remove()
        db.drop_all()
        cls.context.pop()

    def setUp(self):
        db.session.remove()
        db.drop_all()
        db.create_all()

    def test_flexible_two_day_hybrid_policy_marks_remaining_workdays_as_emergency_duty(self):
        user = User(
            email='hybrid-attendance@example.test',
            name='Hybrid Attendance',
            password_hash='not-used-in-test',
            role='EMPLOYEE',
        )
        schedule = WorkSchedule(name='Regular', kind='FIXED')
        policy = WorkPolicy(
            name='Outside Ramallah',
            days_policy='HYBRID_WEEKLY_QUOTA',
            hybrid_office_days=2,
            hybrid_remote_days=3,
            hybrid_selection_mode='FLEXIBLE',
            location_policy='HYBRID',
            is_active=True,
        )
        db.session.add_all((user, schedule, policy, SystemSetting(key='HR_WEEKLY_HOLIDAYS_MASK', value='48')))
        db.session.flush()
        db.session.add(WorkAssignment(
            name='Outside Ramallah employee',
            schedule_id=schedule.id,
            policy_id=policy.id,
            target_type='USER',
            target_user_id=user.id,
            is_active=True,
        ))
        summaries = [
            AttendanceDailySummary(
                user_id=user.id,
                day='2026-03-02',
                first_in=datetime(2026, 3, 2, 8, 0),
                last_out=datetime(2026, 3, 2, 15, 0),
                status='OK',
            ),
            AttendanceDailySummary(
                user_id=user.id,
                day='2026-03-03',
                first_in=datetime(2026, 3, 3, 8, 20),
                last_out=datetime(2026, 3, 3, 15, 0),
                late_minutes=20,
                status='OK',
            ),
        ]
        db.session.add_all(summaries)
        db.session.commit()

        rows = _diwan_official_month_rows(
            [user.id],
            datetime(2026, 3, 2).date(),
            datetime(2026, 3, 8).date(),
            summaries,
            {},
        )

        self.assertEqual(rows[0]['symbols'], ['+', 'X', '*', '*', 'F', 'F', '*'])

    def test_flexible_three_day_hybrid_policy_uses_its_own_quota(self):
        user = User(
            email='inside-hybrid-attendance@example.test',
            name='Inside Hybrid Attendance',
            password_hash='not-used-in-test',
            role='EMPLOYEE',
        )
        schedule = WorkSchedule(name='Regular', kind='FIXED')
        policy = WorkPolicy(
            name='Inside Ramallah',
            days_policy='HYBRID_WEEKLY_QUOTA',
            hybrid_office_days=3,
            hybrid_remote_days=2,
            hybrid_selection_mode='FLEXIBLE',
            location_policy='HYBRID',
            is_active=True,
        )
        db.session.add_all((user, schedule, policy, SystemSetting(key='HR_WEEKLY_HOLIDAYS_MASK', value='48')))
        db.session.flush()
        db.session.add(WorkAssignment(
            name='Inside Ramallah employee',
            schedule_id=schedule.id,
            policy_id=policy.id,
            target_type='USER',
            target_user_id=user.id,
            is_active=True,
        ))
        summaries = [
            AttendanceDailySummary(
                user_id=user.id,
                day=day,
                first_in=datetime(2026, 3, index, 8, 0),
                last_out=datetime(2026, 3, index, 15, 0),
                late_minutes=late_minutes,
                status='OK',
            )
            for index, day, late_minutes in (
                (2, '2026-03-02', 0),
                (3, '2026-03-03', 0),
                (4, '2026-03-04', 15),
            )
        ]
        db.session.add_all(summaries)
        db.session.commit()

        rows = _diwan_official_month_rows(
            [user.id],
            datetime(2026, 3, 2).date(),
            datetime(2026, 3, 8).date(),
            summaries,
            {},
        )

        self.assertEqual(rows[0]['symbols'], ['+', '+', 'X', '*', 'F', 'F', '*'])

    def test_emergency_duty_symbol_is_written_to_the_exported_workbook(self):
        from openpyxl import load_workbook

        workbook = load_workbook(_export_diwan_official_attendance_xlsx(
            date(2026, 3, 1),
            date(2026, 3, 1),
            [{
                'month': 3,
                'year': 2026,
                'ministry_code': '01',
                'employee_no': '100',
                'employee_name': 'Emergency Duty Employee',
                'symbols': ['*'],
                'notes': '',
            }],
        ))

        sheet = workbook['كشف الدوام']
        self.assertEqual(sheet.cell(row=3, column=6).value, '*')
        self.assertEqual(sheet.cell(row=3, column=6).fill.fgColor.rgb, '00D9EAD3')


if __name__ == '__main__':
    unittest.main()
