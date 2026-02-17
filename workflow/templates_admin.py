# workflow/templates_admin.py

from flask import render_template, request, redirect, url_for, flash, send_file
from flask_login import login_required, current_user
from sqlalchemy import func
from io import BytesIO
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


from . import workflow_bp
from extensions import db
from utils.perms import perm_required
from utils.org_dynamic import build_org_node_picker_tree
from models import (
    WorkflowTemplate,
    WorkflowTemplateStep,
    WorkflowTemplateParallelAssignee,
    User,
    Department,
    Directorate,
    Unit,
    Section,
    Division,
    Role,
    Committee,
    CommitteeAssignee,

    OrgNode,
    OrgNodeType,
)


def _to_int(val, default=None):
    try:
        if val is None or str(val).strip() == "":
            return default
        return int(val)
    except Exception:
        return default


def _normalize_kind(kind: str) -> str:
    kind = (kind or "").strip().upper()
    if kind in ("USER", "ROLE", "DEPARTMENT", "DIRECTORATE", "UNIT", "SECTION", "DIVISION", "ORG_NODE", "COMMITTEE"):
        return kind
    return ""


def _normalize_step_mode(mode: str) -> str:
    mode = (mode or "").strip().upper()
    if mode in ("SEQUENTIAL", "PARALLEL_SYNC"):
        return mode
    return "SEQUENTIAL"




def _normalize_committee_delivery_mode(mode: str) -> str:
    """Normalize committee delivery mode.

    Canonical values stored in DB/UI:
      - Committee_ALL
      - Committee_CHAIR
      - Committee_SECRETARY

    Backward-compatible aliases accepted:
      - COMMITTEE_ALL / COMMITTEE_CHAIR / COMMITTEE_SECRETARY
    """
    mode_raw = (mode or '').strip()
    if not mode_raw:
        return 'Committee_ALL'

    up = mode_raw.upper()
    if up in ('COMMITTEE_ALL', 'COMMITTEE_CHAIR', 'COMMITTEE_SECRETARY'):
        return 'Committee_' + up.split('_', 1)[1]

    # Accept already canonical values
    if mode_raw in ('Committee_ALL', 'Committee_CHAIR', 'Committee_SECRETARY'):
        return mode_raw

    return 'Committee_ALL'

def _get_role_choices():
    """Return role list for UI (active roles table + any roles already in DB)."""
    seen = set()
    out = []

    def add(r):
        if not r:
            return
        r = str(r).strip()
        if not r:
            return
        key = r.lower()
        if key in seen:
            return
        seen.add(key)
        out.append(r)

    # 1) Preferred source: active roles from master data
    try:
        for r in Role.query.filter_by(is_active=True).order_by(Role.code.asc()).all():
            add(r.code)
    except Exception:
        pass

    # 2) Fallback: any roles already used by users
    for (r,) in db.session.query(User.role).distinct().all():
        add(r)

    # 3) Ensure core system roles exist
    for r in ["USER", "ADMIN", "SUPER_ADMIN"]:
        add(r)

    return out


def _resequence_steps(template_id: int):
    steps = (
        WorkflowTemplateStep.query.filter_by(template_id=template_id)
        .order_by(WorkflowTemplateStep.step_order.asc())
        .all()
    )
    for i, s in enumerate(steps, start=1):
        s.step_order = i
        # keep denormalized fields for parallel assignees in sync
        try:
            WorkflowTemplateParallelAssignee.query.filter_by(template_step_id=s.id).update({"template_id": s.template_id, "step_order": i})
        except Exception:
            pass
    db.session.commit()

def _xlsx_response(sheet_name: str, headers: list[str], rows: list[list], filename_prefix: str):
    """Create an .xlsx response (download) using openpyxl."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = (sheet_name or "Sheet1")[:31]  # Excel sheet name max = 31

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_alignment

    for row in rows:
        ws.append(row)

    # Simple auto width (capped)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[col_letter]:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len, 60) + 2

    ws.freeze_panes = "A2"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    ts = datetime.now().strftime("%Y-%m-%d_%H-%M")
    filename = f"{filename_prefix}_{ts}.xlsx"

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )



@workflow_bp.route("/templates")
@login_required
@perm_required("WORKFLOW_TEMPLATES_READ")
def templates_list():
    q = (request.args.get("q") or "").strip()

    query = WorkflowTemplate.query
    if q:
        like = f"%{q}%"
        query = query.filter(WorkflowTemplate.name.ilike(like))

    items = query.order_by(WorkflowTemplate.id.desc()).all()
    return render_template("workflow/templates_admin/list.html", items=items, q=q)


@workflow_bp.route("/templates/export-excel")
@login_required
@perm_required("WORKFLOW_TEMPLATES_READ")
def templates_export_excel():
    """Export workflow templates list to Excel. Respects ?q filter."""
    q = (request.args.get("q") or "").strip()

    query = WorkflowTemplate.query
    if q:
        like = f"%{q}%"
        query = query.filter(WorkflowTemplate.name.ilike(like))

    items = query.order_by(WorkflowTemplate.id.desc()).all()

    # created_by email map
    creator_ids = [t.created_by_id for t in items if getattr(t, "created_by_id", None)]
    creators = []
    if creator_ids:
        creators = User.query.filter(User.id.in_(creator_ids)).all()
    creators_map = {u.id: u for u in creators}

    headers = [
        "ID",
        "اسم المسار",
        "نشط",
        "SLA الافتراضي (أيام)",
        "عدد الخطوات",
        "أنشئ بواسطة",
        "تاريخ الإنشاء",
    ]

    rows = []
    for t in items:
        created_at = getattr(t, "created_at", None)
        creator = creators_map.get(getattr(t, "created_by_id", None))
        rows.append([
            getattr(t, "id", ""),
            getattr(t, "name", ""),
            "نعم" if getattr(t, "is_active", False) else "لا",
            getattr(t, "sla_days_default", None) if getattr(t, "sla_days_default", None) is not None else "",
            len(getattr(t, "steps", []) or []),
            getattr(creator, "email", "") if creator else (getattr(t, "created_by_id", "") or ""),
            created_at.strftime("%Y-%m-%d %H:%M") if created_at else "",
        ])

    return _xlsx_response(
        sheet_name="Templates",
        headers=headers,
        rows=rows,
        filename_prefix="workflow_templates",
    )


@workflow_bp.route("/templates/<int:template_id>/steps/export-excel")
@login_required
@perm_required("WORKFLOW_TEMPLATES_READ")
def templates_steps_export_excel(template_id: int):
    """Export steps of a template to Excel.

    ⚠️ Important: this export is designed to be *round‑trippable* with the
    "استيراد خطوات المسار" feature.

    Columns are the same suggested in the UI:
      step_order, mode, approver_kind, target, sla_days, committee_delivery_mode, parallel_assignees

    `target` is exported as a value that the importer can resolve:
      - USER: email
      - ROLE: role code
      - DEPARTMENT/DIRECTORATE/UNIT/SECTION/DIVISION/ORG_NODE/COMMITTEE: code (if exists) else numeric id

    `parallel_assignees` format:
      KIND:VALUE; KIND:VALUE; ...
    """
    t = WorkflowTemplate.query.get_or_404(template_id)

    steps = (
        WorkflowTemplateStep.query.filter_by(template_id=t.id)
        .order_by(WorkflowTemplateStep.step_order.asc())
        .all()
    )

    # Collect referenced ids (including parallel assignees)
    user_ids, dept_ids, dir_ids = set(), set(), set()
    unit_ids, section_ids, division_ids = set(), set(), set()
    org_node_ids, committee_ids = set(), set()

    for s in steps:
        if getattr(s, 'approver_user_id', None):
            user_ids.add(s.approver_user_id)
        if getattr(s, 'approver_department_id', None):
            dept_ids.add(s.approver_department_id)
        if getattr(s, 'approver_directorate_id', None):
            dir_ids.add(s.approver_directorate_id)
        if getattr(s, 'approver_unit_id', None):
            unit_ids.add(s.approver_unit_id)
        if getattr(s, 'approver_section_id', None):
            section_ids.add(s.approver_section_id)
        if getattr(s, 'approver_division_id', None):
            division_ids.add(s.approver_division_id)
        if getattr(s, 'approver_org_node_id', None):
            org_node_ids.add(s.approver_org_node_id)
        if getattr(s, 'approver_committee_id', None):
            committee_ids.add(s.approver_committee_id)

        for a in (getattr(s, 'parallel_assignees', None) or []):
            if getattr(a, 'approver_user_id', None):
                user_ids.add(a.approver_user_id)
            if getattr(a, 'approver_department_id', None):
                dept_ids.add(a.approver_department_id)
            if getattr(a, 'approver_directorate_id', None):
                dir_ids.add(a.approver_directorate_id)
            if getattr(a, 'approver_unit_id', None):
                unit_ids.add(a.approver_unit_id)
            if getattr(a, 'approver_section_id', None):
                section_ids.add(a.approver_section_id)
            if getattr(a, 'approver_division_id', None):
                division_ids.add(a.approver_division_id)
            if getattr(a, 'approver_org_node_id', None):
                org_node_ids.add(a.approver_org_node_id)
            if getattr(a, 'approver_committee_id', None):
                committee_ids.add(a.approver_committee_id)

    users_map = {u.id: u for u in (User.query.filter(User.id.in_(user_ids)).all() if user_ids else [])}
    depts_map = {d.id: d for d in (Department.query.filter(Department.id.in_(dept_ids)).all() if dept_ids else [])}
    dirs_map  = {d.id: d for d in (Directorate.query.filter(Directorate.id.in_(dir_ids)).all() if dir_ids else [])}
    units_map = {u.id: u for u in (Unit.query.filter(Unit.id.in_(unit_ids)).all() if unit_ids else [])}
    sections_map = {s.id: s for s in (Section.query.filter(Section.id.in_(section_ids)).all() if section_ids else [])}
    divisions_map = {d.id: d for d in (Division.query.filter(Division.id.in_(division_ids)).all() if division_ids else [])}
    org_nodes_map = {n.id: n for n in (OrgNode.query.filter(OrgNode.id.in_(org_node_ids)).all() if org_node_ids else [])}
    committees_map = {c.id: c for c in (Committee.query.filter(Committee.id.in_(committee_ids)).all() if committee_ids else [])}

    def _code_or_id(obj):
        if not obj:
            return ''
        code = getattr(obj, 'code', None)
        if code and str(code).strip():
            return str(code).strip()
        oid = getattr(obj, 'id', None)
        return str(oid) if oid is not None else ''

    def _target_for(kind: str, step_or_assignee):
        k = (kind or '').upper().strip()
        if k == 'USER':
            u = users_map.get(getattr(step_or_assignee, 'approver_user_id', None))
            return (getattr(u, 'email', None) or '').strip() if u else (str(getattr(step_or_assignee, 'approver_user_id', '') or '').strip())
        if k == 'ROLE':
            return (getattr(step_or_assignee, 'approver_role', None) or '').strip()
        if k == 'DEPARTMENT':
            return _code_or_id(depts_map.get(getattr(step_or_assignee, 'approver_department_id', None))) or str(getattr(step_or_assignee, 'approver_department_id', '') or '').strip()
        if k == 'DIRECTORATE':
            return _code_or_id(dirs_map.get(getattr(step_or_assignee, 'approver_directorate_id', None))) or str(getattr(step_or_assignee, 'approver_directorate_id', '') or '').strip()
        if k == 'UNIT':
            return _code_or_id(units_map.get(getattr(step_or_assignee, 'approver_unit_id', None))) or str(getattr(step_or_assignee, 'approver_unit_id', '') or '').strip()
        if k == 'SECTION':
            return _code_or_id(sections_map.get(getattr(step_or_assignee, 'approver_section_id', None))) or str(getattr(step_or_assignee, 'approver_section_id', '') or '').strip()
        if k == 'DIVISION':
            return _code_or_id(divisions_map.get(getattr(step_or_assignee, 'approver_division_id', None))) or str(getattr(step_or_assignee, 'approver_division_id', '') or '').strip()
        if k == 'ORG_NODE':
            return _code_or_id(org_nodes_map.get(getattr(step_or_assignee, 'approver_org_node_id', None))) or str(getattr(step_or_assignee, 'approver_org_node_id', '') or '').strip()
        if k == 'COMMITTEE':
            return _code_or_id(committees_map.get(getattr(step_or_assignee, 'approver_committee_id', None))) or str(getattr(step_or_assignee, 'approver_committee_id', '') or '').strip()
        # fallback
        return (getattr(step_or_assignee, 'approver_role', None) or '').strip() or ''

    headers = [
        'template_id',
        'template_name',
        'step_order',
        'mode',
        'approver_kind',
        'target',
        'sla_days',
        'committee_delivery_mode',
        'parallel_assignees',
    ]

    rows = []
    for s in steps:
        kind = getattr(s, 'approver_kind', '')
        target = _target_for(kind, s)

        cmode = getattr(s, 'committee_delivery_mode', None) or ''
        if (kind or '').upper().strip() != 'COMMITTEE':
            cmode = ''

        extras = []
        for a in (getattr(s, 'parallel_assignees', None) or []):
            ak = (getattr(a, 'approver_kind', '') or '').upper().strip() or 'USER'
            av = _target_for(ak, a)
            if not av:
                continue
            extras.append(f"{ak}:{av}")
        extras_str = '; '.join(extras)

        rows.append([
            getattr(t, 'id', ''),
            getattr(t, 'name', ''),
            getattr(s, 'step_order', ''),
            (getattr(s, 'mode', None) or 'SEQUENTIAL'),
            (kind or ''),
            target,
            getattr(s, 'sla_days', None) if getattr(s, 'sla_days', None) is not None else '',
            cmode,
            extras_str,
        ])

    return _xlsx_response(
        sheet_name=f"Steps_{t.id}",
        headers=headers,
        rows=rows,
        filename_prefix=f"template_{t.id}_steps",
    )



@workflow_bp.route("/templates/import-excel", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_import_excel():
    """Import Workflow Templates from Excel.

    Columns (recommended):
      - name (required)
      - is_active (1/0)
      - sla_days_default (int)

    Modes:
      - safe: upsert by name
      - replace: tries to clear templates + steps then re-insert (falls back to soft replace if FK blocks)
    """
    from utils import importer as xl

    mode = (request.form.get("mode") or "safe").strip().lower()
    if mode not in ("replace", "safe"):
        mode = "safe"

    f = request.files.get("file")
    if not f or not getattr(f, "filename", ""):
        flash("اختر ملف Excel (.xlsx)", "danger")
        return redirect(url_for("workflow.templates_list"))

    try:
        _sheet, rows, headers = xl.read_excel_rows(f)
    except Exception as e:
        flash(f"خطأ في قراءة ملف Excel: {e}", "danger")
        return redirect(url_for("workflow.templates_list"))

    created = updated = skipped = 0

    def do_upsert():
        nonlocal created, updated, skipped
        for r in rows:
            name = xl.to_str(xl.pick(r, "name", "template_name", "اسم المسار", "المسار"))
            if not name:
                skipped += 1
                continue
            name = name.strip()
            is_active = xl.to_bool(xl.pick(r, "is_active", "active", "نشط"), default=True)
            sla = xl.to_int(xl.pick(r, "sla_days_default", "sla_days", "sla", "sla الافتراضي"), default=None)

            obj = WorkflowTemplate.query.filter(func.lower(WorkflowTemplate.name) == name.lower()).first()
            if obj:
                obj.name = name
                obj.is_active = bool(is_active) if is_active is not None else obj.is_active
                obj.sla_days_default = sla
                updated += 1
            else:
                obj = WorkflowTemplate(name=name, is_active=bool(is_active) if is_active is not None else True, sla_days_default=sla)
                db.session.add(obj)
                created += 1

        return created, updated

    if mode == "replace":
        try:
            # Clear in dependency order
            WorkflowTemplateParallelAssignee.query.delete(synchronize_session=False)
            WorkflowTemplateStep.query.delete(synchronize_session=False)
            WorkflowTemplate.query.delete(synchronize_session=False)
            db.session.flush()

            do_upsert()
            db.session.commit()
            flash(f"✅ تم الاستيراد (Replace). جديد: {created} / تحديث: {updated} / تخطي: {skipped}", "success")
            return redirect(url_for("workflow.templates_list"))
        except Exception:
            db.session.rollback()
            # Soft fallback: disable all templates then upsert
            try:
                for t in WorkflowTemplate.query.all():
                    t.is_active = False
                db.session.flush()
                do_upsert()
                db.session.commit()
                flash(f"✅ تم الاستيراد (Soft Replace). جديد: {created} / تحديث: {updated} / تخطي: {skipped}", "success")
                return redirect(url_for("workflow.templates_list"))
            except Exception as e:
                db.session.rollback()
                flash(f"❌ فشل الاستيراد: {e}", "danger")
                return redirect(url_for("workflow.templates_list"))

    # Safe mode
    try:
        do_upsert()
        db.session.commit()
        flash(f"✅ تم الاستيراد (Safe Replace). جديد: {created} / تحديث: {updated} / تخطي: {skipped}", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ فشل الاستيراد: {e}", "danger")

    return redirect(url_for("workflow.templates_list"))
@workflow_bp.route("/templates/new", methods=["GET", "POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_CREATE")
def templates_new():
    users = User.query.order_by(User.email.asc()).all()
    departments = (
        Department.query.filter_by(is_active=True).order_by(Department.name_ar.asc()).all()
    )
    directorates = (
        Directorate.query.filter_by(is_active=True)
        .order_by(Directorate.name_ar.asc())
        .all()
    )

    units = (
        Unit.query.filter_by(is_active=True)
        .order_by(Unit.name_ar.asc())
        .all()
    )
    sections = (
        Section.query.filter_by(is_active=True)
        .order_by(Section.name_ar.asc())
        .all()
    )
    divisions = (
        Division.query.filter_by(is_active=True)
        .order_by(Division.name_ar.asc())
        .all()
    )
    role_choices = _get_role_choices()
    committees = Committee.query.filter_by(is_active=True).order_by(Committee.name_ar.asc()).all()

    if request.method == "POST":
        name = (request.form.get("name") or "").strip()
        sla_days_default = _to_int(request.form.get("sla_days_default"), default=None)

        if not name:
            flash("اسم المسار مطلوب.", "danger")
            return redirect(request.url)

        t = WorkflowTemplate(
            name=name,
            is_active=True,  # new.html عادة بدون checkbox
            created_by_id=current_user.id,
            sla_days_default=sla_days_default,
        )
        db.session.add(t)
        db.session.flush()  # نحتاج t.id قبل إضافة الخطوات

        # Arrays from new.html
        kinds = request.form.getlist("approver_kind")
        modes = request.form.getlist("step_mode")
        user_ids = request.form.getlist("approver_user_id")
        roles = request.form.getlist("approver_role")
        dept_ids = request.form.getlist("approver_department_id")
        dir_ids = request.form.getlist("approver_directorate_id")
        committee_ids = request.form.getlist("approver_committee_id")
        committee_modes = request.form.getlist("committee_delivery_mode")
        sla_list = request.form.getlist("step_sla_days")
        unit_ids = request.form.getlist("approver_unit_id")
        section_ids = request.form.getlist("approver_section_id")
        division_ids = request.form.getlist("approver_division_id")

        max_len = max(
            len(kinds),
            len(modes),
            len(user_ids),
            len(roles),
            len(dept_ids),
            len(dir_ids),
            len(unit_ids),
            len(section_ids),
            len(division_ids),
            len(committee_ids),
            len(committee_modes),
            len(sla_list),
            0,
        )

        step_order = 1

        for i in range(max_len):
            kind = _normalize_kind(kinds[i] if i < len(kinds) else "")
            mode = _normalize_step_mode(modes[i] if i < len(modes) else "SEQUENTIAL")
            user_id = _to_int(user_ids[i] if i < len(user_ids) else "", default=None)
            role = (roles[i] if i < len(roles) else "").strip()
            dept_id = _to_int(dept_ids[i] if i < len(dept_ids) else "", default=None)
            dir_id = _to_int(dir_ids[i] if i < len(dir_ids) else "", default=None)
            committee_id = _to_int(committee_ids[i] if i < len(committee_ids) else "", default=None)
            committee_delivery_mode = _normalize_committee_delivery_mode(committee_modes[i] if i < len(committee_modes) else 'Committee_ALL')
            sla_days = _to_int(sla_list[i] if i < len(sla_list) else "", default=None)
            unit_id = _to_int(unit_ids[i] if i < len(unit_ids) else "", default=None)
            section_id = _to_int(section_ids[i] if i < len(section_ids) else "", default=None)
            division_id = _to_int(division_ids[i] if i < len(division_ids) else "", default=None)

            # Skip completely empty row
            if (
                not kind
                and not user_id
                and not role
                and not dept_id
                and not dir_id
                and sla_days is None
            ):
                continue

            if not kind:
                db.session.rollback()
                flash("يرجى اختيار نوع المعتمد لكل خطوة.", "danger")
                return redirect(request.url)

            if kind == "USER" and not user_id:
                db.session.rollback()
                flash("يرجى اختيار مستخدم لخطوة نوع USER.", "danger")
                return redirect(request.url)

            if kind == "ROLE" and not role:
                db.session.rollback()
                flash("يرجى اختيار ROLE لخطوة نوع ROLE.", "danger")
                return redirect(request.url)

            if kind == "DEPARTMENT" and not dept_id:
                db.session.rollback()
                flash("يرجى اختيار دائرة/قسم للخطوة.", "danger")
                return redirect(request.url)

            if kind == "COMMITTEE" and not committee_id:
                db.session.rollback()
                flash("يرجى اختيار لجنة للخطوة.", "danger")
                return redirect(request.url)

            if kind == "UNIT" and not unit_id:
                db.session.rollback()
                flash("يرجى اختيار وحدة للخطوة.", "danger")
                return redirect(request.url)

            if kind == "SECTION" and not section_id:
                db.session.rollback()
                flash("يرجى اختيار قسم للخطوة.", "danger")
                return redirect(request.url)

            if kind == "DIVISION" and not division_id:
                db.session.rollback()
                flash("يرجى اختيار شعبة للخطوة.", "danger")
                return redirect(request.url)

            # Optional validation: CHAIR/SECRETARY must exist
            if kind == "COMMITTEE" and committee_delivery_mode in ("Committee_CHAIR", "Committee_SECRETARY"):
                needed = 'CHAIR' if committee_delivery_mode == "Committee_CHAIR" else 'SECRETARY'
                exists = CommitteeAssignee.query.filter_by(committee_id=committee_id, is_active=True).filter(CommitteeAssignee.member_role.ilike(needed)).first()
                if not exists:
                    db.session.rollback()
                    flash("اللجنة المختارة لا تحتوي على %s نشط." % ('رئيس' if needed=='CHAIR' else 'مقرر'), "danger")
                    return redirect(request.url)

            if kind == "DIRECTORATE" and not dir_id:
                db.session.rollback()
                flash("يرجى اختيار إدارة للخطوة.", "danger")
                return redirect(request.url)

            step = WorkflowTemplateStep(
                template_id=t.id,
                step_order=step_order,
                approver_kind=kind,
                approver_user_id=user_id if kind == "USER" else None,
                approver_role=role if kind == "ROLE" else None,
                approver_department_id=dept_id if kind == "DEPARTMENT" else None,
                approver_directorate_id=dir_id if kind == "DIRECTORATE" else None,
                approver_committee_id=committee_id if kind == "COMMITTEE" else None,
                committee_delivery_mode=committee_delivery_mode if kind == "COMMITTEE" else None,
                sla_days=sla_days,
                approver_unit_id=unit_id if kind == "UNIT" else None,
                approver_section_id=section_id if kind == "SECTION" else None,
                approver_division_id=division_id if kind == "DIVISION" else None,
                mode=mode,
            )
            db.session.add(step)
            step_order += 1

        if step_order == 1:
            db.session.rollback()
            flash("أضف خطوة واحدة على الأقل.", "danger")
            return redirect(request.url)

        db.session.commit()
        flash("تم إنشاء المسار وخطواته.", "success")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    return render_template(
        "workflow/templates_admin/new.html",
        users=users,
        departments=departments,
        directorates=directorates,
        units=units,
        sections=sections,
        divisions=divisions,
        role_choices=role_choices,
        committees=committees,
    )


@workflow_bp.route("/templates/<int:template_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_edit(template_id):
    t = WorkflowTemplate.query.get_or_404(template_id)

    if request.method == "POST":
        t.name = (request.form.get("name") or "").strip()
        t.is_active = request.form.get("is_active") == "1"
        t.sla_days_default = _to_int(request.form.get("sla_days_default"), default=None)

        if not t.name:
            flash("اسم المسار مطلوب.", "danger")
            return redirect(request.url)

        db.session.commit()
        flash("تم حفظ بيانات المسار.", "success")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    steps = (
        WorkflowTemplateStep.query.filter_by(template_id=t.id)
        .order_by(WorkflowTemplateStep.step_order.asc())
        .all()
    )

    users = User.query.order_by(User.email.asc()).all()
    departments = (
        Department.query.filter_by(is_active=True).order_by(Department.name_ar.asc()).all()
    )
    directorates = (
        Directorate.query.filter_by(is_active=True)
        .order_by(Directorate.name_ar.asc())
        .all()
    )

    # Needed for adding steps with kinds: UNIT / SECTION / DIVISION
    units = Unit.query.filter_by(is_active=True).order_by(Unit.name_ar.asc()).all()
    sections = Section.query.filter_by(is_active=True).order_by(Section.name_ar.asc()).all()
    divisions = Division.query.filter_by(is_active=True).order_by(Division.name_ar.asc()).all()

    org_nodes = (
        OrgNode.query
        .join(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
        .filter(
            OrgNode.is_active == True,
            OrgNodeType.is_active == True,
            OrgNodeType.allow_in_approvals == True,
        )
        .order_by(OrgNodeType.sort_order.asc().nullslast(), OrgNode.name_ar.asc())
        .all()
    )
    role_choices = _get_role_choices()
    committees = Committee.query.filter_by(is_active=True).order_by(Committee.name_ar.asc()).all()

    return render_template(
        "workflow/templates_admin/edit.html",
        t=t,
        steps=steps,
        users=users,
        departments=departments,
        directorates=directorates,
        units=units,
        sections=sections,
        divisions=divisions,
        org_nodes=org_nodes,
        role_choices=role_choices,
        committees=committees,
    )


@workflow_bp.route("/templates/<int:template_id>/details")
@login_required
@perm_required("WORKFLOW_TEMPLATES_READ")
def templates_details(template_id):
    t = WorkflowTemplate.query.get_or_404(template_id)

    steps = (
        WorkflowTemplateStep.query.filter_by(template_id=t.id)
        .order_by(WorkflowTemplateStep.step_order.asc())
        .all()
    )

    users = User.query.order_by(User.email.asc()).all()
    users_map = {u.id: u for u in users}

    departments = (
        Department.query.filter_by(is_active=True).order_by(Department.name_ar.asc()).all()
    )
    depts_map = {d.id: d for d in departments}

    directorates = (
        Directorate.query.filter_by(is_active=True)
        .order_by(Directorate.name_ar.asc())
        .all()
    )
    dirs_map = {d.id: d for d in directorates}

    units = Unit.query.filter_by(is_active=True).order_by(Unit.name_ar.asc()).all()
    units_map = {u.id: u for u in units}

    sections = Section.query.filter_by(is_active=True).order_by(Section.name_ar.asc()).all()
    sections_map = {s.id: s for s in sections}

    divisions = Division.query.filter_by(is_active=True).order_by(Division.name_ar.asc()).all()
    divisions_map = {d.id: d for d in divisions}

    committees = Committee.query.order_by(Committee.name_ar.asc()).all()
    committees_map = {c.id: c for c in committees}

    org_nodes = (
        OrgNode.query
        .join(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
        .filter(OrgNode.is_active == True, OrgNodeType.is_active == True)
        .all()
    )
    org_nodes_map = {n.id: n for n in org_nodes}

    return render_template(
        "workflow/templates_admin/details.html",
        t=t,
        steps=steps,
        users_map=users_map,
        depts_map=depts_map,
        dirs_map=dirs_map,
        units_map=units_map,
        sections_map=sections_map,
        divisions_map=divisions_map,
        org_nodes_map=org_nodes_map,
        committees_map=committees_map,
    )


@workflow_bp.route("/templates/<int:template_id>/delete", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_DELETE")
def templates_delete(template_id):
    t = WorkflowTemplate.query.get_or_404(template_id)

    WorkflowTemplateStep.query.filter_by(template_id=t.id).delete()
    db.session.delete(t)
    db.session.commit()

    flash("تم حذف المسار.", "warning")
    return redirect(url_for("workflow.templates_list"))




@workflow_bp.route("/templates/<int:template_id>/steps/import-excel", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_steps_import_excel(template_id: int):
    """Import steps for a specific template from Excel.

    Recommended columns:
      - step_order (required)
      - mode (SEQUENTIAL / PARALLEL_SYNC)
      - approver_kind (USER/ROLE/DEPARTMENT/DIRECTORATE/UNIT/SECTION/DIVISION/COMMITTEE)
      - target (email / role code / dept code / dir code / committee code)
      - sla_days (optional)
      - committee_delivery_mode (optional for COMMITTEE)
      - parallel_assignees (optional; for PARALLEL_SYNC)

    Modes:
      - safe: upsert by step_order
      - replace: deletes existing steps for this template then re-inserts
    """
    from utils import importer as xl

    t = WorkflowTemplate.query.get_or_404(template_id)

    mode = (request.form.get("mode") or "safe").strip().lower()
    if mode not in ("replace", "safe"):
        mode = "safe"

    f = request.files.get("file")
    if not f or not getattr(f, "filename", ""):
        flash("اختر ملف Excel (.xlsx)", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    try:
        _sheet, rows, headers = xl.read_excel_rows(f)
    except Exception as e:
        flash(f"خطأ في قراءة ملف Excel: {e}", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    # caches
    users_by_email = { (u.email or '').strip().lower(): u for u in User.query.all() if u.email }
    roles_set = { (r.code or '').strip(): r for r in Role.query.filter_by(is_active=True).all() if r.code }

    depts = Department.query.all()
    dirs = Directorate.query.all()
    depts_by_id = {d.id: d for d in depts}
    dirs_by_id = {d.id: d for d in dirs}

    units = Unit.query.all()
    sections = Section.query.all()
    divisions = Division.query.all()
    units_by_id = {u.id: u for u in units}
    sections_by_id = {s.id: s for s in sections}
    divisions_by_id = {d.id: d for d in divisions}

    def _key(s):
        return (s or '').strip().lower()

    import re

    def _extract_id(text: str):
        if not text:
            return None
        s = str(text)
        # Common exported format: "Name [#123]"
        m = re.search(r'\[#\s*(\d+)\s*\]', s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        m = re.search(r'#\s*(\d+)', s)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None

    def _strip_bracket_id(text: str) -> str:
        if text is None:
            return ''
        s = str(text)
        s = re.sub(r'\s*\[#\s*\d+\s*\]\s*', ' ', s)
        return ' '.join(s.split()).strip()

    depts_by_code = {_key(getattr(d, 'code', None)): d for d in depts if getattr(d, 'code', None)}
    dirs_by_code  = {_key(getattr(d, 'code', None)): d for d in dirs if getattr(d, 'code', None)}

    depts_by_name = {_key(getattr(d, 'name_ar', None)): d for d in depts if getattr(d, 'name_ar', None)}
    depts_by_name.update({_key(getattr(d, 'name_en', None)): d for d in depts if getattr(d, 'name_en', None)})

    dirs_by_name = {_key(getattr(d, 'name_ar', None)): d for d in dirs if getattr(d, 'name_ar', None)}
    dirs_by_name.update({_key(getattr(d, 'name_en', None)): d for d in dirs if getattr(d, 'name_en', None)})

    units_by_code = {_key(getattr(u, 'code', None)): u for u in units if getattr(u, 'code', None)}
    units_by_name = {_key(getattr(u, 'name_ar', None)): u for u in units if getattr(u, 'name_ar', None)}
    units_by_name.update({_key(getattr(u, 'name_en', None)): u for u in units if getattr(u, 'name_en', None)})

    sections_by_code = {_key(getattr(s, 'code', None)): s for s in sections if getattr(s, 'code', None)}
    sections_by_name = {_key(getattr(s, 'name_ar', None)): s for s in sections if getattr(s, 'name_ar', None)}
    sections_by_name.update({_key(getattr(s, 'name_en', None)): s for s in sections if getattr(s, 'name_en', None)})

    divisions_by_code = {_key(getattr(d, 'code', None)): d for d in divisions if getattr(d, 'code', None)}
    divisions_by_name = {_key(getattr(d, 'name_ar', None)): d for d in divisions if getattr(d, 'name_ar', None)}
    divisions_by_name.update({_key(getattr(d, 'name_en', None)): d for d in divisions if getattr(d, 'name_en', None)})

    committees = Committee.query.all()
    committees_by_id = {c.id: c for c in committees}
    committees_by_code = {_key(getattr(c, 'code', None)): c for c in committees if getattr(c, 'code', None)}
    committees_by_name = {_key(getattr(c, 'name_ar', None)): c for c in committees if getattr(c, 'name_ar', None)}
    committees_by_name.update({_key(getattr(c, 'name_en', None)): c for c in committees if getattr(c, 'name_en', None)})

    # Org nodes (dynamic)
    org_nodes = (
        OrgNode.query
        .join(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
        .filter(OrgNode.is_active == True, OrgNodeType.is_active == True)
        .all()
    )
    org_nodes_by_id = {n.id: n for n in org_nodes}
    org_nodes_by_code = {_key(getattr(n, 'code', None)): n for n in org_nodes if getattr(n, 'code', None)}
    org_nodes_by_name = {_key(getattr(n, 'name_ar', None)): n for n in org_nodes if getattr(n, 'name_ar', None)}
    org_nodes_by_name.update({_key(getattr(n, 'name_en', None)): n for n in org_nodes if getattr(n, 'name_en', None)})

    def _resolve_org_node(val):
        s = xl.to_str(val)
        if not s:
            return None
        s = s.strip()
        if s.isdigit():
            return org_nodes_by_id.get(int(s))
        sid = _extract_id(s)
        if sid is not None:
            obj = org_nodes_by_id.get(sid)
            if obj:
                return obj
        clean = _strip_bracket_id(s)
        return org_nodes_by_code.get(_key(clean)) or org_nodes_by_name.get(_key(clean))

    def _normalize_step_mode(v: str) -> str:
        s = (v or '').strip().upper()
        if not s:
            return 'SEQUENTIAL'
        if s in ('PARALLEL', 'PARALLEL_SYNC', 'SYNC', 'مزامن', 'متزامن'):
            return 'PARALLEL_SYNC'
        if s in ('SEQUENTIAL', 'SEQ', 'تسلسلي'):
            return 'SEQUENTIAL'
        return 'PARALLEL_SYNC' if 'PARALLEL' in s or 'SYNC' in s else 'SEQUENTIAL'

    def _resolve_department(val):
        s = xl.to_str(val)
        if not s:
            return None
        s = s.strip()
        if s.isdigit():
            return depts_by_id.get(int(s))
        sid = _extract_id(s)
        if sid is not None:
            obj = depts_by_id.get(sid)
            if obj:
                return obj
        clean = _strip_bracket_id(s)
        return depts_by_code.get(_key(clean)) or depts_by_name.get(_key(clean))

    def _resolve_directorate(val):
        s = xl.to_str(val)
        if not s:
            return None
        s = s.strip()
        if s.isdigit():
            return dirs_by_id.get(int(s))
        sid = _extract_id(s)
        if sid is not None:
            obj = dirs_by_id.get(sid)
            if obj:
                return obj
        clean = _strip_bracket_id(s)
        return dirs_by_code.get(_key(clean)) or dirs_by_name.get(_key(clean))
    def _resolve_unit(target_val):
        if not target_val:
            return None
        s = str(target_val).strip()
        # allow id
        try:
            sid = int(s)
            if sid in units_by_id:
                return units_by_id[sid]
        except Exception:
            pass
        sid2 = _extract_id(s)
        if sid2 is not None and sid2 in units_by_id:
            return units_by_id[sid2]
        s = _strip_bracket_id(s)
        key = _key(s)
        return units_by_code.get(key) or units_by_name.get(key)

    def _resolve_section(target_val):
        if not target_val:
            return None
        s = str(target_val).strip()
        try:
            sid = int(s)
            if sid in sections_by_id:
                return sections_by_id[sid]
        except Exception:
            pass
        sid2 = _extract_id(s)
        if sid2 is not None and sid2 in sections_by_id:
            return sections_by_id[sid2]
        s = _strip_bracket_id(s)
        key = _key(s)
        return sections_by_code.get(key) or sections_by_name.get(key)

    def _resolve_division(target_val):
        if not target_val:
            return None
        s = str(target_val).strip()
        try:
            sid = int(s)
            if sid in divisions_by_id:
                return divisions_by_id[sid]
        except Exception:
            pass
        sid2 = _extract_id(s)
        if sid2 is not None and sid2 in divisions_by_id:
            return divisions_by_id[sid2]
        s = _strip_bracket_id(s)
        key = _key(s)
        return divisions_by_code.get(key) or divisions_by_name.get(key)



    def _resolve_committee(val):
        s = xl.to_str(val)
        if not s:
            return None
        s = s.strip()
        if s.isdigit():
            return committees_by_id.get(int(s))
        sid = _extract_id(s)
        if sid is not None:
            obj = committees_by_id.get(sid)
            if obj:
                return obj
        clean = _strip_bracket_id(s)
        return committees_by_code.get(_key(clean)) or committees_by_name.get(_key(clean))

    if mode == 'replace':
        # delete existing steps for this template
        WorkflowTemplateParallelAssignee.query.filter_by(template_id=t.id).delete(synchronize_session=False)
        WorkflowTemplateStep.query.filter_by(template_id=t.id).delete(synchronize_session=False)
        db.session.flush()

    created = updated = skipped = 0

    for r in rows:
        step_order = xl.to_int(xl.pick(r, 'step_order', 'step', 'رقم الخطوة', 'ترتيب الخطوة', 'الترتيب', 'step order'), default=None)
        if step_order is None:
            skipped += 1
            continue

        step_mode = _normalize_step_mode(xl.to_str(xl.pick(r, 'mode', 'execution_mode', 'طريقة التنفيذ')) or '')
        kind = _normalize_kind(xl.to_str(xl.pick(r, 'approver_kind', 'kind', 'النوع')) or '')
        target = xl.to_str(xl.pick(r, 'target', 'value', 'المستهدف'))
        sla_days = xl.to_int(xl.pick(r, 'sla_days', 'sla', 'SLA', 'SLA (أيام)', 'SLA(أيام)', 'sla (أيام)', 'sla(أيام)', 'sla أيام', 'sla days'), default=None)
        cmode = xl.to_str(xl.pick(r, 'committee_delivery_mode', 'committee_mode'))
        cmode = _normalize_committee_delivery_mode(cmode)

        if not kind:
            skipped += 1
            continue

        # Find existing step
        step = WorkflowTemplateStep.query.filter_by(template_id=t.id, step_order=step_order).first()
        if step and mode == 'safe':
            updated += 1
        elif step and mode == 'replace':
            updated += 1
        else:
            step = WorkflowTemplateStep(template_id=t.id, step_order=step_order, approver_kind=kind)
            db.session.add(step)
            created += 1

        # reset targets
        step.mode = step_mode
        step.approver_kind = kind
        step.approver_user_id = None
        step.approver_department_id = None
        step.approver_directorate_id = None
        step.approver_unit_id = None
        step.approver_section_id = None
        step.approver_division_id = None
        step.approver_org_node_id = None
        step.approver_role = None
        step.approver_committee_id = None
        step.committee_delivery_mode = None
        step.sla_days = sla_days

        if kind == 'USER':
            u = users_by_email.get((target or '').strip().lower()) if target else None
            if not u:
                skipped += 1
                continue
            step.approver_user_id = u.id

        elif kind == 'ROLE':
            if not target:
                skipped += 1
                continue
            role_code = target.strip()
            # allow role even if not seeded, but prefer existing
            step.approver_role = role_code

        elif kind == 'DIRECTORATE':
            d = _resolve_directorate(target)
            if not d:
                skipped += 1
                continue
            step.approver_directorate_id = d.id

        elif kind == 'DEPARTMENT':
            d = _resolve_department(target)
            if not d:
                skipped += 1
                continue
            step.approver_department_id = d.id

        elif kind == 'UNIT':
            uo = _resolve_unit(target)
            if not uo:
                skipped += 1
                continue
            step.approver_unit_id = uo.id

        elif kind == 'SECTION':
            so = _resolve_section(target)
            if not so:
                skipped += 1
                continue
            step.approver_section_id = so.id

        elif kind == 'DIVISION':
            vo = _resolve_division(target)
            if not vo:
                skipped += 1
                continue
            step.approver_division_id = vo.id

        elif kind == 'ORG_NODE':
            no = _resolve_org_node(target)
            if not no:
                skipped += 1
                continue
            step.approver_org_node_id = no.id

        elif kind == 'COMMITTEE':
            c = _resolve_committee(target)
            if not c:
                skipped += 1
                continue
            step.approver_committee_id = c.id
            step.committee_delivery_mode = cmode

        else:
            # Unknown kinds fall back to role
            step.approver_role = (target or '').strip() or None

        # Optional: PARALLEL_SYNC extra assignees
        extras = xl.to_str(xl.pick(r, "parallel_assignees", "extras", "المتزامنون"))
        if step_mode == "PARALLEL_SYNC" and extras:
            # Ensure step has ID
            db.session.flush()

            # Clear old extras for this step (safe updates)
            WorkflowTemplateParallelAssignee.query.filter_by(
                template_id=t.id, step_order=step_order
            ).delete(synchronize_session=False)

            norm = extras.replace("\n", ";").replace(",", ";")
            items = [x.strip() for x in norm.split(";") if x.strip()]
            for item in items:
                # format: KIND:VALUE (defaults to USER)
                if ":" in item:
                    ek, ev = item.split(":", 1)
                    ek = _normalize_kind(ek)
                    ev = (ev or "").strip()
                else:
                    ek, ev = "USER", item.strip()

                a = WorkflowTemplateParallelAssignee(
                    template_step_id=step.id,
                    template_id=t.id,
                    step_order=step_order,
                    approver_kind=ek,
                )

                if ek == "USER":
                    uu = users_by_email.get(ev.strip().lower())
                    if not uu:
                        continue
                    a.approver_user_id = uu.id
                elif ek == "ROLE":
                    if not ev:
                        continue
                    a.approver_role = ev
                elif ek == "DIRECTORATE":
                    dd = _resolve_directorate(ev)
                    if not dd:
                        continue
                    a.approver_directorate_id = dd.id
                elif ek == "DEPARTMENT":
                    dd = _resolve_department(ev)
                    if not dd:
                        continue
                    a.approver_department_id = dd.id
                elif ek == "COMMITTEE":
                    cc = _resolve_committee(ev)
                    if not cc:
                        continue
                    a.approver_committee_id = cc.id
                    a.committee_delivery_mode = "Committee_ALL"
                else:
                    continue

                db.session.add(a)

    # flush to link parallel assignees to steps
    db.session.flush()
    # update template_step_id for newly added parallel assignees
    if rows:
        # map step_order -> step.id
        step_map = {s.step_order: s.id for s in WorkflowTemplateStep.query.filter_by(template_id=t.id).all()}
        for a in WorkflowTemplateParallelAssignee.query.filter_by(template_id=t.id).all():
            if a.template_step_id:
                continue
            a.template_step_id = step_map.get(a.step_order) or a.template_step_id

    try:
        db.session.commit()
        flash(f"✅ تم استيراد الخطوات. جديد: {created} / تحديث: {updated} / تخطي: {skipped}", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"❌ فشل استيراد الخطوات: {e}", "danger")

    return redirect(url_for('workflow.templates_edit', template_id=t.id))
@workflow_bp.route("/templates/<int:template_id>/steps/add", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_steps_add(template_id):
    t = WorkflowTemplate.query.get_or_404(template_id)

    kind = _normalize_kind(request.form.get("approver_kind"))
    user_id = _to_int(request.form.get("approver_user_id"), default=None)
    dept_id = _to_int(request.form.get("approver_department_id"), default=None)
    dir_id = _to_int(request.form.get("approver_directorate_id"), default=None)
    role = (request.form.get("approver_role") or "").strip()
    unit_id = _to_int(request.form.get("approver_unit_id"), default=None)
    section_id = _to_int(request.form.get("approver_section_id"), default=None)
    division_id = _to_int(request.form.get("approver_division_id"), default=None)
    org_node_id = _to_int(request.form.get("approver_org_node_id"), default=None)

    committee_id = _to_int(request.form.get("approver_committee_id"), default=None)
    committee_delivery_mode = _normalize_committee_delivery_mode(request.form.get("committee_delivery_mode"))

    sla_days = _to_int(request.form.get("sla_days"), default=None)
    mode = _normalize_step_mode(request.form.get("mode"))

    if not kind:
        flash("يرجى اختيار نوع المعتمد (USER/ROLE/DEPARTMENT/DIRECTORATE/UNIT/SECTION/DIVISION/ORG_NODE/COMMITTEE).", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "USER" and not user_id:
        flash("يرجى اختيار مستخدم للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "ROLE" and not role:
        flash("يرجى إدخال ROLE للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "DEPARTMENT" and not dept_id:
        flash("يرجى اختيار دائرة/قسم للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "DIRECTORATE" and not dir_id:
        flash("يرجى اختيار إدارة للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "UNIT" and not unit_id:
        flash("يرجى اختيار وحدة للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "SECTION" and not section_id:
        flash("يرجى اختيار قسم (Section) للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "DIVISION" and not division_id:
        flash("يرجى اختيار شعبة (Division) للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "ORG_NODE" and not org_node_id:
        flash("يرجى اختيار عنصر من الهيكلية الموحدة (Org Node) للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    if kind == "COMMITTEE" and not committee_id:
        flash("يرجى اختيار لجنة للخطوة.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    # Validation: CHAIR/SECRETARY must exist (only for committee steps)
    if kind == "COMMITTEE" and committee_delivery_mode in ("Committee_CHAIR", "Committee_SECRETARY"):
        needed = 'CHAIR' if committee_delivery_mode == "Committee_CHAIR" else 'SECRETARY'
        exists = CommitteeAssignee.query.filter_by(committee_id=committee_id, is_active=True)             .filter(CommitteeAssignee.member_role.ilike(needed)).first()
        if not exists:
            flash("اللجنة المختارة لا تحتوي على %s نشط." % ('رئيس' if needed=='CHAIR' else 'مقرر'), "danger")
            return redirect(url_for("workflow.templates_edit", template_id=t.id))

    max_order = (
        db.session.query(func.max(WorkflowTemplateStep.step_order))
        .filter(WorkflowTemplateStep.template_id == t.id)
        .scalar()
    ) or 0

    step = WorkflowTemplateStep(
        template_id=t.id,
        step_order=int(max_order) + 1,
        approver_kind=kind,
        approver_user_id=user_id if kind == "USER" else None,
        approver_department_id=dept_id if kind == "DEPARTMENT" else None,
        approver_directorate_id=dir_id if kind == "DIRECTORATE" else None,
        approver_unit_id=unit_id if kind == "UNIT" else None,
        approver_section_id=section_id if kind == "SECTION" else None,
        approver_division_id=division_id if kind == "DIVISION" else None,
        approver_org_node_id=org_node_id if kind == "ORG_NODE" else None,
        approver_role=role if kind == "ROLE" else None,
        approver_committee_id=committee_id if kind == "COMMITTEE" else None,
        committee_delivery_mode=committee_delivery_mode if kind == "COMMITTEE" else None,
        sla_days=sla_days,
        mode=mode,
    )
    db.session.add(step)
    db.session.commit()

    flash("تمت إضافة خطوة.", "success")
    return redirect(url_for("workflow.templates_edit", template_id=t.id))


@workflow_bp.route("/templates/steps/<int:step_id>/move/<direction>", methods=["POST"])


@workflow_bp.route("/templates/steps/<int:step_id>/move/<direction>", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_steps_move(step_id, direction):
    s = WorkflowTemplateStep.query.get_or_404(step_id)
    tid = s.template_id

    direction = (direction or "").strip().lower()
    if direction not in ("up", "down"):
        flash("اتجاه غير صحيح.", "danger")
        return redirect(url_for("workflow.templates_edit", template_id=tid))

    target_order = (s.step_order or 1) - 1 if direction == "up" else (s.step_order or 1) + 1

    other = (
        WorkflowTemplateStep.query
        .filter_by(template_id=tid, step_order=target_order)
        .first()
    )
    if other:
        other.step_order, s.step_order = s.step_order, other.step_order
        db.session.commit()
        _resequence_steps(tid)

    flash("تم تحديث ترتيب الخطوات.", "success")
    return redirect(url_for("workflow.templates_edit", template_id=tid))


@workflow_bp.route("/templates/steps/<int:step_id>/delete", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_steps_delete(step_id):
    s = WorkflowTemplateStep.query.get_or_404(step_id)
    tid = s.template_id

    db.session.delete(s)
    db.session.commit()

    _resequence_steps(tid)
    flash("تم حذف الخطوة.", "warning")
    return redirect(url_for("workflow.templates_edit", template_id=tid))


@workflow_bp.route("/templates/steps/<int:step_id>/edit", methods=["GET", "POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_UPDATE")
def templates_steps_edit(step_id):
    s = WorkflowTemplateStep.query.get_or_404(step_id)
    t = WorkflowTemplate.query.get_or_404(s.template_id)

    users = User.query.order_by(User.email.asc()).all()
    departments = Department.query.filter_by(is_active=True).order_by(Department.name_ar.asc()).all()
    directorates = Directorate.query.filter_by(is_active=True).order_by(Directorate.name_ar.asc()).all()
    units = Unit.query.filter_by(is_active=True).order_by(Unit.name_ar.asc()).all()
    sections = Section.query.filter_by(is_active=True).order_by(Section.name_ar.asc()).all()
    divisions = Division.query.filter_by(is_active=True).order_by(Division.name_ar.asc()).all()
    role_choices = _get_role_choices()
    committees = Committee.query.filter_by(is_active=True).order_by(Committee.name_ar.asc()).all()

    org_nodes = (
        OrgNode.query
        .join(OrgNodeType, OrgNode.type_id == OrgNodeType.id)
        .filter(
            OrgNode.is_active == True,
            OrgNodeType.is_active == True,
            OrgNodeType.allow_in_approvals == True,
        )
        .order_by(OrgNodeType.sort_order.asc().nullslast(), OrgNode.name_ar.asc())
        .all()
    )

    if request.method == "POST":
        kind = _normalize_kind(request.form.get("approver_kind"))
        user_id = _to_int(request.form.get("approver_user_id"), default=None)
        dept_id = _to_int(request.form.get("approver_department_id"), default=None)
        dir_id = _to_int(request.form.get("approver_directorate_id"), default=None)
        role = (request.form.get("approver_role") or "").strip()
        unit_id = _to_int(request.form.get("approver_unit_id"), default=None)
        section_id = _to_int(request.form.get("approver_section_id"), default=None)
        division_id = _to_int(request.form.get("approver_division_id"), default=None)
        org_node_id = _to_int(request.form.get("approver_org_node_id"), default=None)

        committee_id = _to_int(request.form.get("approver_committee_id"), default=None)
        committee_delivery_mode = _normalize_committee_delivery_mode(request.form.get("committee_delivery_mode"))

        sla_days = _to_int(request.form.get("sla_days"), default=None)
        mode = _normalize_step_mode(request.form.get("mode"))

        if not kind:
            flash("نوع المعتمد غير صحيح.", "danger")
            return redirect(request.url)

        if kind == "USER" and not user_id:
            flash("يرجى اختيار مستخدم.", "danger")
            return redirect(request.url)

        if kind == "ROLE" and not role:
            flash("يرجى إدخال ROLE.", "danger")
            return redirect(request.url)

        if kind == "DEPARTMENT" and not dept_id:
            flash("يرجى اختيار دائرة/قسم للخطوة.", "danger")
            return redirect(request.url)

        if kind == "DIRECTORATE" and not dir_id:
            flash("يرجى اختيار إدارة للخطوة.", "danger")
            return redirect(request.url)

        if kind == "UNIT" and not unit_id:
            flash("يرجى اختيار وحدة للخطوة.", "danger")
            return redirect(request.url)

        if kind == "SECTION" and not section_id:
            flash("يرجى اختيار قسم (Section) للخطوة.", "danger")
            return redirect(request.url)

        if kind == "DIVISION" and not division_id:
            flash("يرجى اختيار شعبة (Division) للخطوة.", "danger")
            return redirect(request.url)

        if kind == "ORG_NODE" and not org_node_id:
            flash("يرجى اختيار عنصر من الهيكلية الموحدة (Org Node) للخطوة.", "danger")
            return redirect(request.url)

        if kind == "COMMITTEE" and not committee_id:
            flash("يرجى اختيار لجنة للخطوة.", "danger")
            return redirect(request.url)

        if kind == "COMMITTEE" and committee_delivery_mode in ("Committee_CHAIR", "Committee_SECRETARY"):
            needed = 'CHAIR' if committee_delivery_mode == "Committee_CHAIR" else 'SECRETARY'
            exists = CommitteeAssignee.query.filter_by(committee_id=committee_id, is_active=True)                 .filter(CommitteeAssignee.member_role.ilike(needed)).first()
            if not exists:
                flash("اللجنة المختارة لا تحتوي على %s نشط." % ('رئيس' if needed=='CHAIR' else 'مقرر'), "danger")
                return redirect(request.url)

        s.approver_kind = kind
        s.approver_user_id = user_id if kind == "USER" else None
        s.approver_role = role if kind == "ROLE" else None
        s.approver_department_id = dept_id if kind == "DEPARTMENT" else None
        s.approver_directorate_id = dir_id if kind == "DIRECTORATE" else None
        s.approver_unit_id = unit_id if kind == "UNIT" else None
        s.approver_section_id = section_id if kind == "SECTION" else None
        s.approver_division_id = division_id if kind == "DIVISION" else None
        s.approver_org_node_id = org_node_id if kind == "ORG_NODE" else None
        s.approver_committee_id = committee_id if kind == "COMMITTEE" else None
        s.committee_delivery_mode = committee_delivery_mode if kind == "COMMITTEE" else None
        s.sla_days = sla_days
        s.mode = mode

        # If mode is not PARALLEL_SYNC, remove any linked parallel assignees
        if mode != "PARALLEL_SYNC":
            try:
                WorkflowTemplateParallelAssignee.query.filter_by(template_step_id=s.id).delete()
            except Exception:
                pass

        db.session.commit()
        flash("تم تحديث الخطوة.", "success")
        return redirect(url_for("workflow.templates_edit", template_id=t.id))

    return render_template(
        "workflow/templates_admin/step_edit.html",
        t=t,
        s=s,
        users=users,
        departments=departments,
        directorates=directorates,
        units=units,
        sections=sections,
        divisions=divisions,
        org_nodes=org_nodes,
        role_choices=role_choices,
        committees=committees,
    )


@workflow_bp.route("/templates/steps/<int:step_id>/parallel-assignees/add", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_WRITE")
def template_step_parallel_assignee_add(step_id: int):
    s = WorkflowTemplateStep.query.get_or_404(step_id)
    t = WorkflowTemplate.query.get_or_404(s.template_id)

    if (getattr(s, 'mode', '') or '').upper() != 'PARALLEL_SYNC':
        flash("يمكن إضافة مرتبطين فقط للخطوات المتزامنة.", "warning")
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    kind = _normalize_kind(request.form.get('pa_kind'))
    user_id = _to_int(request.form.get('pa_user_id'), default=None)
    dept_id = _to_int(request.form.get('pa_department_id'), default=None)
    dir_id = _to_int(request.form.get('pa_directorate_id'), default=None)
    role = (request.form.get('pa_role') or '').strip()
    unit_id = _to_int(request.form.get('pa_unit_id'), default=None)
    section_id = _to_int(request.form.get('pa_section_id'), default=None)
    division_id = _to_int(request.form.get('pa_division_id'), default=None)
    org_node_id = _to_int(request.form.get('pa_org_node_id'), default=None)

    committee_id = _to_int(request.form.get('pa_committee_id'), default=None)
    committee_delivery_mode = _normalize_committee_delivery_mode(request.form.get('pa_committee_delivery_mode'))

    if not kind:
        flash('نوع المرتبط غير صحيح.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'USER' and not user_id:
        flash('يرجى اختيار مستخدم.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'ROLE' and not role:
        flash('يرجى اختيار Role.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'DEPARTMENT' and not dept_id:
        flash('يرجى اختيار Department.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'DIRECTORATE' and not dir_id:
        flash('يرجى اختيار Directorate.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'UNIT' and not unit_id:
        flash('يرجى اختيار Unit.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'SECTION' and not section_id:
        flash('يرجى اختيار Section.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'DIVISION' and not division_id:
        flash('يرجى اختيار Division.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'ORG_NODE' and not org_node_id:
        flash('يرجى اختيار عنصر هيكلية موحّد.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'COMMITTEE' and not committee_id:
        flash('يرجى اختيار لجنة.', 'danger')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    if kind == 'COMMITTEE' and committee_delivery_mode in ("Committee_CHAIR", "Committee_SECRETARY"):
        needed = 'CHAIR' if committee_delivery_mode == "Committee_CHAIR" else 'SECRETARY'
        exists = CommitteeAssignee.query.filter_by(committee_id=committee_id, is_active=True)             .filter(CommitteeAssignee.member_role.ilike(needed)).first()
        if not exists:
            flash("اللجنة المختارة لا تحتوي على %s نشط." % ('رئيس' if needed=='CHAIR' else 'مقرر'), "danger")
            return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    # Prevent duplicates
    q = WorkflowTemplateParallelAssignee.query.filter_by(
        template_step_id=s.id,
        approver_kind=kind,
        approver_user_id=(user_id if kind == 'USER' else None),
        approver_role=(role if kind == 'ROLE' else None),
        approver_department_id=(dept_id if kind == 'DEPARTMENT' else None),
        approver_directorate_id=(dir_id if kind == 'DIRECTORATE' else None),
        approver_unit_id=(unit_id if kind == 'UNIT' else None),
        approver_section_id=(section_id if kind == 'SECTION' else None),
        approver_division_id=(division_id if kind == 'DIVISION' else None),
        approver_org_node_id=(org_node_id if kind == 'ORG_NODE' else None),
        approver_committee_id=(committee_id if kind == 'COMMITTEE' else None),
        committee_delivery_mode=(committee_delivery_mode if kind == 'COMMITTEE' else None),
    )
    if q.first():
        flash('هذا المرتبط موجود بالفعل.', 'info')
        return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))

    row = WorkflowTemplateParallelAssignee(
        template_step_id=s.id,
        template_id=s.template_id,
        step_order=s.step_order,
        approver_kind=kind,
        approver_user_id=(user_id if kind == 'USER' else None),
        approver_role=(role if kind == 'ROLE' else None),
        approver_department_id=(dept_id if kind == 'DEPARTMENT' else None),
        approver_directorate_id=(dir_id if kind == 'DIRECTORATE' else None),
        approver_unit_id=(unit_id if kind == 'UNIT' else None),
        approver_section_id=(section_id if kind == 'SECTION' else None),
        approver_division_id=(division_id if kind == 'DIVISION' else None),
        approver_org_node_id=(org_node_id if kind == 'ORG_NODE' else None),
        approver_committee_id=(committee_id if kind == 'COMMITTEE' else None),
        committee_delivery_mode=(committee_delivery_mode if kind == 'COMMITTEE' else None),
    )

    db.session.add(row)
    db.session.commit()

    flash('تم ربط المرتبط بالخطوة المتزامنة.', 'success')
    return redirect(url_for('workflow.templates_steps_edit', step_id=s.id))


@workflow_bp.route("/templates/parallel-assignees/<int:pa_id>/delete", methods=["POST"])
@login_required
@perm_required("WORKFLOW_TEMPLATES_WRITE")
def template_step_parallel_assignee_delete(pa_id: int):
    row = WorkflowTemplateParallelAssignee.query.get_or_404(pa_id)
    step_id = int(getattr(row, 'template_step_id'))
    db.session.delete(row)
    db.session.commit()
    flash('تم حذف المرتبط.', 'success')
    return redirect(url_for('workflow.templates_steps_edit', step_id=step_id))


# =========================
# OrgNode picker tree (for approvals)
# =========================

@workflow_bp.route("/pickers/org-node-tree")
@login_required
@perm_required("WORKFLOW_TEMPLATES_READ")
def org_node_picker_tree():
    """Return OrgNode tree HTML snippet for picker modals (AJAX)."""
    mode = (request.args.get("mode") or "approvals").strip().lower()
    sel = (request.args.get("selected") or "").strip()
    selected_id = int(sel) if sel.isdigit() else None

    tree = build_org_node_picker_tree(mode=mode)
    return render_template(
        "components/_org_node_picker_tree.html",
        tree=tree,
        selected_id=selected_id,
        mode=mode,
    )
