"""Excel import helper used across admin pages.

This helper is intentionally lightweight, but consistent across screens.

Features
--------
- Flexible header normalization (Arabic/English, spaces/underscores)
- Read rows from .xlsx into dicts
- Validate required columns
- Generic upsert-by-code
- Generic replace-all (clear then insert)

Each page can still add its own validation and mapping.
"""

from __future__ import annotations

from typing import Any, Callable, Dict, Iterable, List, Optional, Sequence, Tuple


def normalize_header(h: str) -> str:
    """Normalize a header for tolerant matching."""
    if h is None:
        return ""
    s = str(h).strip().lower()
    if not s:
        return ""
    # unify separators
    for ch in ["\u200f", "\u200e", "\ufeff"]:
        s = s.replace(ch, "")
    s = s.replace("-", " ").replace("/", " ")
    s = s.replace("_", " ")
    # collapse spaces
    s = " ".join(s.split())
    return s


def norm_key(key: str) -> str:
    return normalize_header(key).replace(" ", "")


def _is_empty_row(values: Sequence[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def read_excel_rows(file_storage, *, sheet_index: int = 0, max_rows: int = 10000) -> Tuple[str, List[Dict[str, Any]], List[str]]:
    """Read an Excel file from Werkzeug FileStorage.

    Returns:
      (sheet_title, rows, headers_normalized)

    `rows` is a list of dicts where keys are normalized header keys without spaces.
    Example: "RequestType Code" -> "requesttypecode".
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_storage, data_only=True)
    sheetnames = wb.sheetnames
    if not sheetnames:
        raise ValueError("Excel file has no sheets")

    idx = sheet_index if 0 <= sheet_index < len(sheetnames) else 0
    ws = wb[sheetnames[idx]]

    # Read header row
    raw_headers = []
    for cell in ws[1]:
        raw_headers.append(str(cell.value).strip() if cell.value is not None else "")

    headers_norm = [norm_key(h) for h in raw_headers]

    rows: List[Dict[str, Any]] = []
    count = 0
    for r in ws.iter_rows(min_row=2, values_only=True):
        if _is_empty_row(r):
            continue
        row: Dict[str, Any] = {}
        for hn, val in zip(headers_norm, r):
            if hn:
                row[hn] = val
        rows.append(row)
        count += 1
        if count >= max_rows:
            break

    return ws.title, rows, headers_norm


def validate_headers(headers_norm: Sequence[str], required_norm: Sequence[str]) -> None:
    missing = [h for h in required_norm if norm_key(h) not in set(headers_norm)]
    if missing:
        raise ValueError("Missing required columns: " + ", ".join(missing))


def pick(row: Dict[str, Any], *possible_headers: str, default: Any = None) -> Any:
    """Pick a value from a normalized-row dict by trying multiple header names."""
    for h in possible_headers:
        k = norm_key(h)
        if k in row:
            return row.get(k)
    return default


def to_str(val: Any) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


def to_int(val: Any, default: Optional[int] = None) -> Optional[int]:
    if val is None:
        return default
    try:
        s = str(val).strip()
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def to_bool(val: Any, default: Optional[bool] = None) -> Optional[bool]:
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "y", "on", "نعم"):
        return True
    if s in ("0", "false", "no", "n", "off", "لا"):
        return False
    return default


def upsert_by_code(
    session,
    model,
    rows: Iterable[Dict[str, Any]],
    *,
    code_getter: Callable[[Dict[str, Any]], Optional[str]],
    values_getter: Callable[[Dict[str, Any]], Dict[str, Any]],
    code_field: str = "code",
    normalize_code: Callable[[str], str] = lambda s: s.strip(),
) -> Tuple[int, int]:
    """Generic upsert by code.

    - `code_getter` returns code string from the row
    - `values_getter` returns dict of fields to set

    Returns: (created, updated)
    """
    created = updated = 0

    for r in rows:
        code = code_getter(r)
        code = to_str(code)
        if not code:
            continue
        code = normalize_code(code)
        if not code:
            continue

        obj = session.query(model).filter(getattr(model, code_field) == code).first()
        values = values_getter(r) or {}

        if obj:
            for k, v in values.items():
                setattr(obj, k, v)
            updated += 1
        else:
            obj = model(**{code_field: code, **values})
            session.add(obj)
            created += 1

    return created, updated


def replace_all(
    session,
    delete_query,
    insert_fn: Callable[[], Tuple[int, int]],
    *,
    soft_fallback: Optional[Callable[[], Tuple[int, int]]] = None,
) -> Tuple[int, int, bool]:
    """Replace mode helper.

    Tries to delete using `delete_query.delete()` then runs `insert_fn()`.

    If deletion fails (FK constraints etc.), it will rollback and (if provided)
    run `soft_fallback()` instead.

    Returns:
      (created, updated, used_soft_fallback)
    """
    used_soft = False
    try:
        delete_query.delete(synchronize_session=False)
        session.flush()
        created, updated = insert_fn()
        return created, updated, used_soft
    except Exception:
        session.rollback()
        if not soft_fallback:
            raise
        used_soft = True
        created, updated = soft_fallback()
        return created, updated, used_soft
