"""utils/excel.py

Lightweight helpers to export tables as .xlsx using openpyxl.

Design goals:
- Minimal styling (header bold + freeze pane)
- Safe for Arabic/Unicode
- No database dependencies
"""

from __future__ import annotations

from io import BytesIO
from typing import Iterable, List, Sequence


def _require_openpyxl():
    try:
        import openpyxl  # noqa: F401
        from openpyxl import Workbook  # noqa: F401
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required for Excel export. Please install openpyxl."
        ) from e

    # return commonly used symbols
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    return Workbook, Font, Alignment, get_column_letter


def make_xlsx_bytes(
    sheet_name: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
) -> bytes:
    """Create an .xlsx file (bytes) for a single sheet table."""
    Workbook, Font, Alignment, get_column_letter = _require_openpyxl()

    wb = Workbook()
    ws = wb.active

    # Excel sheet name max 31
    sname = (sheet_name or "Sheet1")[:31]
    ws.title = sname

    # Header
    ws.append(list(headers))
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Rows
    for r in rows:
        ws.append(["" if v is None else v for v in r])

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto width (simple heuristic)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            try:
                val = "" if cell.value is None else str(cell.value)
            except Exception:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def make_xlsx_bytes_multi(
    tables: Sequence[tuple[str, Sequence[str], Iterable[Sequence[object]]]],
) -> bytes:
    """Create an .xlsx file (bytes) containing multiple sheets.

    Each element in *tables* is:
      (sheet_name, headers, rows)
    """
    Workbook, Font, Alignment, get_column_letter = _require_openpyxl()

    wb = Workbook()

    # Remove default sheet if we will add our own
    if tables:
        try:
            wb.remove(wb.active)
        except Exception:
            pass

    header_font = Font(bold=True)

    for idx, (sheet_name, headers, rows) in enumerate(tables):
        ws = wb.create_sheet(title=(sheet_name or f"Sheet{idx+1}")[:31])
        ws.append(list(headers))

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for r in rows:
            ws.append(["" if v is None else v for v in r])

        ws.freeze_panes = "A2"

        # Auto width
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    val = "" if cell.value is None else str(cell.value)
                except Exception:
                    val = ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), 55)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()
